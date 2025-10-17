from django.shortcuts import HttpResponse
from django.http import JsonResponse
from rest_framework.decorators import api_view
from .utils import sapnwrfc
from ctypes import *
from rest_framework.response import Response
from rest_framework import status
from hdbcli import dbapi
import sqlite3
from django.db import connections, transaction
from .serlializers import *
from .models import Project,Connection,fields
import json
from django.core.serializers import serialize
import pandas as pd
import re,string
from django.db import connection
from rest_framework.views import APIView
from .serlializers import FileSerializer
from .models import FileConnection
from django.utils import timezone
from django.db.models import Q, Case, When, Value, IntegerField
import LLM_migration
from django.forms.models import model_to_dict
import sys
from datetime import datetime
import numpy as np
# setting path
sys.path.append('../DMtool')
from DMtool.dmtool import DMTool




def get_valid_data_frame(table_name, connection=connection):
    # Load table from DB to DataFrame
    df = pd.read_sql_query(f"SELECT * FROM {table_name}", connection)

    # Replace empty strings and only-whitespace with NaN
    df = df.replace(r'^\s*$', np.nan, regex=True)
    
    # Drop columns where all values are missing (NaN or converted-empty string)
    df_valid = df.dropna(axis=1, how='all')

    return df_valid

@api_view(['GET'])
def return_target_table(request):
    df_valid = get_valid_data_frame("t_113_Product_Basic_Data_mandatory")
    print(df_valid)
    json_data = df_valid.to_dict(orient="records")
    return JsonResponse(json_data,safe=False)


def home(request):
  return HttpResponse("home Page")


def add_prompt_and_last_updated_on(table_name):
    with connection.cursor() as cursor:
        # Add prompt column
        cursor.execute(f"ALTER TABLE {table_name} ADD COLUMN prompt TEXT;")
        # Add last_updated_on column (as TEXT, or replace with DATETIME if you prefer)
        cursor.execute(f"ALTER TABLE {table_name} ADD COLUMN last_updated_on DATETIME;")
    print(f"Columns 'prompt' and 'last_updated_on' added to {table_name}.")


def copy_rows_with_reason_and_timestamp(table1, table2, row_indices, prompt):
    import pandas as pd

    # Load table1 into DataFrame using raw connection for full compatibility
    conn = connection.connection
    df1 = pd.read_sql_query(f"SELECT * FROM {table1}", conn)
    
    if not row_indices:
        print("No row indices provided to copy.")
        return
    
    # Get the rows by index
    try:
        selected_df = df1.iloc[row_indices].copy()
    except Exception as e:
        print(f"Error selecting rows: {e}")
        return

    # Add the extra columns
    selected_df['prompt'] = prompt
    selected_df['last_updated_on'] = datetime.now()  # or .isoformat()
    print(selected_df.head(5))
    # Write to table2 (assume table2 already exists with correct schema)
    selected_df.to_sql(table2, conn, if_exists='append', index=False)
    print(f"Inserted {len(selected_df)} rows into {table2} with reason and last_updated_on.")

 
@api_view(['POST'])
def SAPconn(request):
    class RFC_ERROR_INFO(Structure):
        _fields_ = [("code", c_long),
                    ("group", c_long),
                    ("key", c_wchar * 128),
                    ("message", c_wchar * 512),
                    ("abapMsgClass", c_wchar * 21),
                    ("abapMsgType", c_wchar * 2),
                    ("abapMsgNumber", c_wchar * 4),
                    ("abapMsgV1", c_wchar * 51),
                    ("abapMsgV2", c_wchar * 51),
                    ("abapMsgV3", c_wchar * 51),
                    ("abapMsgV4", c_wchar * 51)]
    class RFC_CONNECTION_PARAMETER(Structure):
        _fields_ = [("name", c_wchar_p),
                    ("value", c_wchar_p)]
    RFC_OK = 0
    RFC_COMMUNICATION_FAILURE = 1
    RFC_LOGON_FAILURE = 2
    RFC_ABAP_RUNTIME_FAILURE = 3
    RFC_ABAP_MESSAGE = 4
    RFC_ABAP_EXCEPTION = 5
    RFC_CLOSED = 6
    RFC_CANCELED = 7
    RFC_TIMEOUT = 8
    RFC_MEMORY_INSUFFICIENT = 9
    RFC_VERSION_MISMATCH = 10
    RFC_INVALID_PROTOCOL = 11
    RFC_SERIALIZATION_FAILURE = 12
    RFC_INVALID_HANDLE = 13
    RFC_RETRY = 14
    RFC_EXTERNAL_FAILURE = 15
    RFC_EXECUTED = 16
    RFC_NOT_FOUND = 17
    RFC_NOT_SUPPORTED = 18
    RFC_ILLEGAL_STATE = 19
    RFC_INVALID_PARAMETER = 20
    RFC_CODEPAGE_CONVERSION_FAILURE = 21
    RFC_CONVERSION_FAILURE = 22
    RFC_BUFFER_TOO_SMALL = 23
    RFC_TABLE_MOVE_BOF = 24
    RFC_TABLE_MOVE_EOF = 25
    RFC_START_SAPGUI_FAILURE = 26
    RFC_ABAP_CLASS_EXCEPTION = 27
    RFC_UNKNOWN_ERROR = 28
    RFC_AUTHORIZATION_FAILURE = 29

    #-RFCTYPE - RFC data types----------------------------------------------
    RFCTYPE_CHAR = 0
    RFCTYPE_DATE = 1
    RFCTYPE_BCD = 2
    RFCTYPE_TIME = 3
    RFCTYPE_BYTE = 4
    RFCTYPE_TABLE = 5
    RFCTYPE_NUM = 6
    RFCTYPE_FLOAT = 7
    RFCTYPE_INT = 8
    RFCTYPE_INT2 = 9
    RFCTYPE_INT1 = 10
    RFCTYPE_NULL = 14
    RFCTYPE_ABAPOBJECT = 16
    RFCTYPE_STRUCTURE = 17
    RFCTYPE_DECF16 = 23
    RFCTYPE_DECF34 = 24
    RFCTYPE_XMLDATA = 28
    RFCTYPE_STRING = 29
    RFCTYPE_XSTRING = 30
    RFCTYPE_BOX = 31
    RFCTYPE_GENERIC_BOX = 32

    #-RFC_UNIT_STATE - Processing status of a background unit---------------
    RFC_UNIT_NOT_FOUND = 0 
    RFC_UNIT_IN_PROCESS = 1 
    RFC_UNIT_COMMITTED = 2 
    RFC_UNIT_ROLLED_BACK = 3 
    RFC_UNIT_CONFIRMED = 4 

    #-RFC_CALL_TYPE - Type of an incoming function call---------------------
    RFC_SYNCHRONOUS = 0 
    RFC_TRANSACTIONAL = 1 
    RFC_QUEUED = 2 
    RFC_BACKGROUND_UNIT = 3 

    #-RFC_DIRECTION - Direction of a function module parameter--------------
    RFC_IMPORT = 1 
    RFC_EXPORT = 2 
    RFC_CHANGING = RFC_IMPORT + RFC_EXPORT 
    RFC_TABLES = 4 + RFC_CHANGING 

    #-RFC_CLASS_ATTRIBUTE_TYPE - Type of an ABAP object attribute-----------
    RFC_CLASS_ATTRIBUTE_INSTANCE = 0 
    RFC_CLASS_ATTRIBUTE_CLASS = 1 
    RFC_CLASS_ATTRIBUTE_CONSTANT = 2 

    #-RFC_METADATA_OBJ_TYPE - Ingroup repository----------------------------
    RFC_METADATA_FUNCTION = 0 
    RFC_METADATA_TYPE = 1 
    RFC_METADATA_CLASS = 2 


    #-Variables-------------------------------------------------------------
    ErrInf = RFC_ERROR_INFO; RfcErrInf = ErrInf()
    ConnParams = RFC_CONNECTION_PARAMETER * 5; RfcConnParams = ConnParams()
    SConParams = RFC_CONNECTION_PARAMETER * 3; RfcSConParams = SConParams()


    SAPNWRFC = "sapnwrfc.dll"
    SAP = windll.LoadLibrary(SAPNWRFC)

    #-Prototypes------------------------------------------------------------
    SAP.RfcAppendNewRow.argtypes = [c_void_p, POINTER(ErrInf)]
    SAP.RfcAppendNewRow.restype = c_void_p

    SAP.RfcCloseConnection.argtypes = [c_void_p, POINTER(ErrInf)]
    SAP.RfcCloseConnection.restype = c_ulong

    SAP.RfcCreateFunction.argtypes = [c_void_p, POINTER(ErrInf)]
    SAP.RfcCreateFunction.restype = c_void_p

    SAP.RfcCreateFunctionDesc.argtypes = [c_wchar_p, POINTER(ErrInf)]
    SAP.RfcCreateFunctionDesc.restype = c_void_p

    SAP.RfcDestroyFunction.argtypes = [c_void_p, POINTER(ErrInf)]
    SAP.RfcDestroyFunction.restype = c_ulong

    SAP.RfcDestroyFunctionDesc.argtypes = [c_void_p, POINTER(ErrInf)]
    SAP.RfcDestroyFunctionDesc.restype = c_ulong

    SAP.RfcGetChars.argtypes = [c_void_p, c_wchar_p, c_void_p, c_ulong, \
    POINTER(ErrInf)]
    SAP.RfcGetChars.restype = c_ulong

    SAP.RfcGetCurrentRow.argtypes = [c_void_p, POINTER(ErrInf)]
    SAP.RfcGetCurrentRow.restype = c_void_p

    SAP.RfcGetFunctionDesc.argtypes = [c_void_p, c_wchar_p, POINTER(ErrInf)]
    SAP.RfcGetFunctionDesc.restype = c_void_p

    SAP.RfcGetRowCount.argtypes = [c_void_p, POINTER(c_ulong), \
    POINTER(ErrInf)]
    SAP.RfcGetRowCount.restype = c_ulong

    SAP.RfcGetStructure.argtypes = [c_void_p, c_wchar_p, \
    POINTER(c_void_p), POINTER(ErrInf)]
    SAP.RfcGetStructure.restype = c_ulong

    SAP.RfcGetTable.argtypes = [c_void_p, c_wchar_p, POINTER(c_void_p), \
    POINTER(ErrInf)]
    SAP.RfcGetTable.restype = c_ulong

    SAP.RfcGetVersion.argtypes = [POINTER(c_ulong), POINTER(c_ulong), \
    POINTER(c_ulong)]
    SAP.RfcGetVersion.restype = c_wchar_p

    SAP.RfcInstallServerFunction.argtypes = [c_wchar_p, c_void_p, \
    c_void_p, POINTER(ErrInf)]
    SAP.RfcInstallServerFunction.restype = c_ulong

    SAP.RfcInvoke.argtypes = [c_void_p, c_void_p, POINTER(ErrInf)]
    SAP.RfcInvoke.restype = c_ulong

    SAP.RfcListenAndDispatch.argtypes = [c_void_p, c_ulong, POINTER(ErrInf)]
    SAP.RfcListenAndDispatch.restype = c_ulong

    SAP.RfcMoveToFirstRow.argtypes = [c_void_p, POINTER(ErrInf)]
    SAP.RfcMoveToFirstRow.restype = c_ulong

    SAP.RfcMoveToNextRow.argtypes = [c_void_p, POINTER(ErrInf)]
    SAP.RfcMoveToNextRow.restype = c_ulong

    SAP.RfcOpenConnection.argtypes = [POINTER(ConnParams), c_ulong, \
    POINTER(ErrInf)]
    SAP.RfcOpenConnection.restype = c_void_p

    SAP.RfcPing.argtypes = [c_void_p, POINTER(ErrInf)]
    SAP.RfcPing.restype = c_ulong

    SAP.RfcRegisterServer.argtypes = [POINTER(SConParams), c_ulong, \
    POINTER(ErrInf)]
    SAP.RfcRegisterServer.restype = c_void_p

    SAP.RfcSetChars.argtypes = [c_void_p, c_wchar_p, c_wchar_p, c_ulong, \
    POINTER(ErrInf)]
    SAP.RfcSetChars.restype = c_ulong

    RfcConnParams[0].name = "ASHOST"; RfcConnParams[0].value = request.data['host']
    RfcConnParams[1].name = "SYSNR" ; RfcConnParams[1].value = request.data['sysnr']            
    RfcConnParams[2].name = "CLIENT"; RfcConnParams[2].value = request.data['client']      
    RfcConnParams[3].name = "USER"  ; RfcConnParams[3].value = request.data['username']     
    RfcConnParams[4].name = "PASSWD"; RfcConnParams[4].value = request.data['password']  


    hRFC = SAP.RfcOpenConnection(RfcConnParams, 5, RfcErrInf)
    # hRFC = ""
    if hRFC != None:
        return Response(status=status.HTTP_200_OK)
    else:
        print(RfcErrInf.message)
    return Response(status=status.HTTP_404_NOT_FOUND)







#To Fetch SAP Tables Names and Description
# @api_view(['GET'])
def saptables_to_sqlite(connection_id):
    try:
        ashost = ""
        sysnr = ""
        client = ""
        user = ""
        password = ""
        connection_object = ""
        if Connection.objects.filter(connection_id=connection_id).exists():
            conn = Connection.objects.get(connection_id=connection_id)
            if conn.status == 'InActive':
                return Response(status=status.HTTP_406_NOT_ACCEPTABLE , data = "Connection is InActive")
            else:
                ashost = conn.host
                sysnr = conn.sysnr
                client = conn.client
                user = conn.username
                password = conn.password
                connection_object = conn
        else:
            return Response(status=status.HTTP_404_NOT_FOUND,data = "Connection Not Found")
    
    
    
        class RFC_ERROR_INFO(Structure):
            _fields_ = [("code", c_ulong),
                        ("group", c_ulong),
                        ("key", c_wchar * 128),
                        ("message", c_wchar * 512),
                        ("abapMsgClass", c_wchar * 21),
                        ("abapMsgType", c_wchar * 2),
                        ("abapMsgNumber", c_wchar * 4),
                        ("abapMsgV1", c_wchar * 51),
                        ("abapMsgV2", c_wchar * 51),
                        ("abapMsgV3", c_wchar * 51),
                        ("abapMsgV4", c_wchar * 51)]
    
        class RFC_CONNECTION_PARAMETER(Structure):
            _fields_ = [("name", c_wchar_p),
                        ("value", c_wchar_p)]
    
    
        #-Constants-------------------------------------------------------------
    
        #-RFC_RC - RFC return codes---------------------------------------------
        RFC_OK = 0
        RFC_COMMUNICATION_FAILURE = 1
        RFC_LOGON_FAILURE = 2
        RFC_ABAP_RUNTIME_FAILURE = 3
        RFC_ABAP_MESSAGE = 4
        RFC_ABAP_EXCEPTION = 5
        RFC_CLOSED = 6
        RFC_CANCELED = 7
        RFC_TIMEOUT = 8
        RFC_MEMORY_INSUFFICIENT = 9
        RFC_VERSION_MISMATCH = 10
        RFC_INVALID_PROTOCOL = 11
        RFC_SERIALIZATION_FAILURE = 12
        RFC_INVALID_HANDLE = 13
        RFC_RETRY = 14
        RFC_EXTERNAL_FAILURE = 15
        RFC_EXECUTED = 16
        RFC_NOT_FOUND = 17
        RFC_NOT_SUPPORTED = 18
        RFC_ILLEGAL_STATE = 19
        RFC_INVALID_PARAMETER = 20
        RFC_CODEPAGE_CONVERSION_FAILURE = 21
        RFC_CONVERSION_FAILURE = 22
        RFC_BUFFER_TOO_SMALL = 23
        RFC_TABLE_MOVE_BOF = 24
        RFC_TABLE_MOVE_EOF = 25
        RFC_START_SAPGUI_FAILURE = 26
        RFC_ABAP_CLASS_EXCEPTION = 27
        RFC_UNKNOWN_ERROR = 28
        RFC_AUTHORIZATION_FAILURE = 29
    
        #-RFCTYPE - RFC data types----------------------------------------------
        RFCTYPE_CHAR = 0
        RFCTYPE_DATE = 1
        RFCTYPE_BCD = 2
        RFCTYPE_TIME = 3
        RFCTYPE_BYTE = 4
        RFCTYPE_TABLE = 5
        RFCTYPE_NUM = 6
        RFCTYPE_FLOAT = 7
        RFCTYPE_INT = 8
        RFCTYPE_INT2 = 9
        RFCTYPE_INT1 = 10
        RFCTYPE_NULL = 14
        RFCTYPE_ABAPOBJECT = 16
        RFCTYPE_STRUCTURE = 17
        RFCTYPE_DECF16 = 23
        RFCTYPE_DECF34 = 24
        RFCTYPE_XMLDATA = 28
        RFCTYPE_STRING = 29
        RFCTYPE_XSTRING = 30
        RFCTYPE_BOX = 31
        RFCTYPE_GENERIC_BOX = 32
    
        #-RFC_UNIT_STATE - Processing status of a background unit---------------
        RFC_UNIT_NOT_FOUND = 0
        RFC_UNIT_IN_PROCESS = 1
        RFC_UNIT_COMMITTED = 2
        RFC_UNIT_ROLLED_BACK = 3
        RFC_UNIT_CONFIRMED = 4
    
        #-RFC_CALL_TYPE - Type of an incoming function call---------------------
        RFC_SYNCHRONOUS = 0
        RFC_TRANSACTIONAL = 1
        RFC_QUEUED = 2
        RFC_BACKGROUND_UNIT = 3
    
        #-RFC_DIRECTION - Direction of a function module parameter--------------
        RFC_IMPORT = 1
        RFC_EXPORT = 2
        RFC_CHANGING = RFC_IMPORT + RFC_EXPORT
        RFC_TABLES = 4 + RFC_CHANGING
    
        #-RFC_CLASS_ATTRIBUTE_TYPE - Type of an ABAP object attribute-----------
        RFC_CLASS_ATTRIBUTE_INSTANCE = 0
        RFC_CLASS_ATTRIBUTE_CLASS = 1
        RFC_CLASS_ATTRIBUTE_CONSTANT = 2
    
        #-RFC_METADATA_OBJ_TYPE - Ingroup repository----------------------------
        RFC_METADATA_FUNCTION = 0
        RFC_METADATA_TYPE = 1
        RFC_METADATA_CLASS = 2
    
    
        #-Variables-------------------------------------------------------------
        ErrInf = RFC_ERROR_INFO; RfcErrInf = ErrInf()
        ConnParams = RFC_CONNECTION_PARAMETER * 5; RfcConnParams = ConnParams()
        SConParams = RFC_CONNECTION_PARAMETER * 3; RfcSConParams = SConParams()
    
    
        #-Library---------------------------------------------------------------
        # if str(platform.architecture()[0]) == "32bit":
        #   os.environ['PATH'] += ";C:\\SAPRFCSDK\\32bit"
        #   SAPNWRFC = "C:\\SAPRFCSDK\\32bit\\sapnwrfc.dll"
        # elif str(platform.architecture()[0]) == "64bit":
        #   os.environ['PATH'] += ";C:\\SAPRFCSDK\\64bit"
        #   SAPNWRFC = "C:\\SAPRFCSDK\\64bit\\sapnwrfc.dll"
    
        SAPNWRFC = "sapnwrfc.dll"
    
        SAP = windll.LoadLibrary(SAPNWRFC)
    
        #-Prototypes------------------------------------------------------------
        SAP.RfcAppendNewRow.argtypes = [c_void_p, POINTER(ErrInf)]
        SAP.RfcAppendNewRow.restype = c_void_p
    
        SAP.RfcCloseConnection.argtypes = [c_void_p, POINTER(ErrInf)]
        SAP.RfcCloseConnection.restype = c_ulong
    
        SAP.RfcCreateFunction.argtypes = [c_void_p, POINTER(ErrInf)]
        SAP.RfcCreateFunction.restype = c_void_p
    
        SAP.RfcSetInt.argtypes = [c_void_p, c_wchar_p, c_ulong, POINTER(ErrInf)]
        SAP.RfcSetInt.restype = c_ulong
    
        SAP.RfcCreateFunctionDesc.argtypes = [c_wchar_p, POINTER(ErrInf)]
        SAP.RfcCreateFunctionDesc.restype = c_void_p
    
        SAP.RfcDestroyFunction.argtypes = [c_void_p, POINTER(ErrInf)]
        SAP.RfcDestroyFunction.restype = c_ulong
    
        SAP.RfcDestroyFunctionDesc.argtypes = [c_void_p, POINTER(ErrInf)]
        SAP.RfcDestroyFunctionDesc.restype = c_ulong
    
        SAP.RfcGetChars.argtypes = [c_void_p, c_wchar_p, c_void_p, c_ulong, \
        POINTER(ErrInf)]
        SAP.RfcGetChars.restype = c_ulong
    
        SAP.RfcGetCurrentRow.argtypes = [c_void_p, POINTER(ErrInf)]
        SAP.RfcGetCurrentRow.restype = c_void_p
    
        SAP.RfcGetFunctionDesc.argtypes = [c_void_p, c_wchar_p, POINTER(ErrInf)]
        SAP.RfcGetFunctionDesc.restype = c_void_p
    
        SAP.RfcCreateTable.argtypes = [c_void_p, POINTER(ErrInf)]
        SAP.RfcCreateTable.restype = c_void_p
    
    
        SAP.RfcGetRowCount.argtypes = [c_void_p, POINTER(c_ulong), \
        POINTER(ErrInf)]
        SAP.RfcGetRowCount.restype = c_ulong
    
        SAP.RfcGetStructure.argtypes = [c_void_p, c_wchar_p, \
        POINTER(c_void_p), POINTER(ErrInf)]
        SAP.RfcGetStructure.restype = c_ulong
    
        SAP.RfcGetTable.argtypes = [c_void_p, c_wchar_p, POINTER(c_void_p), \
        POINTER(ErrInf)]
        SAP.RfcGetTable.restype = c_ulong
    
        SAP.RfcGetVersion.argtypes = [POINTER(c_ulong), POINTER(c_ulong), \
        POINTER(c_ulong)]
        SAP.RfcGetVersion.restype = c_wchar_p
    
        SAP.RfcInstallServerFunction.argtypes = [c_wchar_p, c_void_p, \
        c_void_p, POINTER(ErrInf)]
        SAP.RfcInstallServerFunction.restype = c_ulong
    
        SAP.RfcInvoke.argtypes = [c_void_p, c_void_p, POINTER(ErrInf)]
        SAP.RfcInvoke.restype = c_ulong
    
        SAP.RfcListenAndDispatch.argtypes = [c_void_p, c_ulong, POINTER(ErrInf)]
        SAP.RfcListenAndDispatch.restype = c_ulong
    
        SAP.RfcMoveToFirstRow.argtypes = [c_void_p, POINTER(ErrInf)]
        SAP.RfcMoveToFirstRow.restype = c_ulong
    
        SAP.RfcMoveToNextRow.argtypes = [c_void_p, POINTER(ErrInf)]
        SAP.RfcMoveToNextRow.restype = c_ulong
    
        SAP.RfcOpenConnection.argtypes = [POINTER(ConnParams), c_ulong, \
        POINTER(ErrInf)]
        SAP.RfcOpenConnection.restype = c_void_p
    
        SAP.RfcPing.argtypes = [c_void_p, POINTER(ErrInf)]
        SAP.RfcPing.restype = c_ulong
    
        SAP.RfcRegisterServer.argtypes = [POINTER(SConParams), c_ulong, \
        POINTER(ErrInf)]
        SAP.RfcRegisterServer.restype = c_void_p
    
        SAP.RfcSetChars.argtypes = [c_void_p, c_wchar_p, c_wchar_p, c_ulong, \
        POINTER(ErrInf)]
        SAP.RfcSetChars.restype = c_ulong
    
    
    
    
    
            
    
    
        # FileName = "sapnwrfc.py"
        # exec(compile(open(FileName).read(), FileName, "exec"))
    
        #-Main------------------------------------------------------------------
    
        RfcConnParams[0].name = "ASHOST"; RfcConnParams[0].value = ashost
        RfcConnParams[1].name = "SYSNR" ; RfcConnParams[1].value = sysnr
        RfcConnParams[2].name = "CLIENT"; RfcConnParams[2].value = client
        RfcConnParams[3].name = "USER"  ; RfcConnParams[3].value = user
        RfcConnParams[4].name = "PASSWD"; RfcConnParams[4].value = password 
    
        tables = []
        res = []
        val = 50
        hRFC = SAP.RfcOpenConnection(RfcConnParams, 5, RfcErrInf)
        if hRFC != None:
    
            charBuffer = create_unicode_buffer(1048576 + 1)
            charBuffer1 = create_unicode_buffer(1048576 + 1)
    
        hFuncDesc = SAP.RfcGetFunctionDesc(hRFC, "ZTABLE_NAMES_DESC", RfcErrInf)
        if hFuncDesc != 0:
            hFunc = SAP.RfcCreateFunction(hFuncDesc, RfcErrInf)
            if hFunc != 0:
                rc = SAP.RfcSetInt(hFunc, "N",val, RfcErrInf)
                # print(SAP.RfcInvoke(hRFC, hFunc, RfcErrInf))
                if SAP.RfcInvoke(hRFC, hFunc, RfcErrInf) == RFC_OK:
        
                    hTable = c_void_p(0)
                    print(SAP.RfcGetTable(hFunc, "DATA", hTable, RfcErrInf))
                    if SAP.RfcGetTable(hFunc, "DATA", hTable, RfcErrInf) == RFC_OK:
                        RowCount = c_ulong(0)
                    rc = SAP.RfcGetRowCount(hTable, RowCount, RfcErrInf)
                    print(RowCount)
                    rc = SAP.RfcMoveToFirstRow(hTable, RfcErrInf)
                    for i in range(0, RowCount.value):
                        hRow = SAP.RfcGetCurrentRow(hTable, RfcErrInf)
                        rc = SAP.RfcGetChars(hRow, "TAB", charBuffer, 512, RfcErrInf)
                        rc = SAP.RfcGetChars(hRow, "DESC", charBuffer1, 512, RfcErrInf)
                        # print(str(charBuffer.value))
                        # tables.append(dict(table = str(charBuffer.value).strip(),desc = str(charBuffer1.value)))
                        res.append(str(charBuffer.value) + "~" + str(charBuffer1.value)) # Print as a dictionary
                        if i < RowCount.value:
                            rc = SAP.RfcMoveToNextRow(hTable, RfcErrInf)
        
                rc = SAP.RfcDestroyFunction(hFunc, RfcErrInf)
        
            rc = SAP.RfcCloseConnection(hRFC, RfcErrInf)
    
            erp_tables_description.objects.all().delete()
    
            print("Hello Yash")
            # print(res[0][0])
            customers_to_create=[]
            for i in range(RowCount.value):
                result = res[i].split("~")
                Dd02 = erp_tables_description(
                            connection_id = connection_object,
                            table = result[0],
                            description = result[1]
                        )
            
                # print(i+" "+RowCount.value)
                # break
    
                customers_to_create.append(Dd02)
                if i%1000 == 0:
                    erp_tables_description.objects.bulk_create(customers_to_create, ignore_conflicts=True)
                    customers_to_create = []
            # Customer.objects.all().delete()
    
            # print(tables)
            return Response("Tables Fetched Successfully",status=status.HTTP_200_OK)
        else:
            print(RfcErrInf.key)    
            print(RfcErrInf.message)
    
        del SAP
    except Exception as e:
        print("Error in getting table names from ERP: ", e)
        return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data=str(e))
 
 
@api_view(['GET'])
def SAPtables(request,load,connection_id):
 
    print("Hello called Get Api")
    print("**********")
 
    try:
        print("upup")
        if erp_tables_description.objects.filter(connection_id=connection_id).exists():
            load = load * 50
            sorted_objects = erp_tables_description.objects.filter(connection_id = connection_id).order_by('table')[:load]
            # projects = project_details.objects.all()
            serializer = ErpTablesSerializer(sorted_objects, many=True)
            return Response(serializer.data)
 
 
        if Connection.objects.filter(connection_id=connection_id).exists():
            conn = Connection.objects.get(connection_id=connection_id)
            if conn.status == 'InActive':
                return Response(status=status.HTTP_406_NOT_ACCEPTABLE , data = "Connection is InActive")
            else:
                saptables_to_sqlite(connection_id)
                load = load * 50
                print("up")
                sorted_objects = erp_tables_description.objects.filter(connection_id = connection_id).order_by('table')[:load]
                # projects = project_details.objects.all()
                print("down")
                serializer = ErpTablesSerializer(sorted_objects, many=True)
                return Response(serializer.data)
 
        else:
            return Response(status=status.HTTP_404_NOT_FOUND,data = "Connection Not Found")
   
    except Exception as e:
        print(e)
        return Response(status=status.HTTP_404_NOT_FOUND,data = "Error while fetching tables")
   
   
@api_view(['GET'])
def SAPTableSearch(request,tab,connection_id):
    print("Hello called search Get Api")
 
      # 1. Query for starts with
    starts_with_objects = erp_tables_description.objects.filter(connection_id  = connection_id , table__istartswith=tab)
 
    # 2. Query for contains (excluding starts with to avoid duplicates)
    contains_objects = erp_tables_description.objects.filter(
        connection_id  = connection_id,
        table__icontains=tab
    ).exclude(table__istartswith=tab)  # Exclude the starts_with results
 
    # 3. Combine and order the results
    combined_objects = (starts_with_objects.annotate(order_priority=Value(0, output_field=IntegerField()))  # starts with priority 0
                        .union(contains_objects.annotate(order_priority=Value(1, output_field=IntegerField()))) # contains priority 1
                        .order_by('order_priority', 'table')) # order by priority and then table name
 
    serializer = ErpTablesSerializer(combined_objects, many=True)
    return Response(serializer.data)











l=[False,""]

@api_view(['POST'])
def HANAconn(request):
    print("Hana")
    print(request.data)
    try:
        conn = dbapi.connect(
            # address="10.56.7.40",
            address = request.data['host'],
            # port=30015,
            port=int(request.data['port']),
            # user="SURYAC",
            # password="Surya@2727",
            # user="SAMPATHS",
            # password="Sampath@123",
            # user="RUPAM",
            user = request.data['username'],
            password= request.data['password'],
            # password="Mrupa09$",
            encrypt='true',
            sslValidateCertificate='false'
        )
        print(conn.isconnected())
        l[0]=conn
        l[1] = request.data['username']
    #     cursor = conn.cursor()
    #     cursor.execute("SELECT TABLE_NAME FROM SYS.TABLES WHERE SCHEMA_NAME = '"+l[1]+"'")
    #     rows = cursor.fetchall()
    #     rows=list(rows)
    #     tables = [dict(var = str(row[0]).strip()) for row in rows]
 
    #     print(tables)
    #     # return Response(tables)
       
    #     return Response(tables,status=status.HTTP_200_OK)
    except:
        # return Response("failure")  
        return Response(status=status.HTTP_404_NOT_FOUND)
   
    if(conn.isconnected):  
        # return HttpResponse("success")
        return Response(status=status.HTTP_200_OK)
    # return HttpResponse("failure")
    return Response(status=status.HTTP_404_NOT_FOUND)


@api_view(['GET'])
def HANAtables(request,p_id,c_name):
    connection = Connection.objects.filter(project_id=p_id,connection_name=c_name)
    json_data = serialize('json', list(connection))
    json_data = json.loads(json_data)[0]['fields']
    print(json_data)
    conn = conn = dbapi.connect(
            # address="10.56.7.40",
            address = json_data['host'],
            # port=30015,
            port=int(json_data['port']),
            # user="SURYAC",
            # password="Surya@2727",
            # user="SAMPATHS",
            # password="Sampath@123",
            # user="RUPAM",
            user = json_data['username'],
            password= json_data['password'],
            # password="Mrupa09$",
            encrypt='true',
            sslValidateCertificate='false'
        )
    cursor = conn.cursor()
    cursor.execute("SELECT TABLE_NAME FROM SYS.TABLES WHERE SCHEMA_NAME = '"+json_data['username']+"'")
    rows = cursor.fetchall()
    rows=list(rows)
    tables = [dict(table = str(row[0]).strip(),desc="") for row in rows]
 
    print(tables)
    return Response(tables)  


@api_view(['POST'])
def ProjectCreate(request):
    print("Hello called Post")
    print(request.data)
    project = ProjectSerializer(data=request.data)

    # validating for already existing data
    # print("varun : ",Project.objects.filter(project_name=request.data['project_name']))

    if Project.objects.filter(project_name=request.data['project_name']):
        return Response(status=status.HTTP_406_NOT_ACCEPTABLE)

    if project.is_valid():
        project.save()
        # proj = project_details.objects.get(project_name=request.data['project_name'])
        # print("Id : ",proj.proj_id)
        return Response(project.data)
    else:
        return Response(status=status.HTTP_409_CONFLICT)


@api_view(['GET'])
def ProjectGet(request):
    print("Hello called Get Api")
    sorted_objects = Project.objects.order_by('-created_at')
    # projects = project_details.objects.all()
    serializer = ProjectSerializer(sorted_objects, many=True)
    return Response(serializer.data)


@api_view(['PUT'])
def projectUpdate(request,pk):
    print("Hello called update")
    print(request.data)
    project = Project.objects.get(project_id=pk)
    data = ProjectSerializer(instance=project, data=request.data)

    

    if data.is_valid():
        data.save()
        print("edjnkfhjrvfh")
        return Response(data.data)
    else:
        return Response(status=status.HTTP_404_NOT_FOUND)



@api_view(['DELETE'])
def project_delete(request,pk):
 
    # pk = request.data['project_name']
    print("Hello called Delete")
    if Project.objects.filter(project_id=pk).exists():
        project = Project.objects.get(project_id=pk)
        if project:
            obj = objects.objects.filter(project_id = pk)
            if obj:
                for object in obj:
                    seg = segments.objects.filter(project_id = pk , obj_id = object.obj_id)
                    for segment in seg:
                        deleteSqlLiteTable(segment.table_name)
                        deleteSqlLiteTable(segment.table_name+"_src")
                        deleteSqlLiteTable(segment.table_name+"_val")
                        deleteSqlLiteTable(segment.table_name+"_err")

            serializer = ProjectSerializer(project)
            project.delete()
            return Response(serializer.data,status=status.HTTP_202_ACCEPTED)
        else:
            return Response(status=status.HTTP_404_NOT_FOUND)
    else:
        return Response(status=status.HTTP_404_NOT_FOUND)





 
@api_view(['POST'])
def ConnectionCreate(request):
    # request.data['connection_type']=""
    try:
        print(request.data)
        data = request.data
        prjid = data['project_id']
        data['connection_type'] = data['connection_type'].upper()
        if Project.objects.filter(project_id = prjid):
            prj = Project.objects.get(project_id = prjid)
            prjName = prj.project_name
            data['project_name'] = prjName
        connection = ConnectionSerializer(data=data)
        print("Hello post connection called")
        if Connection.objects.filter(project_id=request.data["project_id"],connection_name = request.data["connection_name"]).exists():
            return Response(status=status.HTTP_406_NOT_ACCEPTABLE)
        if connection.is_valid():
           
            connection.save()
            return Response(connection.data,status=status.HTTP_201_CREATED)
        else:
            return Response(status=status.HTTP_409_CONFLICT)
    except Exception as e:
        print(f"Error in connection creation: {e}")  # Log the error for debugging
        return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)
   
@api_view(['GET'])
def ConnectionGet(request):
    try:
        print("hii")
        connections = Connection.objects.all()
        serializer = ConnectionSerializer(connections,many=True)
        return Response(serializer.data)
    except Exception as e:
        print(f"Error in Getting Connections: {e}")  # Log the error for debugging
        return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)
 
@api_view(['PUT'])
def ConnectionUpdate(request,p_id,c_name):
    try:
        print(request.data)
        print(p_id,c_name)
        connection = Connection.objects.filter(project_id=p_id,connection_name=c_name)
        if connection:
            connection = Connection.objects.get(project_id=p_id,connection_name=c_name)
            info = request.data
            prjid = info['project_id']
            info['connection_type'] = info['connection_type'].upper()
            if Project.objects.filter(project_id = prjid):
                prj = Project.objects.get(project_id = prjid)
                prjName = prj.project_name
                info['project_name'] = prjName
            data = ConnectionSerializer(instance=connection, data=info)
            if data.is_valid():
                data.save()
                return Response(data.data,status=status.HTTP_202_ACCEPTED)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND)
        else:
            return Response(status=status.HTTP_404_NOT_FOUND)
       
    except Exception as e:
        print(f"Error in Updating connection: {e}")  # Log the error for debugging
        return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)
   
@api_view(['DELETE'])
def connectionDelete(request,p_id,c_name):
    if Connection.objects.filter(project_id=p_id,connection_name=c_name).exists():
        connection = Connection.objects.get(project_id=p_id,connection_name=c_name)
        if connection:
            connection.delete()
            print("ssssssuccesssss")
            return Response(c_name,status=status.HTTP_202_ACCEPTED)
        else:
            return Response(status=status.HTTP_404_NOT_FOUND)
    else:
        return Response(status=status.HTTP_404_NOT_FOUND)  
 
@api_view(['GET'])
def ConnectionGetSingle(request,p_id,c_name):
    if Connection.objects.filter(project_id=p_id,connection_name=c_name).exists():
        connection = Connection.objects.get(project_id=p_id,connection_name=c_name)
        if connection:
            serializer = ConnectionSerializer(connection)
            return Response(serializer.data)
        else:
            return Response(status=status.HTTP_404_NOT_FOUND)
    else:
        return Response(status=status.HTTP_404_NOT_FOUND)  
   
@api_view(['GET'])
def ProjectGetSingle(request,p_id):
 
    try:
        if Project.objects.filter(project_id=p_id).exists():
            project = Project.objects.get(project_id=p_id)
            if project:
                serializer = ProjectSerializer(project)
                return Response(serializer.data)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND)
        else:
            return Response(status=status.HTTP_404_NOT_FOUND)  
       
    except Exception as e:
        print(f"Error in project_dataObject: {e}")  # Log the error for debugging
        return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)
 
 
@api_view(['PUT'])
def connectionRename(request,re_val,p_id,c_name):
    # print(request.data)
 
    try:
 
        connection = Connection.objects.get(project_id=p_id,connection_name=c_name)
        request.data['connection_name'] = re_val
        data = ConnectionSerializer(instance=connection, data=request.data)
        if data.is_valid():
            try:
                data.save()
                return  Response(c_name,status=status.HTTP_202_ACCEPTED)
            except:
                return Response(re_val,status=status.HTTP_404_NOT_FOUND)
        else:
            print(data.errors)
            return Response(re_val,status=status.HTTP_404_NOT_FOUND)
   
    except Exception as e:
        print(f"Error in project_dataObject: {e}")  # Log the error for debugging
        return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)
   
       
 
    
        


# @api_view(['GET'])  # This function is for dynamically creating tables and you have to pass
def create_table(table_name,columns):
 
    try:
        with connection.cursor() as cursor:
                # 1. Check if the table exists

                cursor.execute(f"SELECT name FROM sqlite_master WHERE type='table' AND name='{table_name}';")
                table_exists = cursor.fetchone() is not None
 
                if not table_exists:
                    create_table_sql = f"CREATE TABLE {table_name} ("
                    for col_name, col_type in columns:
                        # valid_col_name = col_name.replace("/", "_")
                        # create_table_sql += f"{col_name} {col_type},"
                        create_table_sql += f"\"{col_name}\" {col_type},"
                    create_table_sql = create_table_sql[:-1] + ")"
                    print(create_table_sql)
                    # with transaction.atomic(using='default'):  
                    cursor.execute(create_table_sql)
                    print(f"Table '{table_name}' created.")
                    return "success"
 
    except Exception as e:
        print(f"Error creating table: {e}")
        connection.rollback()
        return "Error"
        # return Response(f"Error creating/inserting data: {e}", status=500)
 
 

 
@api_view(['GET'])
def viewDynamic(request):
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
            tables = [row[0] for row in cursor.fetchall()]  # Extract table names
            return Response(tables)
 
 
        #     table_name = "bala7"
        #     cursor.execute(f"SELECT * FROM {table_name}")
        #     rows = cursor.fetchall()
 
        # # Print the data (or process it as needed)
        # ans = []
        # for row in rows:
        #     ans.append(row)
        # return Response(ans)
 
 
 
            # table_name = "bala5"
            #   # Method 1: Using PRAGMA table_info (recommended)
            # cursor.execute(f"PRAGMA table_info({table_name});")
            # columns_info = cursor.fetchall()
            # for column in columns_info:
            #     print(column)  # Print all column details
            #     print(f"Column Name: {column[1]}")
            # return Response("Hello")
 
    except Exception as e:
        print(f"Error creating/checking table: {e}")  # Print the error to the console
        connection.rollback()  # Rollback any partial changes on error
        return Response(f"Error creating/checking table: {e}", status=500)
   
 
 
def deleteSqlLiteTable(table_name):
 
    # table_name = "demo"
 
    try:
            with connection.cursor() as cursor:
                # Use parameterized query to prevent SQL injection
                cursor.execute(f"DROP TABLE IF EXISTS  {table_name}") # Correct: parameterized query
                # or cursor.execute(f"DROP TABLE IF EXISTS {table_name}") # Less secure way
                print(f"Table '{table_name}' dropped (IF EXISTS).")
    except Exception as e:
            print(f"Error dropping table '{table_name}': {e}")
   
 
 
   
 
    return Response("Hii")
 
    # columns = [
    #     ("id", "INTEGER PRIMARY KEY AUTOINCREMENT"),
    #     ("productname", "TEXT")
    #     ("price", "REAL"),
    #     ("description", "TEXT"),
    #     ("is_active", "BOOLEAN"),
    #     ("created_at", "DATETIME"),
    # ]
 
    # table_name = "bala8"
 
    # create_table(columns,table_name)
 
    # data_to_insert =
    # [
    #     {
    #         "productname": "varun A"
    #         "price": 10.99,
    #         "description": "Product A description",
    #         "is_active": True,
    #         "created_at": "2024-10-29 17:00:00",
    #     },
    #     {
    #         "productname": "Product B"
    #         "price": 20.00,
    #         "description": "Product B description",
    #         "is_active": False,
    #         "created_at": "2024-10-29 18:00:00",
    #     }
    # ]
 
 
def TableName_Modification(text):
    
    allowed_chars = string.ascii_letters + string.digits + ' '  # Add space if needed
 
    # Filter out characters not in the allowed set
    cleaned_text = ''.join(char for char in text if char in allowed_chars)
    
    return re.sub(r'\s+', '_', cleaned_text)
 
import pandas as pd
from openpyxl import load_workbook

def sheet_get(df,sheet_data,obj_id,file):
 
    print("came to sheet_get")
    # deleteSqlLiteTable()
    try:
        project_id = sheet_data['project_id']
        obj_name = sheet_data['obj_name']
        template_name  = sheet_data['template_name']
        excel_file = pd.ExcelFile(file)
        sheet_names = excel_file.sheet_names
        sheet_index = 2
       
   
        x=0
        columns = []
        # segment = "Additional Descriptions"
        group = ""
        field_data = []
        for ind,i in df.iterrows():
            col = []
            data = []
            # print("namesss",i['Sheet Name'] , " : " , i['Sheet Name']!="" )
            if i['Sheet Name']=="":
                print(i['SAP Field'])
                if i['SAP Field'] !="":
                    col.append(i['SAP Field'])
                    data.append(i['SAP Field'])
                    # if i['Type'].lower() == 'text':
                    #     col.append("TEXT")
                    # elif i['Type'].lower() == 'Number':
                    #     col.append("INTEGER")
                    # elif i['Type'].lower() == 'date':
                    #     col.append("DATE")
                    # elif i['Type'].lower() == 'boolean':
                    #     col.append("BOOLEAN")
                    # elif i['Type'].lower() == 'datetime':
                    #     col.append("DATETIME")
                    # else:
                    #     
                    col.append("TEXT")
                    columns.append(col)
                    data.append(i['Field Description'])
                    if i['Importance'] != "":
                        data.append("True")
                    else:
                        data.append("False")
                    data.append(i['SAP Structure'])
                    if(i['Group Name']=="Key"):
                        data.append("True")
                        group = "Key"
                    elif i['Group Name'] != "":
                        group = i['Group Name']
                        data.append("False")
                    elif i['Group Name'] == "":
                        if group == "Key":
                            data.append("True")
                        else:
                            data.append("False")
   
                    field_data.append(data)
            else:
                print("Columns  : ",len(columns))
                if len(columns) == 0:
                    seg_name =TableName_Modification(i['Sheet Name'])
                    tab ="t"+"_"+project_id+"_"+str(obj_name)+"_"+str(seg_name)
                    seg = i['Sheet Name']
                    seg_obj = {
                        "project_id" : project_id,
                        "obj_id" : obj_id,
                        "segement_name":seg,
                        "table_name" : tab
                    }
                    seg_instance = SegementSerializer(data=seg_obj)
                    if seg_instance.is_valid():
                        seg_id_get = seg_instance.save()
                        segment_id = seg_id_get.segment_id
   
                    else:
                        print("")
                        local_objects_delete(obj_id)
                        return "Error"
                if len(columns) != 0:
                   
                    print("creeeeeeadye")
                    status_create_table = create_table(tab,columns)
                    create_table(tab+"_src",columns)
                    # create_table(tab+"_val",columns)
                    # add_prompt_and_last_updated_on(tab+"_val")
                    # create_table(tab+"_err",columns)

                    if status_create_table == "Error":
                        print("cameeee innnnnn")
                        local_objects_delete(obj_id)
                        return "Error"
 
                    print("HIIII")
                    file_path = file        # Your Excel file
                    sheet = sheet_names[sheet_index]
                    skip_rows = [0, 1, 2, 3, 5, 6, 7]

                    # Load workbook
                    wb = load_workbook(file_path, data_only=False)  # Needed to get raw display formatting
                    ws = wb[sheet]

                    data = []
                    for ii, row in enumerate(ws.iter_rows()):
                        if ii in skip_rows:
                            continue
                        row_data = []
                        for cell in row:
                            if cell.value is None:
                                row_data.append("")
                            else:
                                # NEW: get exactly how Excel displays it
                                if hasattr(cell, 'number_format') and cell.is_date:
                                    row_data.append(cell.value.strftime('%d.%m.%Y'))
                                else:
                                    # Use the formatted value from Excel's UI representation
                                    try:
                                        display_val = cell.number_format
                                        # Trick: use cell._value as stored raw, but better: formatted
                                        s = cell._value if isinstance(cell._value, str) else str(cell._value)
                                        row_data.append(s)
                                    except:
                                        row_data.append(str(cell.value))
                        data.append(row_data)

                    df1 = pd.DataFrame(data)

                    # Set first row as header
                    df1.columns = df1.iloc[0]
                    df1 = df1[1:].reset_index(drop=True)
                    sheet_index+=1
                    print("HIIII")
                    status_inserting_data = insert_data_from_dataframe(df1,tab)
                    status_inserting_data = insert_data_from_dataframe(df1,tab+"_src")

                    if status_inserting_data == "Error":
                        local_objects_delete(obj_id)
                        return "Error"
                    # print("came1")
               
   
                    for d in field_data:
                        field_obj = {
                            "project_id" : project_id,
                            "obj_id" : obj_id,
                            "segement_id" : segment_id,
                            "sap_structure" : d[3],
                            "fields" : d[0],
                            "description" : d[1],
                            "isMandatory" : d[2],
                            "isKey" : d[4]
                        }
                        field_instance = FieldSerializer(data=field_obj)
                        if field_instance.is_valid():
                            field_id_get = field_instance.save()
                            field_id = field_id_get.field_id
                            # print("came2")
                        else:
                            local_objects_delete(obj_id)
                            return "Error"
   
                    seg = i['Sheet Name']
                    seg_name = TableName_Modification(i['Sheet Name'])
                    tab ="t"+"_"+project_id+"_"+str(obj_name)+"_"+str(seg_name)
                    seg_obj = {
                        "project_id" : project_id,
                        "obj_id" : obj_id,
                        "segement_name":seg,
                        "table_name" : tab
                    }
                    # break
                    seg_instance = SegementSerializer(data=seg_obj)
                    if seg_instance.is_valid():
                        seg_id_get = seg_instance.save()
                        segment_id = seg_id_get.segment_id
                   
                    else:
                        local_objects_delete(obj_id)
                        return Response("Error")
                    columns=[]
                    field_data=[]

        status_create_table = create_table(tab,columns)
        create_table(tab+"_src",columns)
        # create_table(tab+"_val",columns)
        # add_prompt_and_last_updated_on(tab+"_val")
        # create_table(tab+"_err",columns)
        if status_create_table == "Error":
            local_objects_delete(obj_id)
            return "Error"
       
        file_path = file        # Your Excel file
        sheet = sheet_names[sheet_index]
        skip_rows = [0, 1, 2, 3, 5, 6, 7]

        # Load workbook
        wb = load_workbook(file_path, data_only=False)  # Needed to get raw display formatting
        ws = wb[sheet]

        data = []
        for ii, row in enumerate(ws.iter_rows()):
            if ii in skip_rows:
                continue
            row_data = []
            for cell in row:
                if cell.value is None:
                    row_data.append("")
                else:
                    # NEW: get exactly how Excel displays it
                    if hasattr(cell, 'number_format') and cell.is_date:
                        row_data.append(cell.value.strftime('%d.%m.%Y'))
                    else:
                        # Use the formatted value from Excel's UI representation
                        try:
                            display_val = cell.number_format
                            # Trick: use cell._value as stored raw, but better: formatted
                            s = cell._value if isinstance(cell._value, str) else str(cell._value)
                            row_data.append(s)
                        except:
                            row_data.append(str(cell.value))
            data.append(row_data)

        df1 = pd.DataFrame(data)
        df1.columns = df1.iloc[0]
        df1 = df1[1:].reset_index(drop=True)
        print("toooppppp",df1.head())
        sheet_index+=1
        status_inserting_data = insert_data_from_dataframe(df1,tab)
        status_inserting_data = insert_data_from_dataframe(df1,tab+"_src")

        if status_inserting_data == "Error":
            local_objects_delete(obj_id)
            return "Error"
 
        for d in field_data:
            field_obj = {
                "project_id" : project_id,
                "obj_id" : obj_id,
                "segement_id" : segment_id,
                "sap_structure" : d[3],
                "fields" : d[0],
                "description" : d[1],
                "isMandatory" : d[2],
                "isKey" : d[4]
            }
            field_instance = FieldSerializer(data=field_obj)
            if field_instance.is_valid():
                field_id_get = field_instance.save()
                field_id = field_id_get.field_id
            else:
                local_objects_delete(obj_id)
                return Response("Error")
        return "Success"
    except Exception as e:
       
        local_objects_delete(obj_id)
 
        print(f"Error in object creation: {e}")  # Log the error for debugging
        return "Error"
 
 
 
 
 

@api_view(['GET'])
def project_dataObject(request,pid,ptype):
    try:
        if pid == 0:
            obj = objects.objects.all()
            if obj:
                serializer = ObjectSerializer(obj,many=True)
                return Response(serializer.data)
            else:
                print("errors")
                return Response(status=status.HTTP_404_NOT_FOUND)
        else:
            obj = objects.objects.filter(project_id=pid) 
            if obj:
                serializer = ObjectSerializer(obj,many=True)
                return Response(serializer.data)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        print(f"Error in project_dataObject: {e}")  # Log the error for debugging
        return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)
   
 
@api_view(['GET'])
def DataObject_Segements(request,pid,oid):
 
    connections = segments.objects.filter(project_id=pid,obj_id=oid)
    if connections:
        serializer = SegementSerializer(connections,many=True)
        return Response(serializer.data)
    else:
        return Response(status=status.HTTP_404_NOT_FOUND)
   
 
 
@api_view(['GET'])
def Segements_Fields(request,pid,oid,sid):
 
    connections = fields.objects.filter(project_id=pid,obj_id=oid,segement_id=sid)
    if connections:
        serializer = FieldSerializer(connections,many=True)
        # print(serializer.data)
        return Response(serializer.data)
    else:
        return Response(status=status.HTTP_404_NOT_FOUND)
 

@api_view(['POST'])
def objects_create(request):
    print("Hello called Objects Post")
    file = request.FILES['file']
    obj_name = request.data['obj_name']
    project_id = request.data['project_id']
    template_name = request.data['template_name']
    ob_name = obj_name.strip()  
   
    prjName = ""
    project_type = ""
    if Project.objects.filter(project_id = project_id):
        prj = Project.objects.get(project_id = project_id)
        prjName = prj.project_name
        project_type = prj.project_type
        
    sess_id = DMTool().create_session_id()
    obj_data = {
        "obj_name" : ob_name,
        "project_id" : project_id,
        "template_name" : template_name,
        "project_name" : prjName,
        "project_type" : project_type,
        "session_id" : sess_id
    }
    print(obj_data)
    print("Heloooooooooooooo")
    obj = ObjectSerializer(data=obj_data)
 
   
    if objects.objects.filter(project_id=obj_data['project_id'],obj_name = obj_data['obj_name']):
        return Response(status=status.HTTP_406_NOT_ACCEPTABLE)
 
    if obj.is_valid():
        obj_instance=obj.save()
        objid = obj_instance.obj_id
 
        df = pd.read_excel(file,sheet_name="Field List",skiprows=[0,1,2],na_filter=False)
        status_creation = sheet_get(df,obj_data,objid,file)

        if status_creation == "Error":
            return Response(status=status.HTTP_404_NOT_FOUND)
 
 
        return Response(obj.data)
    else:
        print(obj.error_messages)
        return Response(status=status.HTTP_409_CONFLICT)
 
 
 
@api_view(['PUT'])
def objects_update(request,oid):
 
    print("Hello called objects update")
    # return Response("Hello")
    # print(request.data)
   
 
    file = request.FILES['file']
    obj_name = request.data['obj_name']
    project_id = request.data['project_id']
    template_name = request.data['file_name']
 
    prjName = ""
    project_type = ""
    if Project.objects.filter(project_id = project_id):
        prj = Project.objects.get(project_id = project_id)
        prjName = prj.project_name
        project_type = prj.project_type
 
    obj_data = {
        "obj_name" : obj_name,
        "project_id" : project_id,
        "template_name" : template_name,
        "project_name" : prjName,
        "project_type" : project_type
    }
 
    if objects.objects.filter(obj_id=oid).exists():
        obj = objects.objects.get(obj_id=oid)
        if obj.obj_name == obj_name:
 
            if obj:
               
                #Deleting existing segements and tables
                seg = segments.objects.filter(project_id=obj.project_id,obj_id=oid)
                for s in seg:
                    deleteSqlLiteTable(s.table_name)
                    # segSerializer = SegementSerializer(s)
                    # s.delete()
 
 
                #Creating new excel tables and details into segements and fields tables
                data = ObjectSerializer(instance=obj, data=obj_data)
                if data.is_valid():
                    obj_instance=data.save()
                    objid = obj_instance.obj_id
 
                    df = pd.read_excel(file,sheet_name="Field List",skiprows=[0,1,2],na_filter=False)
                    # print(df)
                    sheet_delete(df,obj_data,objid)
                    sheet_update(df,obj_data,objid,file)
 
                    return Response(data.data)
                else:
                    return Response(status=status.HTTP_406_NOT_ACCEPTABLE)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND)
 
 
 
                # return Response(serializer.data,status=status.HTTP_202_ACCEPTED)
        else:
                return Response(status=status.HTTP_401_UNAUTHORIZED)
    else:
        return Response(status=status.HTTP_404_NOT_FOUND)
 
 
 
 
 
 
 
def local_objects_delete(oid):
     print("Deletion Method called")
     if objects.objects.filter(obj_id=oid).exists():
        obj = objects.objects.get(obj_id=oid)
        print("Yes object Exists")
        if obj:
            seg = segments.objects.filter(project_id=obj.project_id,obj_id=oid)
            if seg:
                for s in seg:
                    deleteSqlLiteTable(s.table_name)
                    deleteSqlLiteTable(s.table_name+"_src")
                    deleteSqlLiteTable(s.table_name+"_val")
                    deleteSqlLiteTable(s.table_name+"_err")
            obj.delete()
            print("object deleted Successfully")

 


@api_view(['DELETE'])
def objects_delete(request,oid):
    print("Hello called object Delete")
    if objects.objects.filter(obj_id=oid).exists():
        obj = objects.objects.get(obj_id=oid)
        if obj:
 
            seg = segments.objects.filter(project_id=obj.project_id,obj_id=oid)
            for s in seg:
                deleteSqlLiteTable(s.table_name)
                deleteSqlLiteTable(s.table_name+"_src")
                deleteSqlLiteTable(s.table_name+"_err")
                deleteSqlLiteTable(s.table_name+"_val")
            serializer = ObjectSerializer(obj)
            obj.delete()
            return Response(serializer.data,status=status.HTTP_202_ACCEPTED)
        else:
            return Response(status=status.HTTP_404_NOT_FOUND)
    else:
        return Response(status=status.HTTP_404_NOT_FOUND)
 
 
@api_view(['GET'])
def objects_get(request,oid):
    print("Hello called object Get Api")
    obj = objects.objects.get(obj_id=oid)
    if obj:
        serializer = ObjectSerializer(obj)
        print(serializer.data['project_id'])   
        return Response(serializer.data)
    else:
        return Response(status=status.HTTP_404_NOT_FOUND)
 

 
@api_view(['POST'])
def xls_read(request):
    file = request.FILES['file']
    excel_file = pd.ExcelFile(file)
    # Get the sheet names
    sheet_names = excel_file.sheet_names
    # Print the sheet names
    print(len(sheet_names))
    print(sheet_names)
    if len(sheet_names) <= 1 :
        return Response(status=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE)
    else:
        if 'Field List' in sheet_names:
            # print("Yes Iam in ...")
            df = pd.read_excel(file, sheet_name='Field List')
           
            val = df.columns[0].split(':')
            return Response(val[1].strip())
        else:
            return Response(status=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE)
 
 
@api_view(['GET'])
def tableDelete(request):
 
    # lst = ['demo0','demo138','demo143','demo154','demo171','demo177','demo185','demo193','demo201'
    #        ,'demo206','demo218','demo227','demo278','demo290','demo312','demo321','demo496','demo490'
    #        ,'demo496','demo521','demo553','demo561','demo618','demo644','demo656','demo698']
   
    # for l in lst:
    deleteSqlLiteTable('demo469')
    return Response("Hello Deleted")



def insert_data_from_dataframe(dataframe, table_name, database_name='default'):
    print("prinnnnting tttt",dataframe)
    cursor = connections[database_name].cursor()
    cursor.execute(f"PRAGMA table_info({table_name})")
    number_columns = len(cursor.fetchall())
    dataframe = dataframe.iloc[:, :number_columns]
    print(dataframe.dtypes)
    try:
        with connections[database_name].cursor() as cursor:
            for index, row in dataframe.iterrows():
                # Construct the INSERT INTO statement
                # column_names = ', '.join(dataframe.columns)
                quoted_column_names = ', '.join(f'"{col}"' for col in dataframe.columns)
                column_names = quoted_column_names
                placeholders = ', '.join(['%s'] * len(dataframe.columns))
                insert_sql = f"INSERT INTO {table_name} ({column_names}) VALUES ({placeholders});"
 
                # Execute the INSERT statement with data from the row
                cursor.execute(insert_sql, tuple(row))
 
            # Commit the changes within a transaction
            with transaction.atomic(using=database_name):
                cursor.execute("COMMIT;")
 
        print(f"Data inserted successfully into '{table_name}' in {database_name} database.")
        return "Success"
       
 
    except Exception as e:
        print("errorrrr",dataframe.describe)
        print(f"Error inserting data: {e}")
        return "Error"


def create_table_dynamically(table_name, fields, database_name='default'):
    try:
        with connections[database_name].cursor() as cursor:
            cursor.execute(f"SELECT name FROM sqlite_master WHERE type='table' AND name='{table_name}';")
            if cursor.fetchone():
                print(f"Table '{table_name}' already exists in {database_name} database.")
                return  
            create_table_sql = f"CREATE TABLE {table_name} ("
            for field_name, field_type in fields.items():
                create_table_sql += f"\"{field_name}\" {field_type},"
            create_table_sql = create_table_sql[:-1] + ");"
            print(create_table_sql)
            with transaction.atomic(using=database_name):  
                cursor.execute(create_table_sql)
                return 1
            print(f"Table '{table_name}' created successfully in {database_name} database.")
 
    except sqlite3.Error as e:
        print(f"SQLite3 Error: {e}")
    except Exception as e:
        print(f"Error creating table: {e}")
 
def convert_list_to_fields(field_list):
    field_dict = {}
    for field_name, field_type in field_list:
        # if field_type.lower() == 'text':
        #     field_dict[field_name] = 'TEXT'
        # elif field_type.lower() == 'date':
        #     field_dict[field_name] = 'DATE'
        # elif field_type.lower() == 'integer':
        #     field_dict[field_name] = 'INTEGER'
        # elif field_type.lower() == 'real':
        #     field_dict[field_name] = 'REAL'
        # elif field_type.lower() == 'boolean':
        #     field_dict[field_name] = 'BOOLEAN'
        # elif field_type.lower() == 'datetime':
        #     field_dict[field_name] = 'DATETIME'
        # else:
        #     
        field_dict[field_name] = 'TEXT'  
    return field_dict
 
def drop_table_dynamically(table_name, database_name='default'):
    try:
        with connections[database_name].cursor() as cursor:
            cursor.execute(f"SELECT name FROM sqlite_master WHERE type='table' AND name='{table_name}';")
            if not cursor.fetchone():
                print(f"Table '{table_name}' does not exist in {database_name} database.")
                return
 
            drop_table_sql = f"DROP TABLE {table_name};"
 
            with transaction.atomic(using=database_name):
                cursor.execute(drop_table_sql)
                return 1
 
            print(f"Table '{table_name}' dropped successfully from {database_name} database.")
 
    except Exception as e:
        print(f"Error dropping table: {e}")
       
@api_view(['POST'])
def fileCreate(request):
    # request.data['connection_type']=""
    connection = FileSerializer(data=request.data)
    print("Hello post file called")
    if FileConnection.objects.filter(project_id=request.data["project_id"],fileName = request.data["fileName"]).exists():
        return Response(status=status.HTTP_406_NOT_ACCEPTABLE)
    if connection.is_valid():
       
        connection.save()
        return Response(connection.data,status=status.HTTP_201_CREATED)
    else:
        return Response(status=status.HTTP_409_CONFLICT)
   
@api_view(['GET'])
def fileGet(request):
    print("hii")
    connections = FileConnection.objects.all()
    serializer = FileSerializer(connections,many=True)
    return Response(serializer.data)
 
@api_view(['PUT'])
def fileUpdate(request,p_id,f_name):
    print(request.data)
    print(p_id,f_name)
    connection = FileConnection.objects.get(project_id=p_id,fileName=f_name)
    data = FileSerializer(instance=connection, data=request.data)
    if data.is_valid():
        print("jfnjkjefkjfkjnrkj")
        data.save()
        return Response(data.data,status=status.HTTP_202_ACCEPTED)
    else:
        print("ffffffffffffffffffffffffffffffffffffff")
        return Response(status=status.HTTP_404_NOT_FOUND)
   
@api_view(['DELETE'])
def fileDelete(request,p_id,f_name):
    if FileConnection.objects.filter(project_id=p_id,fileName=f_name).exists():
        connection = FileConnection.objects.get(project_id=p_id,fileName=f_name)
        if connection:
            connection.delete()
            print("ssssssuccesssss")
            return Response(f_name,status=status.HTTP_200_OK)
        else:
            return Response(status=status.HTTP_404_NOT_FOUND)
    else:
        return Response(status=status.HTTP_404_NOT_FOUND)  
 
@api_view(['GET'])
def fileGetSingle(request,p_id,f_name):
    if FileConnection.objects.filter(project_id=p_id,fileName=f_name).exists():
        connection = FileConnection.objects.get(project_id=p_id,fileName=f_name)
        if connection:
            serializer = FileSerializer(connection)
            return Response(serializer.data)
        else:
            return Response(status=status.HTTP_404_NOT_FOUND)
    else:
        return Response(status=status.HTTP_404_NOT_FOUND)  
 
@api_view(['PUT'])
def fileRename(request,re_val,p_id,f_name):
    # print(request.data)
    connection = FileConnection.objects.get(project_id=p_id,fileName=f_name)
    request.data['fileName'] = re_val
    d={}
    d['project_id'] = request.data['project_id']
    d['fileName'] = re_val
    d['tableName'] = request.data['table_name']
    d['fileType'] = request.data['file_type']
    d['sheet'] = request.data['sheet']
    data = FileSerializer(instance=connection, data=d)
    print(data)
    if data.is_valid():
        try:
            data.save()
            return  Response(f_name,status=status.HTTP_200_OK)
        except:
            return Response(re_val,status=status.HTTP_404_NOT_FOUND)
    else:
        return Response(re_val,status=status.HTTP_404_NOT_FOUND)
 
 
 
 
class GetXL(APIView):
    def post(self, request):
        file = request.FILES['file']
        excel_file = pd.ExcelFile(file)
        # Get the sheet names
        sheet_names = excel_file.sheet_names
        # Print the sheet names
        print(sheet_names)
        # data = pd.read_excel(file)
        # data = pd.DataFrame(data)
        # column_names_list = data.columns.tolist()
        d = []
        for i in sheet_names:
            d.append(i)
        print(d)
        return Response(d)
 
class GetXLSheet(APIView):
    def post(self, request):
        data = request.data.copy()  # Create a mutable copy of request.data
 
        # Ensure project_id is an integer
        try:
            project_id = data.get('projectID')
        except (ValueError, TypeError):
            return Response({"projectID": ["Must be a valid integer."]}, status=status.HTTP_400_BAD_REQUEST)
 
        # Assign int value to project_id, where FileSerializer can get this field name
        data['project_id'] = project_id
        data['fileType'] = 'Excel'  # Set fileType directly in the data
        print(data)
        serializer = FileSerializer(data=data)
        

        isTableExists = FileConnection.objects.filter(tableName = request.data['tableName'])

        if isTableExists :
            return Response(serializer.errors, status=status.HTTP_409_CONFLICT)

        if not serializer.is_valid():
            print(serializer.errors)  # Very important for debugging
            return Response(serializer.errors, status=status.HTTP_423_LOCKED)
       
        serializer.save()
        print(request.data.get('sheet'))
        df = pd.read_excel(data['file'], sheet_name = data['sheet'])
        df = pd.DataFrame(df)
        columns = list(df.columns)
        feilds= {}
        for i in columns:
            feilds[i] = "TEXT"
        tablename=request.data['tableName']
        flag = drop_table_dynamically(str(tablename))
        print(flag)
        flag = create_table_dynamically(str(tablename),feilds,"default")
        print(flag)
        insert_data_from_dataframe(dataframe=df,table_name=tablename,database_name='default')
        return Response()
 
class GetTXT(APIView):
    def post(self, request):
        file = request.FILES['file']
        delim = request.data.get('delimiter')
        data = request.data.copy()
        try:
            project_id = data.get('projectID')
        except (ValueError, TypeError):
            return Response({"projectID": ["Must be a valid integer."]}, status=status.HTTP_400_BAD_REQUEST)
 
        # Assign int value to project_id, where FileSerializer can get this field name
        data['project_id'] = project_id
        data['fileType'] = 'Text'  # Set fileType directly in the data
        print(data)
        serializer = FileSerializer(data=data)
 
        if not serializer.is_valid():
            print(serializer.errors)  # Very important for debugging
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
       
        serializer.save()
        print(delim)
        data = pd.read_table(file)
        print(data)
        # df = pd.read_excel(data['file'], sheet_name = data['sheet'])
        df = pd.DataFrame(data)
        columns = list(df.columns)
        n, flag = len(columns), 0
        for i in columns:
            if ':' in i:
                flag = 1
                break
        if flag:
            columns = []
            for i in range(n):
                s = 'Column' + str(i)
                columns.append(s)
        print(columns)
        feilds= {}
        for i in columns:
            feilds[i] = "TEXT"
        tablename=request.data['tableName']
        flag = drop_table_dynamically(str(tablename))
        print(flag)
        flag = create_table_dynamically(str(tablename),feilds,"default")
        print(flag)
        insert_data_from_dataframe(dataframe=df,table_name=tablename,database_name='default')
        return Response()
   
class GetFile(APIView):
    def post(self, request):
        data = request.data.copy()  # Create a mutable copy of request.data
 
        # Ensure project_id is an integer
        try:
            project_id = data.get('projectID')
        except (ValueError, TypeError):
            return Response({"projectID": ["Must be a valid integer."]}, status=status.HTTP_400_BAD_REQUEST)
 
        # Assign int value to project_id, where FileSerializer can get this field name
        data['project_id'] = project_id
        data['fileType'] = 'CSV'  # Set fileType directly in the data
        print(data)
        serializer = FileSerializer(data=data)
 
        if not serializer.is_valid():
            print(serializer.errors)  # Very important for debugging
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
       
        serializer.save()
        file = request.FILES['file']
        data = pd.read_csv(file)
        columns = list(data.columns)
        print(*columns, sep = ', ')
        data = pd.DataFrame(data)
 
        df = pd.DataFrame(data)
        columns = list(df.columns)
        feilds= {}
        for i in columns:
            feilds[i] = "TEXT"
        tablename=request.data['tableName']
        flag = drop_table_dynamically(str(tablename))
        print(flag)
        flag = create_table_dynamically(str(tablename),feilds,"default")
        print(flag)
        insert_data_from_dataframe(dataframe=df,table_name=tablename,database_name='default')
        return Response()
 




class ReuploadXLSheet(APIView):
    def post(self, request):
        data = request.data.copy()  # Create a mutable copy of request.data
 
        # Ensure project_id is an integer
        try:
            project_id = data.get('projectID')
        except (ValueError, TypeError):
            return Response({"projectID": ["Must be a valid integer."]}, status=status.HTTP_400_BAD_REQUEST)
 
        # Assign int value to project_id, where FileSerializer can get this field name
        data['project_id'] = project_id
        data['fileType'] = 'Excel'  # Set fileType directly in the data
        print(data)
        serializer = FileSerializer(data=data)
        

        isTableExists = FileConnection.objects.filter(tableName = request.data['tableName'])

        if isTableExists :
            return Response(serializer.errors, status=status.HTTP_409_CONFLICT)

        if not serializer.is_valid():
            print(serializer.errors)  # Very important for debugging
            return Response(serializer.errors, status=status.HTTP_423_LOCKED)
       
        serializer.save()
        print(request.data.get('sheet'))
        df = pd.read_excel(data['file'], sheet_name = data['sheet'])
        df = pd.DataFrame(df)
        columns = list(df.columns)
        feilds= {}
        for i in columns:
            feilds[i] = "TEXT"
        tablename=request.data['tableName']
        flag = drop_table_dynamically(str(tablename))
        print(flag)
        flag = create_table_dynamically(str(tablename),feilds,"default")
        print(flag)
        insert_data_from_dataframe(dataframe=df,table_name=tablename,database_name='default')
        return Response()
 
class ReuploadTXT(APIView):
    def post(self, request):
        file = request.FILES['file']
        delim = request.data.get('delimiter')
        data = request.data.copy()
        try:
            project_id = data.get('projectID')
        except (ValueError, TypeError):
            return Response({"projectID": ["Must be a valid integer."]}, status=status.HTTP_400_BAD_REQUEST)
 
        # Assign int value to project_id, where FileSerializer can get this field name
        data['project_id'] = project_id
        data['fileType'] = 'Text'  # Set fileType directly in the data
        print(data)
        serializer = FileSerializer(data=data)
 
        if not serializer.is_valid():
            print(serializer.errors)  # Very important for debugging
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
       
        serializer.save()
        print(delim)
        data = pd.read_table(file)
        print(data)
        # df = pd.read_excel(data['file'], sheet_name = data['sheet'])
        df = pd.DataFrame(data)
        columns = list(df.columns)
        n, flag = len(columns), 0
        for i in columns:
            if ':' in i:
                flag = 1
                break
        if flag:
            columns = []
            for i in range(n):
                s = 'Column' + str(i)
                columns.append(s)
        print(columns)
        feilds= {}
        for i in columns:
            feilds[i] = "TEXT"
        tablename=request.data['tableName']
        flag = drop_table_dynamically(str(tablename))
        print(flag)
        flag = create_table_dynamically(str(tablename),feilds,"default")
        print(flag)
        insert_data_from_dataframe(dataframe=df,table_name=tablename,database_name='default')
        return Response()
   
class ReuploadCSV(APIView):
    def post(self, request):
        data = request.data.copy()  # Create a mutable copy of request.data
 
        # Ensure project_id is an integer
        try:
            project_id = data.get('projectID')
        except (ValueError, TypeError):
            return Response({"projectID": ["Must be a valid integer."]}, status=status.HTTP_400_BAD_REQUEST)
 
        # Assign int value to project_id, where FileSerializer can get this field name
        data['project_id'] = project_id
        data['fileType'] = 'CSV'  # Set fileType directly in the data
        print(data)
        serializer = FileSerializer(data=data)
 
        if not serializer.is_valid():
            print(serializer.errors)  # Very important for debugging
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
       
        serializer.save()
        file = request.FILES['file']
        data = pd.read_csv(file)
        columns = list(data.columns)
        print(*columns, sep = ', ')
        data = pd.DataFrame(data)
 
        df = pd.DataFrame(data)
        columns = list(df.columns)
        feilds= {}
        for i in columns:
            feilds[i] = "TEXT"
        tablename=request.data['tableName']
        flag = drop_table_dynamically(str(tablename))
        print(flag)
        flag = create_table_dynamically(str(tablename),feilds,"default")
        print(flag)
        insert_data_from_dataframe(dataframe=df,table_name=tablename,database_name='default')
        return Response()
 












def sheet_update(df,sheet_data,obj_id,file):
   
    project_id = sheet_data['project_id']
    obj_name = sheet_data['obj_name']
    template_name  = sheet_data['template_name']
 
    excel_file = pd.ExcelFile(file)
    sheet_names = excel_file.sheet_names
    sheet_index = 2
   
 
    x=0
    is_seg=0
    columns = []
    # segment = "Additional Descriptions"
    group = ""
    customers_to_create=[]
    field_data = []
    for ind,i in df.iterrows():
        col = []
        data = []
        # print(i['Sheet Name'] , " : " , i['Sheet Name']!="" and i['Sheet Name'] == segment)
        if i['Sheet Name']=="":
 
            if i['SAP Field'] !="":
                col.append(i['SAP Field'])
                data.append(i['SAP Field'])
                # if i['Type'].lower() == 'text':
                #     col.append("TEXT")
                # elif i['Type'].lower() == 'Number':
                #     col.append("INTEGER")
                # elif i['Type'].lower() == 'date':
                #     col.append("DATE")
                # elif i['Type'].lower() == 'boolean':
                #     col.append("BOOLEAN")
                # elif i['Type'].lower() == 'datetime':
                #     col.append("DATETIME")
                # else:
                #     
                col.append("TEXT")
                columns.append(col)
                data.append(i['Field Description'])
                if i['Importance'] != "":
                    data.append("True")
                else:
                    data.append("False")
                data.append(i['SAP Structure'])
                if(i['Group Name']=="Key"):
                    data.append("True")
                    group = "Key"
                elif i['Group Name'] != "":
                    group = i['Group Name']
                    data.append("False")
                elif i['Group Name'] == "":
                    if group == "Key":
                        data.append("True")
                    else:
                        data.append("False")
                field_data.append(data)
        else:
            # print("Columns varun : ",len(columns))
            if len(columns) == 0:
                seg_name =TableName_Modification(i['Sheet Name'])
                tab ="t"+"_"+project_id+"_"+str(obj_name)+"_"+str(seg_name)
                seg = i['Sheet Name']
                seg_obj = {
                    "project_id" : project_id,
                    "obj_id" : obj_id,
                    "segement_name":seg,
                    "table_name" : tab
                }
 
                seg_instance = segments.objects.filter(project_id=project_id,obj_id=obj_id)
                x=0
                for s in seg_instance:
                    if s.segement_name == seg:
                        x=1
                        is_seg=0
                        break
                if x==0:
                    seg_instanc = SegementSerializer(data=seg_obj)
                    if seg_instance.is_valid():
                        seg_id_get = seg_instanc.save()
                        segment_id = seg_id_get.segment_id
                        is_seg=1
                    else:
                        return Response("Error at first segement creation")
                else:
                    segment_id = s.segment_id
            if len(columns) != 0:
                   
 
                create_table(tab,columns)
                create_table(tab+"_src",columns)
                # create_table(tab+"_val",columns)
                # add_prompt_and_last_updated_on(tab+"_val")
                # create_table(tab+"_err",columns)
 
                df1 = pd.read_excel(file,sheet_name=sheet_names[sheet_index],skiprows=[0,1,2,3,5,6,7],na_filter=False)
                sheet_index+=1
                insert_data_from_dataframe(df1,tab)
                insert_data_from_dataframe(df1,tab+"_src")
 
                field_names = []
                for fie in columns:
                    field_names.append(fie[0])
               
                fields_in_table = fields.objects.filter(project_id=project_id,obj_id=obj_id,segement_id=segment_id)
               
                for v in fields_in_table:
                    if v.fields in field_names:
                        pass
                    else:
                        serlzr = FieldSerializer(v)
                        v.delete()
               
 
                # if is_seg == 1:
                for d in field_data:
                    field_obj = {
                        "project_id" : project_id,
                        "obj_id" : obj_id,
                        "segement_id" : segment_id,
                        "sap_structure" : d[3],
                        "fields" : d[0],
                        "description" : d[1],
                        "isMandatory" : d[2],
                        "isKey" : d[4]
                    }
                    field_check = fields.objects.filter(project_id=project_id,obj_id=obj_id,segement_id=segment_id)
                    y=0
                    for f in field_check:
                        if f.fields == d[0]:
                            y=1
                            break
                    if y==0:
                        field_instance = FieldSerializer(data=field_obj)
                        if field_instance.is_valid():
                            field_id_get = field_instance.save()
                            field_id = field_id_get.field_id
                        else:
                            return Response("Error at Field Creation")
                    else:
                        field_obj = {
                            "field_id" : f.field_id,
                            "project_id" : project_id,
                            "obj_id" : obj_id,
                            "segement_id" : segment_id,
                            "sap_structure" : d[3],
                            "fields" : d[0],
                            "description" : d[1],
                            "isMandatory" : d[2],
                            "isKey" : d[4]
                        }
                        field = fields.objects.get(field_id=f.field_id)
                        data = FieldSerializer(instance=field, data=field_obj)
                        # print("Fields : ",field_obj)
                        # print(data)
                        if data.is_valid():
                            # print("Valid field")
                            # print(data)
                            data.save()
                        else:
                            # print("Error : ",data.error_messages)
                            return Response("Error at Field Creation")
 
 
                field_inst = Rule.objects.filter(project_id=project_id,object_id=obj_id,segment_id=segment_id)
                if field_inst:
                    latest_version = Rule.objects.filter(
                        project_id=project_id,  # Assuming all items have the same IDs
                        object_id=obj_id,
                        segment_id=segment_id
                    ).order_by('-version_id').first()
                    field_inst = fields.objects.filter(project_id=project_id,obj_id=obj_id,segement_id=segment_id)
                    for fi in field_inst:
                        fields_tab = Rule.objects.filter(project_id=project_id,object_id=obj_id,segment_id=segment_id,version_id=latest_version.version_id,field_id=fi.field_id).first()
                        if fields_tab:
                            rule = {
                                "project_id" : project_id,
                                "object_id" : obj_id,
                                "segment_id" : segment_id,
                                "field_id" : fi.field_id,
                                "version_id" : latest_version.version_id+1,
                                "target_sap_table" : fi.sap_structure,
                                "target_sap_field" : fi.fields,
                                "source_table" : fields_tab.source_table,
                                "source_field_name" : fields_tab.source_field_name,
                                "data_mapping_rules": fields_tab.data_mapping_rules,
                                "text_description" : fi.description,
                                "isMandatory" :fi.isMandatory
                            }
                            sezr = RuleSerializer(data=rule)
                            if sezr.is_valid():
                                sezr.save()
                        else:
                            rule = {
                                "project_id" : project_id,
                                "object_id" : obj_id,
                                "segment_id" : segment_id,
                                "field_id" : fi.field_id,
                                "version_id" : latest_version.version_id+1,
                                "target_sap_table" : fi.sap_structure,
                                "target_sap_field" : fi.fields,
                                "text_description" : fi.description,
                                "isMandatory" :fi.isMandatory
                            }
                            sezr = RuleSerializer(data=rule)
                            if sezr.is_valid():
                                sezr.save()      
 
 
                seg = i['Sheet Name']
                seg_name = TableName_Modification(i['Sheet Name'])
                tab ="t"+"_"+project_id+"_"+str(obj_name)+"_"+str(seg_name)
                seg_obj = {
                    "project_id" : project_id,
                    "obj_id" : obj_id,
                    "segement_name":seg,
                    "table_name" : tab
                }
                # break
 
                seg_instance = segments.objects.filter(project_id=project_id,obj_id=obj_id)
                x=0
                for s in seg_instance:
                    if s.segement_name == seg:
                        x=1
                        is_seg=0
                        break
                if x==0:
                    seg_instanc = SegementSerializer(data=seg_obj)
                    if seg_instanc.is_valid():
                        seg_id_get = seg_instanc.save()
                        segment_id = seg_id_get.segment_id
                        is_seg=1
 
                    else:
                        return Response("Error at first segement creation")
                else:
                    segment_id = s.segment_id
 
                columns=[]
                field_data=[]
    create_table(tab,columns)
    create_table(tab+"_src",columns)
    # create_table(tab+"_val",columns)
    # add_prompt_and_last_updated_on(tab+"_val")

    # create_table(tab+"_err",columns)

    # if is_seg==1:
 
    df1 = pd.read_excel(file,sheet_name=sheet_names[sheet_index],skiprows=[0,1,2,3,5,6,7],na_filter=False)
    sheet_index+=1
    insert_data_from_dataframe(df1,tab)
    insert_data_from_dataframe(df1,tab+"_src")
 
    field_names = []
    for fie in columns:
        field_names.append(fie[0])
   
    fields_in_table = fields.objects.filter(project_id=project_id,obj_id=obj_id,segement_id=segment_id)
   
    for v in fields_in_table:
        if v.fields in field_names:
            pass
        else:
            serlzr = FieldSerializer(v)
            v.delete()
 
 
    for d in field_data:
        field_obj = {
            "project_id" : project_id,
            "obj_id" : obj_id,
            "segement_id" : segment_id,
            "sap_structure" : d[3],
            "fields" : d[0],
            "description" : d[1],
            "isMandatory" : d[2],
            "isKey" : d[4]
        }
        field_check = fields.objects.filter(project_id=project_id,obj_id=obj_id,segement_id=segment_id)
        y=0
        for f in field_check:
            if f.fields == d[0]:
                y=1
                break
        if y==0:
            field_instance = FieldSerializer(data=field_obj)
            if field_instance.is_valid():
                field_id_get = field_instance.save()
                field_id = field_id_get.field_id
            else:
                return Response("Error at Field Creation")
        else:
            field_obj = {
                            "field_id" : f.field_id,
                            "project_id" : project_id,
                            "obj_id" : obj_id,
                            "segement_id" : segment_id,
                            "sap_structure" : d[3],
                            "fields" : d[0],
                            "description" : d[1],
                            "isMandatory" : d[2],
                            "isKey" : d[4]
                        }
            field = fields.objects.get(field_id=f.field_id)
            # print("Fields : ", field_obj)
            data = FieldSerializer(instance=field, data=field_obj)
            if data.is_valid():
                # print("Valid data")
                data.save()
            else:
                return Response("Error at Field Creation")
    field_inst = Rule.objects.filter(project_id=project_id,object_id=obj_id,segment_id=segment_id)
    if field_inst:
        latest_version = Rule.objects.filter(
            project_id=project_id,  # Assuming all items have the same IDs
            object_id=obj_id,
            segment_id=segment_id
        ).order_by('-version_id').first()
        field_inst = fields.objects.filter(project_id=project_id,obj_id=obj_id,segement_id=segment_id)
        for fi in field_inst:
            fields_tab = Rule.objects.filter(project_id=project_id,object_id=obj_id,segment_id=segment_id,version_id=latest_version.version_id,field_id=fi.field_id).first()
            if fields_tab:
                rule = {
                    "project_id" : project_id,
                    "object_id" : obj_id,
                    "segment_id" : segment_id,
                    "field_id" : fi.field_id,
                    "version_id" : latest_version.version_id+1,
                    "target_sap_table" : fi.sap_structure,
                    "target_sap_field" : fi.fields,
                    "source_table" : fields_tab.source_table,
                    "source_field_name" : fields_tab.source_field_name,
                    "data_mapping_rules": fields_tab.data_mapping_rules,
                    "text_description" : fi.description
                }
                sezr = RuleSerializer(data=rule)
                if sezr.is_valid():
                    sezr.save()
            else:
                rule = {
                    "project_id" : project_id,
                    "object_id" : obj_id,
                    "segment_id" : segment_id,
                    "field_id" : fi.field_id,
                    "version_id" : latest_version.version_id+1,
                    "target_sap_table" : fi.sap_structure,
                    "target_sap_field" : fi.fields,
                    "text_description" : fi.description
                }
                sezr = RuleSerializer(data=rule)
                if sezr.is_valid():
                    sezr.save()    
 
 
 
 
def sheet_delete(df,sheet_data,obj_id):
 
 
    # deleteSqlLiteTable()
   
    project_id = sheet_data['project_id']
    obj_name = sheet_data['obj_name']
    template_name  = sheet_data['template_name']
 
    x=0
    is_seg=0
    columns = []
    # segment = "Additional Descriptions"
    sheet_names = []
    for ind,i in df.iterrows():
        # print(i['Sheet Name'] , " : " , i['Sheet Name']!="" and i['Sheet Name'] == segment)
        if i['Sheet Name']=="":
            pass
        else:
            sheet_names.append(i['Sheet Name'])
 
    print("Sheets : ",sheet_names)
 
    segment_instance = segments.objects.filter(project_id=project_id,obj_id=obj_id)
    for s in segment_instance:
        if s.segement_name in sheet_names:
            pass
        else:
            seg_delete = SegementSerializer(s)
            s.delete()




def RuleVersions(pid,oid,sid):
    latest_version = Rule.objects.filter(
                project_id=pid,  # Assuming all items have the same IDs
                object_id=oid,
                segment_id=sid
            ).order_by('-version_id').first()
 
    # print(latest_version.version_id)
    versions=[]
    if latest_version:
        for i in range(latest_version.version_id):
            versions.append({'ind':i+1})
        # print("Hello : ",versions)
        return versions
    else:
        return versions
 
 
@api_view(['GET'])
def VerisonData(request,pid,oid,sid,vid):
    print(vid)
    versiondata = Rule.objects.filter(
                project_id=pid,  # Assuming all items have the same IDs
                object_id=oid,
                segment_id=sid,
        version_id = vid    )
 
    versiondata = RuleSerializer(versiondata,many=True)
    return Response(versiondata.data)
 
@api_view(['POST'])
def SaveRuleCreate(request):
    try:
        data = request.data
        print("Save Rules:", data)

        pid = oid = sid = 0

        for item in data:
            if pid == 0 or oid == 0 or sid == 0:
                pid = item.get('project_id')
                oid = item.get('object_id')
                sid = item.get('segment_id')

            existing_record = SaveRule.objects.filter(
                project_id=item['project_id'],
                object_id=item['object_id'],
                segment_id=item['segment_id'],
                field_id=item['field_id']
            ).first()

            item['check_box'] = False
            item['last_updated_on'] = timezone.now().strftime("%Y-%m-%d %H:%M:%S")

            if existing_record:
                serializer = SaveRuleSerializer(existing_record, data=item, partial=True)
                if serializer.is_valid():
                    serializer.save()
                else:
                    print(serializer.errors)
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            else:
                serializer = SaveRuleSerializer(data=item)
                if serializer.is_valid():
                    serializer.save()
                else:
                    print(serializer.errors)
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        serializer = SaveRule.objects.filter(
            project_id=pid,
            object_id=oid,
            segment_id=sid
        )
        if serializer:
            rule = SaveRuleSerializer(serializer, many=True)
            return Response(rule.data)

        return Response(status=status.HTTP_200_OK)

    except Exception as e:   # ✅ fixed typo
        print("SaveRules Exception:", e)
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


 
 
@api_view(['GET'])
def GetSaveRule(request,pid,oid,sid):

    serializer = SaveRule.objects.filter(
            project_id = pid,
            object_id = oid,
            segment_id = sid
        )
    versions = RuleVersions(pid,oid,sid)
    rule_with_versions = []
    if serializer:
        rule = SaveRuleSerializer(serializer,many=True)
        rule_with_versions.append(rule.data)
        rule_with_versions.append(versions)
        return Response(rule_with_versions,status=status.HTTP_200_OK)
    else:
        connections = fields.objects.filter(project_id=pid,obj_id=oid,segement_id=sid)
        if connections:
            serializer = FieldSerializer(connections,many=True) 
            print(serializer.data)
            data = serializer.data
            final_data = []
            for d in data:
                item = {
                    "version_id": 0,
                    "source_table": "",
                    "source_field_name": "",
                    "data_mapping_type": "",
                    "data_mapping_rules": "",
                    "project_id": d["project_id"], 
                    "object_id": d["obj_id"],
                    "field_id": d["field_id"],
                    "segment_id": d["segement_id"], 
                    "target_sap_table": d["sap_structure"],
                    "target_sap_field": d["fields"], 
                    "text_description": d["description"],
                    # "lookup_table": "",
                    # "lookup_field": "",
                    "lookup" : False,
                    "last_updated_by": "System",
                    "last_updated_on": "",
                    "isKey": d["isKey"],
                    "check_box": False,
                    "isMandatory": d.get("isMandatory", False)
                }
                final_data.append(item)
            
            rule_with_versions.append(final_data)
            rule_with_versions.append(versions)
            return Response(rule_with_versions)
        else:
            return Response(status=status.HTTP_404_NOT_FOUND)



def LocalSaveRuleCreate(data,pid,oid,sid):
    # print("Hello called SaveRuleCreate")
    for item in data:
        # 1. Check if a record with the same criteria exists
        # print("Hello : ",item)

        if pid == 0 or oid == 0 or  sid == 0:
            pid = item['project_id']
            oid = item['object_id']
            sid = item['segment_id']

        existing_record = SaveRule.objects.filter(
            project_id=item['project_id'],
            object_id=item['object_id'],
            segment_id=item['segment_id'],
            field_id = item['field_id']
        ).first()
        print(existing_record)
    
        item['check_box'] = False
 
 
        now = timezone.now()
        formatted_datetime = now.strftime("%Y-%m-%d %H:%M:%S")  # yyyy-mm-dd hh:mm:ss
        item['last_updated_on'] = formatted_datetime
 
        if existing_record:
            # 2. Update existing record
            serializer = SaveRuleSerializer(existing_record, data=item, partial=True)  # partial=True for partial updates
            # print("Hello Existing : ",serializer)
            if serializer.is_valid():
                serializer.save()
               
               
            else:
                print(serializer.error_messages)
                return "Error"
 
        else:
            # 3. Create a new record
            serializer = SaveRuleSerializer(data=item)
            if serializer.is_valid():
                serializer.save()
               
            else:
                print(serializer.error_messages)
                return "Error"
    

    return "Success"
 

'''
UPDATE {target_table_name}
SET {rule.target_sap_field} = (
    SELECT l.To_Val
    FROM LookUp_table l
    WHERE l.Field_Name = '{rule.target_sap_field}'
    AND l.From_Val = {target_table_name}.{rule.target_sap_field}
    LIMIT 1
)
WHERE EXISTS (
    SELECT 1+
    FROM LookUp_table l
    WHERE l.Field_Name = '{rule.target_sap_field}'
    AND l.From_Val = {target_table_name}.{rule.target_sap_field}
);
''' 

def LocalapplyOneToOne(pid,oid,sid):
    try:
        tar_table=""
        latest_version = Rule.objects.filter(
                    project_id=pid,  
                    object_id=oid,
                    segment_id=sid
                ).order_by('-version_id').first()
        rules = Rule.objects.filter(project_id=pid,  
                    object_id=oid,
                    segment_id=sid,version_id=latest_version.version_id).all()
        segmentForTable = segments.objects.filter(segment_id=sid,project_id=pid,obj_id=oid).first()
        target_table_name = segmentForTable.table_name
        for rule in rules :
            if(rule.source_table !="" and not(table_exists(rule.source_table.upper())) and rule.data_mapping_type != ""):
                for each_table in rule.source_table.upper().split(","):
                    each_table = each_table.strip()
                    jsonPrimary=func(each_table)      
                    create_and_insert_data(each_table,jsonPrimary)
            if(rule.data_mapping_type == "LKT"):
                print("LKT",rule.target_sap_field,target_table_name)
                cursor = connection.cursor()
                cursor.execute(f"""
                               UPDATE {target_table_name}
                                SET {rule.target_sap_field} = (
                                    SELECT l.To_Val
                                    FROM LookUp_table l
                                    WHERE l.Field_Name = '{rule.target_sap_field}'
                                    AND l.From_Val = {target_table_name}.{rule.target_sap_field}
                                    LIMIT 1
                                )
                                WHERE EXISTS (
                                    SELECT 1+
                                    FROM LookUp_table l
                                    WHERE l.Field_Name = '{rule.target_sap_field}'
                                    AND l.From_Val = {target_table_name}.{rule.target_sap_field}
                                );
                                """)
                cursor.commit()
                cursor.close()
        field_mapping={}
        print(latest_version.version_id)
        if(latest_version.version_id==1):
            for rule in rules:
                curr_field = fields.objects.filter(field_id=rule.field_id).first()
                isMandt=curr_field.isKey
                # print("came1")
                if(rule.source_table!="" and rule.source_field_name!="" and isMandt != "False" and rule.data_mapping_type == "1:1"):    
                    print(rule.source_table,rule.source_field_name,target_table_name,rule.target_sap_field)
                    if(not(table_exists(rule.source_table.upper()))):
                        jsonPrimary=func(rule.source_table.upper())      
                        create_and_insert_data(rule.source_table.upper(),jsonPrimary)
                    # print("came2")
                    src_table=rule.source_table.upper()
                    tar_table=target_table_name
                    field_mapping[rule.source_field_name] = rule.target_sap_field
            # print(src_table,tar_table,field_mapping)
            if(src_table!="" and tar_table!=""):
                copy_data_between_tables_with_field_mapping(src_table,tar_table,field_mapping)
                copy_data_between_tables_with_field_mapping(src_table,tar_table+"_src",field_mapping)
                remove_duplicate_rows_group_by_all(tar_table)
                remove_duplicate_rows_group_by_all(tar_table+"_src")
        field_mapping={}  
        print("hii")  
        rules = Rule.objects.filter(project_id=pid,  
                    object_id=oid,
                    segment_id=sid,version_id=1).all()
        for rule in rules :
            print("in cur rule")
            curr_field = fields.objects.filter(field_id=rule.field_id).first()
            isMandt=curr_field.isKey
            print("before")
            print(isMandt)
            if(isMandt != "False"):
                print("inside")
                print(rule.target_sap_field)
                field_mapping[rule.source_field_name] = rule.target_sap_field
        # field mapping contains all the mandetory fields from segment along with their techinal filed name
        rules = Rule.objects.filter(project_id=pid,  
                    object_id=oid,
                    segment_id=sid,version_id=latest_version.version_id).all()
        prev_rules=""
        if(latest_version.version_id!=1):
            prev_rules = Rule.objects.filter(project_id=pid,  
                    object_id=oid,
                    segment_id=sid,version_id=(latest_version.version_id)-1).all()
        print(field_mapping)   
        pkcol1=list(field_mapping.keys())
        pkcol2=list(field_mapping.values())
        table_grouping={}   
        rules_indexing=-1   
        for rule in rules:
            print(table_grouping)
            curr_field = fields.objects.filter(field_id=rule.field_id).first()
            isMandt=curr_field.isKey
            rules_indexing+=1
            prev_rule=""
            if(prev_rules!=""):
                prev_rule=prev_rules[rules_indexing]
            if(rule.source_table!="" and rule.source_field_name!="" and isMandt == "False" and rule.data_mapping_type == "1:1"
               and prev_rule!="" and prev_rule.source_table!=rule.source_table and prev_rule.source_field_name!=rule.source_field_name
               ):
                if(not(table_exists(target_table_name.upper()))):
                    jsonPrimary=func(rule.source_table)  
                    # print(jsonPrimary[0])
                    create_and_insert_data(rule.source_table.upper(),jsonPrimary)
                print(rule.source_table,rule.source_field_name,rule.target_sap_table,rule.target_sap_field)
                print("tablename",segmentForTable.table_name)
                table_grouping[rule.source_table] = table_grouping.get(rule.source_table,{}) 
                table_grouping[rule.source_table][rule.source_field_name]=rule.target_sap_field
            elif(prev_rule=="" and rule.source_table!="" and rule.source_field_name!="" and isMandt == "False" and rule.data_mapping_type == "1:1"
               ):
                if(not(table_exists(target_table_name.upper()))):
                    jsonPrimary=func(rule.source_table)  
                    # print(jsonPrimary[0])
                    create_and_insert_data(rule.source_table.upper(),jsonPrimary)
                print(rule.source_table,rule.source_field_name,rule.target_sap_table,rule.target_sap_field)
                
                print("tablename",segmentForTable.table_name)
                table_grouping[rule.source_table] = table_grouping.get(rule.source_table,{}) 
                table_grouping[rule.source_table][rule.source_field_name]=rule.target_sap_field

        print("hackkkkkk",table_grouping,pkcol1,pkcol2)        
        for table_name in table_grouping.keys(): 
            try:
                update_related_data_with_mapping_and_composite_pks(table_name,segmentForTable.table_name,
                                                    table_grouping[table_name] , " 1 = 1 ",pkcol1,pkcol2)
                update_related_data_with_mapping_and_composite_pks(table_name,segmentForTable.table_name+"_src",
                                                    table_grouping[table_name] , " 1 = 1 ",pkcol1,pkcol2)
            except Exception as e:
                return Response(str(e))
        print("hackkkkkk",table_grouping,pkcol1,pkcol2)     
        latest_version = Rule.objects.filter(
                        project_id=pid,  
                        object_id=oid,
                        segment_id=sid
                    ).order_by('-version_id').first()
        rules = Rule.objects.filter(project_id=pid,
                    object_id=oid,
                    segment_id=sid,version_id=latest_version.version_id).all()
        segmentForTable = segments.objects.filter(segment_id=sid,project_id=pid,obj_id=oid).first()
        target_table_name = segmentForTable.table_name
        for rule in rules:
            print(rule.data_mapping_type,rule.data_mapping_rules)  
            if(rule.data_mapping_type=='Constant'):
                update_column_with_constant(target_table_name,rule.target_sap_field,rule.data_mapping_rules)
                update_column_with_constant(target_table_name+"_src",rule.target_sap_field,rule.data_mapping_rules)
        for rule in rules:
            if(rule.lookup == True):
                try:    
                    print("LKT",rule.target_sap_field,target_table_name)
                    cursor = connection.cursor()
                    cursor.execute(f"""
                                UPDATE {target_table_name}
                                    SET {rule.target_sap_field} = (
                                        SELECT l.From_Val
                                        FROM LookUp_table l
                                        WHERE l.Field_Name = '{rule.target_sap_field}'
                                        AND l.To_Val = {target_table_name}.{rule.target_sap_field}
                                        LIMIT 1
                                    )
                                    WHERE EXISTS (
                                        SELECT 1
                                        FROM LookUp_table l
                                        WHERE l.Field_Name = '{rule.target_sap_field}'
                                        AND l.To_Val = {target_table_name}.{rule.target_sap_field}
                                    );
                                    """)
                    connection.commit()
                    cursor.close()
                except Exception as e:
                    print(e) 
                    
        
        
        
        return "Success"
    except Exception as e:
        print(e)
        return "Error"   



@api_view(['POST'])
def VersionRuleCreate(request):
    print("Hello called Version create")
    data=request.data
    pid = data[0]['project_id']
    oid = data[0]['object_id']
    sid = data[0]['segment_id']
    save_response = LocalSaveRuleCreate(data,pid,oid,sid)
    # if save_response == "Error":
    #     return Response(status=status.HTTP_405_METHOD_NOT_ALLOWED)


    latest_version = Rule.objects.filter(
                project_id=data[0]['project_id'],  # Assuming all items have the same IDs
                object_id=data[0]['object_id'],
                segment_id=data[0]['segment_id']
            ).order_by('-version_id').first()
    next_version_id = 1  # Default if no previous versions
    if latest_version:
        next_version_id = latest_version.version_id + 1
    for item in data:
            item['version_id'] = next_version_id
            item['check_box'] = False
    serializer = RuleSerializer(data=data, many=True)
    print("Hello outside if")
    if serializer.is_valid():
        serializer.save()
        oneToone_response = LocalapplyOneToOne(pid,oid,sid)
        print("one to one response :" ,oneToone_response)
        # if oneToone_response == "Error":
        #     return Response(status=status.HTTP_406_NOT_ACCEPTABLE)
        return Response(serializer.data)
    else:
        return Response(status=status.HTTP_404_NOT_FOUND)
 


def LocalgetTableData(sid):

    try:
        segment = segments.objects.filter(segment_id=sid).first()  # Use .first() to get a single object
        if not segment:  # Handle the case where no segment is found.
            return Response({"error": "Segment not found"}, status=404)
 
        table_name = segment.table_name # Access table_name directly from the segment object
        print(table_name)
        with connections["default"].cursor() as cursor:
            cursor.execute(f"SELECT * FROM {table_name}")
            data = cursor.fetchall()
            cursor.execute(f"SELECT * FROM {table_name} LIMIT 0")
            column_names = [col[0] for col in cursor.description]
 
            # 1. Convert to list of dictionaries (JSON-like)    
            results = []
            for row in data:
                row_dict = dict(zip(column_names, row))  
                results.append(row_dict)
            return results  
 
    except Exception as e:
        print(f"Error fetching data: {e}")
        return Response({"error": str(e)}, status=500)


def table_to_custom_json(table_name):
    # Get column names
    with connection.cursor() as cursor:
        cursor.execute(f"PRAGMA table_info({table_name})")
        columns_info = cursor.fetchall()
    columns = [col[1] for col in columns_info]
    columns.append(columns.pop(0))
    # Fetch all rows into pandas DataFrame
    query = f"SELECT * FROM {table_name}"
    df = pd.read_sql_query(query, connection)

    # Function to determine if column should be displayed
    def should_display(col):
        # Check if any value in the column is neither null nor empty (after stripping spaces)
        return int(df[col].dropna().astype(str).str.strip().replace('', None).notnull().any())

    # Build json_columns as list of {key, value, display}
    json_columns = [{'key': i, 'value': col, 'display': should_display(col)} for i, col in enumerate(columns)]

    # Convert DataFrame rows into dict list
    json_data = json.loads(df.to_json(orient='records'))

    # Prepare return JSON structure
    return_json = {
        "rows": json_data,
        "columns": json_columns
    }
    
    print("hiiiii",json_columns)  # for debug/logging

    return JsonResponse(return_json, safe=False)


@api_view(['GET'])
def getTableData(request,sid):

    print("HII from GetTableData")

    try:
        segment = segments.objects.filter(segment_id=sid).first()  # Use .first() to get a single object
        if not segment:  # Handle the case where no segment is found.
            return Response({"error": "Segment not found"}, status=404)
 
        table_name = segment.table_name # Access table_name directly from the segment object
        print(table_name)
        return table_to_custom_json(table_name)
        # df_valid = get_valid_data_frame(table_name)
        # print(df_valid)
        # json_data = df_valid.to_dict(orient="records")
        # columns=list(df_valid.columns)
        # json_columns = [{'key': i, 'value': col} for i, col in enumerate(columns)]
        # return_json = {
        # "rows": json_data,
        # "columns": json_columns
        # }
        
        # return JsonResponse(return_json,safe=False)
        # print(table_name)
        # with connections["default"].cursor() as cursor:
        #     cursor.execute(f"SELECT * FROM {table_name}")
        #     data = cursor.fetchall()
        #     cursor.execute(f"SELECT * FROM {table_name} LIMIT 0")
        #     column_names = [col[0] for col in cursor.description]
 
        #     # 1. Convert to list of dictionaries (JSON-like)    
        #     results = []
        #     for row in data:
        #         row_dict = dict(zip(column_names, row))  
        #         results.append(row_dict)
        #     return Response(results)  
 
    except Exception as e:
        print(f"Error fetching data: {e}")
        return Response({"error": str(e)}, status=500)
 

@api_view(['POST'])
def execute_selection_criteria(request,pid,oid,sid):
    print("Execute Queries Called...")
    selection_criteria = request.data['prompt']
    segmentForTable = segments.objects.filter(segment_id=sid,project_id=pid,obj_id=oid).first()
    target_table_name = segmentForTable.table_name
    print(selection_criteria," ",pid," ",oid," ",sid)
    try:
        system = DMTool()
        print("HIIIIII")
        # query = "Check Materials which you have got from Transaofmration rule 1 In MARA_500 table and IF matching Entries found, then bring Unit of Measure   field from MARA_500 table to the Target Table ELSE, If no entries found in MARA_500, then check ROH  Material  ( found in Transformation 2 ) in MARA_700 Table and bring the Unit of Measure ELSE, If no entries found in MARA_700, then bring the Unit of measure from MARA table"
        # query = "Bring PRODUCT with MTART = ROH from MARA Table"
        obj = objects.objects.get(obj_id=oid)
        sess_id = obj.session_id
        result, affected_indexes = system.process_selection_criteria(
            selection_criteria,
            object_id=oid,
            segment_id=sid,
            project_id=pid,
            session_id = sess_id
        )
        copy_rows_with_reason_and_timestamp(target_table_name, 
                                    target_table_name+"_val", 
                                    affected_indexes,
                                    f"selection cretria : {selection_criteria}")
        print("hiiii",result)
        print(f"rows effected indexes : {affected_indexes}")
        return Response("Demo")
    except Exception as e:
        connection.rollback() #rollback the changes if an error occurs.
        print(e)
        return Response({"message": f"Error executing query: {str(e)}"}, status=500)



@api_view(['GET'])
def upload_Bussiness_rules(request):
    # file = request.FILES['file']
    file=r"C:\Users\gajananda.mallidi\Downloads\NEWPUSH\NEWPUSH\LLM_Backend\connection\view_rules.xlsx"
    df = pd.read_excel(file)
    df=df.to_dict(orient='records')
    return JsonResponse(df,safe=False)
    





@api_view(['POST'])
def execute_queries(request,pid,oid,sid):
    print("Execute Queries Called...")
    prompt = request.data['prompt']
    print(prompt," ",pid," ",oid," ",sid)
    try:
       
        system = DMTool()
        print("HIIIIII")
        # query = "Check Materials which you have got from Transaofmration rule 1 In MARA_500 table and IF matching Entries found, then bring Unit of Measure   field from MARA_500 table to the Target Table ELSE, If no entries found in MARA_500, then check ROH  Material  ( found in Transformation 2 ) in MARA_700 Table and bring the Unit of Measure ELSE, If no entries found in MARA_700, then bring the Unit of measure from MARA table"
        # query = "Bring PRODUCT with MTART = ROH from MARA Table"
        segmentForTable = segments.objects.filter(segment_id=sid,project_id=pid,obj_id=oid).first()
        target_table_name = segmentForTable.table_name
        obj = objects.objects.get(obj_id=oid)
        sess_id = obj.session_id
        print(prompt,
            oid,
            sid,
            pid,
            sess_id)
        result, affected_indexes = system.process_sequential_query(
            prompt,
            object_id=oid,
            segment_id=sid,
            project_id=pid,
            session_id = sess_id
        )
        copy_rows_with_reason_and_timestamp(target_table_name, 
                                    target_table_name+"_val", 
                                    affected_indexes,
                                    prompt)
        print("hiiii",result)
        print(f"rows effected indexes : {affected_indexes}")
        return Response("Demo")
    except Exception as e:
        connection.rollback() #rollback the changes if an error occurs.
        print(e)
        return Response({"message": f"Error executing query: {str(e)}"}, status=500)

    






    # LLM_migration.project_id = pid
    # LLM_migration.object_id = oid
    # LLM_migration.segment_id = sid
    # LLM_migration.user_prompt_from_backend = prompt

    # LLM_migration.main()
    # table_data = LocalgetTableData(sid)
    # return Response(table_data)


@api_view(['GET'])
def delete_table_data(request):

    table_name ="t_24_Product_Plant_Data_Ext"

    try:
        with connection.cursor() as cursor:
                
            sql = f"DELETE FROM {table_name};"

            cursor.execute(sql)
            print(f"Data deleted from table '{table_name}' successfully.")
            return Response("Data deleted")
            # return Response("Success")

    except sqlite3.Error as e:
        print(f"An error occurred: {e}")
        # return Response("Error")

def update_related_data_with_mapping_and_composite_pks(table1_name, table2_name, field_mapping, condition1, pk_columns1, pk_columns2):
    """
    Updates table2 rows by setting columns based on values from table1 using correlated subqueries in SQLite.
    Only rows matching composite primary keys and condition1 in table1 are updated.
    """
    try:
        # Compose SET clause with correlated subqueries for each mapped field
        set_clauses = []
        for t1_field, t2_field in field_mapping.items():
            # Build the correlated subquery using composite keys
            # WHERE clause to match all composite key parts between table1 and table2
            join_conditions = []
            for pk1, pk2 in zip(pk_columns1, pk_columns2):
                join_conditions.append(f"{table1_name}.{pk1} = {table2_name}.{pk2}")
            join_condition_str = " AND ".join(join_conditions)

            # Add optional condition1 from table1
            if condition1:
                full_condition = f"{join_condition_str} AND ({condition1})"
            else:
                full_condition = join_condition_str

            set_clause = (
                f"{t2_field} = (SELECT {t1_field} FROM {table1_name} WHERE {full_condition} LIMIT 1)"
            )
            set_clauses.append(set_clause)

        set_clause_str = ", ".join(set_clauses)

        # Compose WHERE EXISTS to update only rows having matching row(s) in table1 with condition1
        join_conditions_for_exists = []
        for pk1, pk2 in zip(pk_columns1, pk_columns2):
            join_conditions_for_exists.append(f"{table1_name}.{pk1} = {table2_name}.{pk2}")
        join_condition_exists_str = " AND ".join(join_conditions_for_exists)
        if condition1:
            exists_condition = f"EXISTS (SELECT 1 FROM {table1_name} WHERE {join_condition_exists_str} AND ({condition1}))"
        else:
            exists_condition = f"EXISTS (SELECT 1 FROM {table1_name} WHERE {join_condition_exists_str})"

        sql = f"""
        UPDATE {table2_name}
        SET {set_clause_str}
        WHERE {exists_condition}
        """

        print(sql)  # For debugging

        with connections["default"].cursor() as cursor:
            cursor.execute(sql)
            connections["default"].commit()
            print(f"Successfully updated {cursor.rowcount} rows in {table2_name}.")

    except Exception as e:
        connections["default"].rollback()
        print(f"Error updating related data: {e}")



# def update_related_data_with_mapping_and_composite_pks(table1_name, table2_name, field_mapping, condition1, pk_columns1, pk_columns2):
#     print(table1_name, table2_name, field_mapping, condition1, pk_columns1, pk_columns2)
#     try:
#         with connections["default"].cursor() as cursor:
#             # 1. Get data from table1 along with composite primary key
#             select_columns = ", ".join(pk_columns1 + list(field_mapping.keys()))  # Combine PKs and fields
#             select_sql = f"SELECT {select_columns} FROM {table1_name}"
#             cursor.execute(select_sql)
#             rows = cursor.fetchall()
#             # print(len(rows))
#             if not rows:
#                 print(f"No matching data found in {table1_name} for condition: {condition1}")
#                 return

#             # 2. Update table2 for each row fetched from table1
#             for row in rows:
#                 # print(row)
#                 pk_values = row[:len(pk_columns1)]  # Extract primary key values (tuple)
#                 values_to_update = row[len(pk_columns1):]  # Extract values to update
#                 # print("primary keys",pk_values)
#                 # print("values to update",values_to_update)
#                 set_clause_parts = []
#                 for i, table1_field in enumerate(field_mapping.keys()):
#                     table2_field = field_mapping[table1_field]
#                     set_clause_parts.append(f"{table2_field} = %s")
#                 set_clause = ", ".join(set_clause_parts)
#                 pk_condition_parts = []
#                 for i, pk_col in enumerate(pk_columns2):
#                     pk_condition_parts.append(f"{pk_col} = %s")
#                 condition2 = " AND ".join(pk_condition_parts)
#                 update_sql = f"UPDATE {table2_name} SET {set_clause} WHERE {condition2}"
#                 print(update_sql)
#                 cursor.execute(update_sql, list(values_to_update) + list(pk_values))  
#             rows_updated = cursor.rowcount  
#             # print(rows_updated)
#             connections["default"].commit()
#             print(f"Successfully updated data in {table2_name} based on {table1_name}")

#     except Exception as e:
#         connections["default"].rollback()
#         print(f"Error updating related data: {e}")


def remove_duplicate_rows_group_by_all(table_name):
    try:
        with connections["default"].cursor() as cursor:
            # 1. Get all column names
            cursor.execute(f"PRAGMA table_info({table_name})") #sqlite command
            columns = [row[1] for row in cursor.fetchall()] #get the column names

            if not columns:
                print(f"Table '{table_name}' not found or has no columns.")
                return

            # 2. Construct the SQL query
            columns_str = ", ".join(columns)
            query = f"""
                DELETE FROM {table_name}
                WHERE ROWID NOT IN (
                    SELECT MIN(ROWID)
                    FROM {table_name}
                    GROUP BY {columns_str}
                );
            """

            # 3. Execute the query
            cursor.execute(query)
            rows_deleted = cursor.rowcount
            connections["default"].commit()
            print(f"Removed {rows_deleted} duplicate rows from '{table_name}'.")

    except Exception as e:
        connections["default"].rollback()
        print(f"Error removing duplicates: {e}")



def copy_data_between_tables_with_field_mapping(table1_name, table2_name, field_mapping):
    try:
        with connections["default"].cursor() as cursor:
            # 1. Construct the INSERT and SELECT queries dynamically
            select_clause = ", ".join(field_mapping.keys())  # Select from table1
            insert_columns = ", ".join(field_mapping.values())  # Insert into table2

            insert_sql = f"INSERT INTO {table2_name} ({insert_columns}) SELECT {select_clause} FROM {table1_name}"

            # 2. Execute the query
            cursor.execute(insert_sql)

            rows_copied = cursor.rowcount
            connections["default"].commit()
            print(f"Successfully copied {rows_copied} rows from {table1_name} to {table2_name}")

    except Exception as e:
        connections["default"].rollback()
        print(f"Error copying data: {e}")



def create_and_insert_data(table_name, data):
    if not data:
        print("No data provided. Table creation and insertion skipped.")
        return
    try:
        with connections["default"].cursor() as cursor:
            # 1. Create Table (Simplified - Same structure assumed)
            first_row = data[0]
            drop_table_dynamically(table_name=table_name)
            create_table_sql = f"CREATE TABLE IF NOT EXISTS {table_name} ("

            columns = []
            for key, value in first_row.items():  # Iterate through the first row only
                column_name = key
                column_type = "TEXT"  # Default type (you can refine this based on value types)
                columns.append(f'"{column_name}" {column_type}')

            create_table_sql += ", ".join(columns) + ")"
            cursor.execute(create_table_sql)


            # 2. Insert Data (Same as before)
            insert_sql = f"INSERT INTO {table_name} (" + ", ".join(f'"{key}"' for key in first_row) + ") VALUES (" + ", ".join(["%s"] * len(first_row)) + ")"

            for row in data:
                values = list(row.values())
                cursor.execute(insert_sql, values)

            connections["default"].commit()
            print(f"Table '{table_name}' created and data inserted successfully.")

    except Exception as e:
        connections["default"].rollback()
        print(f"Error creating/inserting data: {e}")



def func(table_name):
  table_name = table_name.upper()
  class RFC_ERROR_INFO(Structure):
      _fields_ = [("code", c_long),
                  ("group", c_long),
                  ("key", c_wchar * 128),
                  ("message", c_wchar * 512),
                  ("abapMsgClass", c_wchar * 21),
                  ("abapMsgType", c_wchar * 2),
                  ("abapMsgNumber", c_wchar * 4),
                  ("abapMsgV1", c_wchar * 51),
                  ("abapMsgV2", c_wchar * 51),
                  ("abapMsgV3", c_wchar * 51),
                  ("abapMsgV4", c_wchar * 51)]

  class RFC_CONNECTION_PARAMETER(Structure):
      _fields_ = [("name", c_wchar_p),
                  ("value", c_wchar_p)]


  #-Constants-------------------------------------------------------------

  #-RFC_RC - RFC return codes---------------------------------------------
  RFC_OK = 0
  RFC_COMMUNICATION_FAILURE = 1
  RFC_LOGON_FAILURE = 2
  RFC_ABAP_RUNTIME_FAILURE = 3
  RFC_ABAP_MESSAGE = 4
  RFC_ABAP_EXCEPTION = 5
  RFC_CLOSED = 6
  RFC_CANCELED = 7
  RFC_TIMEOUT = 8
  RFC_MEMORY_INSUFFICIENT = 9
  RFC_VERSION_MISMATCH = 10
  RFC_INVALID_PROTOCOL = 11
  RFC_SERIALIZATION_FAILURE = 12
  RFC_INVALID_HANDLE = 13
  RFC_RETRY = 14
  RFC_EXTERNAL_FAILURE = 15
  RFC_EXECUTED = 16
  RFC_NOT_FOUND = 17
  RFC_NOT_SUPPORTED = 18
  RFC_ILLEGAL_STATE = 19
  RFC_INVALID_PARAMETER = 20
  RFC_CODEPAGE_CONVERSION_FAILURE = 21
  RFC_CONVERSION_FAILURE = 22
  RFC_BUFFER_TOO_SMALL = 23
  RFC_TABLE_MOVE_BOF = 24
  RFC_TABLE_MOVE_EOF = 25
  RFC_START_SAPGUI_FAILURE = 26
  RFC_ABAP_CLASS_EXCEPTION = 27
  RFC_UNKNOWN_ERROR = 28
  RFC_AUTHORIZATION_FAILURE = 29

  #-RFCTYPE - RFC data types----------------------------------------------
  RFCTYPE_CHAR = 0
  RFCTYPE_DATE = 1
  RFCTYPE_BCD = 2
  RFCTYPE_TIME = 3
  RFCTYPE_BYTE = 4
  RFCTYPE_TABLE = 5
  RFCTYPE_NUM = 6
  RFCTYPE_FLOAT = 7
  RFCTYPE_INT = 8
  RFCTYPE_INT2 = 9
  RFCTYPE_INT1 = 10
  RFCTYPE_NULL = 14
  RFCTYPE_ABAPOBJECT = 16
  RFCTYPE_STRUCTURE = 17
  RFCTYPE_DECF16 = 23
  RFCTYPE_DECF34 = 24
  RFCTYPE_XMLDATA = 28
  RFCTYPE_STRING = 29
  RFCTYPE_XSTRING = 30
  RFCTYPE_BOX = 31
  RFCTYPE_GENERIC_BOX = 32

  #-RFC_UNIT_STATE - Processing status of a background unit---------------
  RFC_UNIT_NOT_FOUND = 0 
  RFC_UNIT_IN_PROCESS = 1 
  RFC_UNIT_COMMITTED = 2 
  RFC_UNIT_ROLLED_BACK = 3 
  RFC_UNIT_CONFIRMED = 4 

  #-RFC_CALL_TYPE - Type of an incoming function call---------------------
  RFC_SYNCHRONOUS = 0 
  RFC_TRANSACTIONAL = 1 
  RFC_QUEUED = 2 
  RFC_BACKGROUND_UNIT = 3 

  #-RFC_DIRECTION - Direction of a function module parameter--------------
  RFC_IMPORT = 1 
  RFC_EXPORT = 2 
  RFC_CHANGING = RFC_IMPORT + RFC_EXPORT 
  RFC_TABLES = 4 + RFC_CHANGING 

  #-RFC_CLASS_ATTRIBUTE_TYPE - Type of an ABAP object attribute-----------
  RFC_CLASS_ATTRIBUTE_INSTANCE = 0 
  RFC_CLASS_ATTRIBUTE_CLASS = 1 
  RFC_CLASS_ATTRIBUTE_CONSTANT = 2 

  #-RFC_METADATA_OBJ_TYPE - Ingroup repository----------------------------
  RFC_METADATA_FUNCTION = 0 
  RFC_METADATA_TYPE = 1 
  RFC_METADATA_CLASS = 2 


  #-Variables-------------------------------------------------------------
  ErrInf = RFC_ERROR_INFO; RfcErrInf = ErrInf()
  ConnParams = RFC_CONNECTION_PARAMETER * 5; RfcConnParams = ConnParams()
  SConParams = RFC_CONNECTION_PARAMETER * 3; RfcSConParams = SConParams()


  #-Library---------------------------------------------------------------
  # if str(platform.architecture()[0]) == "32bit":
  #   os.environ['PATH'] += ";C:\\SAPRFCSDK\\32bit"
  #   SAPNWRFC = "C:\\SAPRFCSDK\\32bit\\sapnwrfc.dll"
  # elif str(platform.architecture()[0]) == "64bit":
  #   os.environ['PATH'] += ";C:\\SAPRFCSDK\\64bit"
  #   SAPNWRFC = "C:\\SAPRFCSDK\\64bit\\sapnwrfc.dll"

  SAPNWRFC = "sapnwrfc.dll"
  SAP = windll.LoadLibrary(SAPNWRFC)

  #-Prototypes------------------------------------------------------------
  SAP.RfcAppendNewRow.argtypes = [c_void_p, POINTER(ErrInf)]
  SAP.RfcAppendNewRow.restype = c_void_p

  SAP.RfcCreateTable.argtypes = [c_void_p, POINTER(ErrInf)]
  SAP.RfcCreateTable.restype = c_void_p

  SAP.RfcCloseConnection.argtypes = [c_void_p, POINTER(ErrInf)]
  SAP.RfcCloseConnection.restype = c_ulong

  SAP.RfcCreateFunction.argtypes = [c_void_p, POINTER(ErrInf)]
  SAP.RfcCreateFunction.restype = c_void_p

  SAP.RfcCreateFunctionDesc.argtypes = [c_wchar_p, POINTER(ErrInf)]
  SAP.RfcCreateFunctionDesc.restype = c_void_p

  SAP.RfcDestroyFunction.argtypes = [c_void_p, POINTER(ErrInf)]
  SAP.RfcDestroyFunction.restype = c_ulong

  SAP.RfcDestroyFunctionDesc.argtypes = [c_void_p, POINTER(ErrInf)]
  SAP.RfcDestroyFunctionDesc.restype = c_ulong

  SAP.RfcGetChars.argtypes = [c_void_p, c_wchar_p, c_void_p, c_ulong, \
    POINTER(ErrInf)]
  SAP.RfcGetChars.restype = c_ulong

  SAP.RfcGetCurrentRow.argtypes = [c_void_p, POINTER(ErrInf)]
  SAP.RfcGetCurrentRow.restype = c_void_p

  SAP.RfcGetFunctionDesc.argtypes = [c_void_p, c_wchar_p, POINTER(ErrInf)]
  SAP.RfcGetFunctionDesc.restype = c_void_p

  SAP.RfcGetRowCount.argtypes = [c_void_p, POINTER(c_ulong), \
    POINTER(ErrInf)]
  SAP.RfcGetRowCount.restype = c_ulong

  SAP.RfcGetStructure.argtypes = [c_void_p, c_wchar_p, \
    POINTER(c_void_p), POINTER(ErrInf)]
  SAP.RfcGetStructure.restype = c_ulong

  SAP.RfcGetTable.argtypes = [c_void_p, c_wchar_p, POINTER(c_void_p), \
    POINTER(ErrInf)]
  SAP.RfcGetTable.restype = c_ulong

  SAP.RfcGetVersion.argtypes = [POINTER(c_ulong), POINTER(c_ulong), \
    POINTER(c_ulong)]
  SAP.RfcGetVersion.restype = c_wchar_p

  SAP.RfcInstallServerFunction.argtypes = [c_wchar_p, c_void_p, \
    c_void_p, POINTER(ErrInf)]
  SAP.RfcInstallServerFunction.restype = c_ulong

  SAP.RfcInvoke.argtypes = [c_void_p, c_void_p, POINTER(ErrInf)]
  SAP.RfcInvoke.restype = c_ulong

  SAP.RfcListenAndDispatch.argtypes = [c_void_p, c_ulong, POINTER(ErrInf)]
  SAP.RfcListenAndDispatch.restype = c_ulong

  SAP.RfcMoveToFirstRow.argtypes = [c_void_p, POINTER(ErrInf)]
  SAP.RfcMoveToFirstRow.restype = c_ulong

  SAP.RfcMoveToNextRow.argtypes = [c_void_p, POINTER(ErrInf)]
  SAP.RfcMoveToNextRow.restype = c_ulong

  SAP.RfcOpenConnection.argtypes = [POINTER(ConnParams), c_ulong, \
    POINTER(ErrInf)]
  SAP.RfcOpenConnection.restype = c_void_p

  SAP.RfcPing.argtypes = [c_void_p, POINTER(ErrInf)]
  SAP.RfcPing.restype = c_ulong

  SAP.RfcRegisterServer.argtypes = [POINTER(SConParams), c_ulong, \
    POINTER(ErrInf)]
  SAP.RfcRegisterServer.restype = c_void_p

  SAP.RfcSetChars.argtypes = [c_void_p, c_wchar_p, c_wchar_p, c_ulong, \
    POINTER(ErrInf)]
  SAP.RfcSetChars.restype = c_ulong
  def join_json_objects_multiple_keys(obj1, obj2, primary_keys):
      result = []

      # Create a dictionary to efficiently look up items in obj2 by combined primary keys
      obj2_lookup = {}
      for item2 in obj2:
          key = tuple(item2[key] for key in primary_keys)  # Create a tuple key
          obj2_lookup[key] = item2

      for item1 in obj1:
          key = tuple(item1[key] for key in primary_keys)
          item2 = obj2_lookup.get(key)  # Efficient lookup

          if item2:
              merged_object = {**item1, **item2}
              result.append(merged_object)
          else:
              result.append(item1)  # Keep item1 if no match
              print(f"No match found for {key}")

      return result


  #-Main------------------------------------------------------------------

  RfcConnParams[0].name = "ASHOST"; RfcConnParams[0].value = "34.194.191.113"
  RfcConnParams[1].name = "SYSNR" ; RfcConnParams[1].value = "01"
  RfcConnParams[2].name = "CLIENT"; RfcConnParams[2].value = "100"
  RfcConnParams[3].name = "USER"  ; RfcConnParams[3].value = "GAJANANDAM"
  RfcConnParams[4].name = "PASSWD"; RfcConnParams[4].value = "SunR!se@6789"

  TableName = table_name
  keyFields = []
  cnt = 0

  hRFC = SAP.RfcOpenConnection(RfcConnParams, 5, RfcErrInf)
  if hRFC != None:

    charBuffer = create_unicode_buffer(1048576 + 1)

    hFuncDesc = SAP.RfcGetFunctionDesc(hRFC, "CACS_GET_TABLE_FIELD450", RfcErrInf)
    if hFuncDesc != 0:
      hFunc = SAP.RfcCreateFunction(hFuncDesc, RfcErrInf)
      if hFunc != 0:

        rc = SAP.RfcSetChars(hFunc, "I_TABNAME", TableName, \
          len(TableName), RfcErrInf)
        print(SAP.RfcInvoke(hRFC, hFunc, RfcErrInf) == RFC_OK)
        if SAP.RfcInvoke(hRFC, hFunc, RfcErrInf) == RFC_OK:

          hTable = c_void_p(0)
          if SAP.RfcGetTable(hFunc, "T_KEYFIELD", hTable, RfcErrInf) == RFC_OK:
            RowCount = c_ulong(0)
            rc = SAP.RfcGetRowCount(hTable, RowCount, RfcErrInf)
            print(RowCount, 1)
            rc = SAP.RfcMoveToFirstRow(hTable, RfcErrInf)
            for i in range(0, RowCount.value):
              hRow = SAP.RfcGetCurrentRow(hTable, RfcErrInf)
              rc = SAP.RfcGetChars(hRow, "FIELDNAME", charBuffer, 512, RfcErrInf)
              # print(str(charBuffer.value), end="    ")
              fieldName = str(charBuffer.value)
              # rc = SAP.RfcGetChars(hRow, "LENGTH", charBuffer, 512, RfcErrInf)
              # val = int(charBuffer.value)
              # if (sum + val < 512):
              #   sum += val
              #   l1.append(fieldName.strip())
              #   # print(sum)
              # else:
              keyFields.append(fieldName.strip())
                # l1 = [fieldName.strip()]
                # sum = val
              if i < RowCount.value:
                rc = SAP.RfcMoveToNextRow(hTable, RfcErrInf)

        rc = SAP.RfcDestroyFunction(hFunc, RfcErrInf)

    # rc = SAP.RfcCloseConnection(hRFC, RfcErrInf)

    print(*keyFields)

    keyFieldsCnt = len(keyFields)
    print(keyFieldsCnt)
  else:
    print(RfcErrInf.key)
    print(RfcErrInf.message)


  ind, keyDict = 0, {}

  # hRFC = SAP.RfcOpenConnection(RfcConnParams, 5, RfcErrInf)
  if hRFC != None:

    charBuffer = create_unicode_buffer(1048576 + 1)

    hFuncDesc = SAP.RfcGetFunctionDesc(hRFC, "Z450RFC_READ_TABLE", RfcErrInf)
    if hFuncDesc != 0:
      hFunc = SAP.RfcCreateFunction(hFuncDesc, RfcErrInf)
      if hFunc != 0:

        rc = SAP.RfcSetChars(hFunc, "QUERY_TABLE", TableName, \
          len(TableName), RfcErrInf)
        rc = SAP.RfcSetChars(hFunc, "DELIMITER", "~", 1, RfcErrInf)
        if SAP.RfcInvoke(hRFC, hFunc, RfcErrInf) == RFC_OK:

          hTable = c_void_p(0)
          if SAP.RfcGetTable(hFunc, "FIELDS", hTable, RfcErrInf) == RFC_OK:
            
            
            sum, l, l1 = 0, [], keyFields.copy()
            keyFieldsLen = 0
            RowCount = c_ulong(0)
            rc = SAP.RfcGetRowCount(hTable, RowCount, RfcErrInf)
            print(RowCount)
            rc = SAP.RfcMoveToFirstRow(hTable, RfcErrInf)
            for i in range(0, RowCount.value):
              hRow = SAP.RfcGetCurrentRow(hTable, RfcErrInf)
              rc = SAP.RfcGetChars(hRow, "FIELDNAME", charBuffer, 512, RfcErrInf)
              # print(str(charBuffer.value), end="    ")
              fieldName = str(charBuffer.value)
              rc = SAP.RfcGetChars(hRow, "LENGTH", charBuffer, 512, RfcErrInf)
              val = int(charBuffer.value)
              cnt += 1
              # print(fieldName.strip(), cnt)
              if (i < keyFieldsCnt):
                print(i)
                i += 1
                keyFieldsLen += val
              else:
                if (sum + val + keyFieldsLen < 400):
                  sum += val
                  l1.append(fieldName.strip())
                  # print(sum)
                else:
                  l.append(l1)
                  l1 = keyFields.copy()
                  l1.append(fieldName.strip())
                  # print(sum + keyFieldsLen)
                  sum = val
                  
              if i < RowCount.value:
                rc = SAP.RfcMoveToNextRow(hTable, RfcErrInf)
            l.append(l1)
        rc = SAP.RfcDestroyFunction(hFunc, RfcErrInf)

    # rc = SAP.RfcCloseConnection(hRFC, RfcErrInf)

    # print(l)
  else:
    print(RfcErrInf.key)
    print(RfcErrInf.message)

  # for i in l:
  #   print(i[:2])

  length = 0
  for ii in l:
    for jj in ii:
      if (jj == 'MANDT' or jj == 'MATNR'): continue
      length += 1
  print(length)

  jsonTemp = jsonPrimary = []
  for splittedFields in l:
    # hRFC = SAP.RfcOpenConnection(RfcConnParams, 5, RfcErrInf)
    if hRFC != None:

      charBuffer = create_unicode_buffer(1048576 + 1)

      hFuncDesc = SAP.RfcGetFunctionDesc(hRFC, "Z450RFC_READ_TAB_DATA", RfcErrInf)
      if hFuncDesc != 0:
        hFunc = SAP.RfcCreateFunction(hFuncDesc, RfcErrInf)
        if hFunc != 0:

          rc = SAP.RfcSetChars(hFunc, "QUERY_TABLE", TableName, \
            len(TableName), RfcErrInf)
          rc = SAP.RfcSetChars(hFunc, "DELIMITER", "~", 1, RfcErrInf)

          #MATNR,MTART,ATTYP,SATNR,MATKL,MBRSH,MEINS,SPART,BISMT,DATAB,LIQDT,NORMT,GROES,LABOR,BRGEW,NTGEW,GEWEI,LAENG,BREIT,HOEHE,MEABM,VOLUM,VOLEH,KZKFG,IPRKZ,RAUBE,TEMPB,BEHVO,STOFF,ETIAR,ETIFO,WESCH,XGCHP,MHDHB,MHDRZ,SLED_BBD

          field = ','.join(splittedFields)
          # print(field)
          rc = SAP.RfcSetChars(hFunc, "FIELDNAME", field, len(field), RfcErrInf)

          # print(SAP.RfcInvoke(hRFC, hFunc, RfcErrInf) == RFC_OK)
          if SAP.RfcInvoke(hRFC, hFunc, RfcErrInf) == RFC_OK:

            hTable = c_void_p(0)
            if SAP.RfcGetTable(hFunc, "DATA", hTable, RfcErrInf) == RFC_OK:

              RowCount = c_ulong(0)
              rc = SAP.RfcGetRowCount(hTable, RowCount, RfcErrInf)
              rc = SAP.RfcMoveToFirstRow(hTable, RfcErrInf)
              for i in range(0, RowCount.value):
                hRow = SAP.RfcGetCurrentRow(hTable, RfcErrInf)
                rc = SAP.RfcGetChars(hRow, "WA", charBuffer, 1024, RfcErrInf)
                data_row=charBuffer.value
                # data_dict = {field: value for field, value in zip(splittedFields, data_row)}
                # print(data_dict)

                data_row = charBuffer.value.split("~")
  
                                  # Create dictionary using only requested fields
                              # data_dict = {field: value for field, value in zip(field, data_row)}
                              # # print(charBuffer.value)
                              # res.append(data_dict)
                fields = field.split(",")
                data_dict = {f: v.strip() for f, v in zip(fields, data_row)}
                jsonTemp.append(data_dict)

                if i < RowCount.value:
                  rc = SAP.RfcMoveToNextRow(hTable, RfcErrInf)

          rc = SAP.RfcDestroyFunction(hFunc, RfcErrInf)
      # print(jsonTemp)
      if (jsonPrimary == []): 
        jsonPrimary = jsonTemp
      else:
        jsonPrimary = join_json_objects_multiple_keys(jsonPrimary, jsonTemp, keyFields)
      jsonTemp = []
      # rc = SAP.RfcCloseConnection(hRFC, RfcErrInf)

    else:
      print(RfcErrInf.key)
      print(RfcErrInf.message)


  del SAP

  # for d in jsonPrimary:
  #   for j in d:
  #     d[j] = d[j].strip()

  return jsonPrimary


def table_exists(table_name):
    try:
        with connections["default"].cursor() as cursor:
            # SQLite specific query to check for table existence
            cursor.execute(f"SELECT * FROM {table_name}")
            return True

    except Exception as e:
        print(f"Error checking table existence: {e}")
        return False  


@api_view(['GET'])
def getChat(request,pid,oid,sid):
    chats = Chat.objects.filter(project_id=pid,object_id=oid,segment_id=sid).all()
    print(chats)
    chats = ChatSerializer(data=chats,many=True)
    chats.is_valid()
    return Response(chats.data)
   
 
@api_view(['POST'])
def CreateChat(request):
    data=request.data
    serializer_data = ChatSerializer(data=data)
    if(serializer_data.is_valid()):
        serializer_data.save()
        return Response("success")
    return Response("fail")


def update_column_with_constant(table_name, column_name, constant_value):
    try:
        with connection.cursor() as cursor:
            # Use parameterized query to prevent SQL injection
            cursor.execute(f"UPDATE {table_name} SET {column_name} = %s", [constant_value])
            print(f"Successfully updated {cursor.rowcount} rows in {table_name}.{column_name}")
 
    except Exception as e:
        print(f"Error updating column: {e}")


@api_view(['GET'])
def applyOneToOne(request,pid,oid,sid):
    try:
        tar_table=""
        latest_version = Rule.objects.filter(
                    project_id=pid,  
                    object_id=oid,
                    segment_id=sid
                ).order_by('-version_id').first()
        rules = Rule.objects.filter(project_id=pid,  
                    object_id=oid,
                    segment_id=sid,version_id=latest_version.version_id).all()
        field_mapping={}
        segmentForTable = segments.objects.filter(segment_id=sid,project_id=pid,obj_id=oid).first()
        target_table_name = segmentForTable.table_name
        print(latest_version.version_id)
        if(latest_version.version_id==1):
            for rule in rules:
                curr_field = fields.objects.filter(field_id=rule.field_id).first()
                isMandt=curr_field.isKey
                # print("came1")
                if(rule.source_table!="" and rule.source_field_name!="" and isMandt != "False"):    
                    print(rule.source_table,rule.source_field_name,target_table_name,rule.target_sap_field)
                    if(not(table_exists(rule.source_table.upper()))):
                        jsonPrimary=func(rule.source_table.upper())      
                        create_and_insert_data(rule.source_table.upper(),jsonPrimary)
                    # print("came2")
                    src_table=rule.source_table.upper()
                    tar_table=target_table_name
                    field_mapping[rule.source_field_name] = rule.target_sap_field
            # print(src_table,tar_table,field_mapping)
            if(src_table!="" and tar_table!=""):
                copy_data_between_tables_with_field_mapping(src_table,tar_table,field_mapping)
                remove_duplicate_rows_group_by_all(tar_table)
           
 
        field_mapping={}  
        print("hii")  
        rules = Rule.objects.filter(project_id=pid,  
                    object_id=oid,
                    segment_id=sid,version_id=1).all()
        for rule in rules :
            print("in cur rule")
            curr_field = fields.objects.filter(field_id=rule.field_id).first()
            isMandt=curr_field.isKey
            print("before")
            print(isMandt)
            if(isMandt != "False"):
                print("inside")
                print(rule.target_sap_field)
                field_mapping[rule.source_field_name] = rule.target_sap_field
        # field mapping contains all the mandetory fields from segment along with their techinal filed name
        rules = Rule.objects.filter(project_id=pid,  
                    object_id=oid,
                    segment_id=sid,version_id=latest_version.version_id).all()
        print(field_mapping)  
        pkcol1=list(field_mapping.keys())
        pkcol2=list(field_mapping.values()) 
        table_grouping = {}     
        for rule in rules:
            curr_field = fields.objects.filter(field_id=rule.field_id).first()
            isMandt=curr_field.isKey
            if(rule.source_table!="" and rule.source_field_name!="" and isMandt == "False"):
                if(not(table_exists(target_table_name.upper()))):
                    jsonPrimary=func(rule.source_table)  
                    print(jsonPrimary[0])
                    create_and_insert_data(rule.source_table.upper(),jsonPrimary)
                print(rule.source_table,rule.source_field_name,rule.target_sap_table,rule.target_sap_field)
                
                print("tablename",segmentForTable.table_name)
                table_grouping[rule.source_table] = table_grouping.get(rule.source_table,{}) 
                table_grouping[rule.source_table][rule.source_field_name]=rule.target_sap_field
        print("hackkkkkk",table_grouping,pkcol1,pkcol2)        
        for table_name in table_grouping.keys(): 
            try:
                update_related_data_with_mapping_and_composite_pks(table_name,segmentForTable.table_name,
                                                    table_grouping[table_name] , " 1 = 1 ",pkcol1,pkcol2)
            except Exception as e:
                return Response(str(e))
    except Exception as e:
        print(e)   
    latest_version = Rule.objects.filter(
                    project_id=pid,  
                    object_id=oid,
                    segment_id=sid
                ).order_by('-version_id').first()
    rules = Rule.objects.filter(project_id=pid,
                object_id=oid,
                segment_id=sid,version_id=latest_version.version_id).all()
    segmentForTable = segments.objects.filter(segment_id=sid,project_id=pid,obj_id=oid).first()
    target_table_name = segmentForTable.table_name
    for rule in rules:
        print(rule.data_mapping_type,rule.data_mapping_rules)  
        if(rule.data_mapping_type=='Constant'):
            update_column_with_constant(target_table_name,rule.target_sap_field,rule.data_mapping_rules)
    return Response("Hii")


@api_view(['GET']) 
def getLatestVersion(request,pid,oid,sid):
    print("came to latestVersion")
    latest_version = Rule.objects.filter(
                project_id=pid, 
                object_id=oid,
                segment_id=sid
            ).order_by('-version_id').first()
    if(not(latest_version)):
        return Response([])
    versiondata = Rule.objects.filter(
                project_id=pid,  
                object_id=oid,
                segment_id=sid,
        version_id = latest_version.version_id).all()
    versiondata = RuleSerializer(versiondata,many=True)
    temp=[]
    for i in versiondata.data:
        if i['data_mapping_type'] != "":
            temp.append(i)
    print("rrrrrrrrttttttt",versiondata)        
    return Response(temp)



@api_view(['POST'])
def saveSuccessFactors(request):
    project_id = request.data['project_id']
    template_name = request.data['template_name']
    template_name = TableName_Modification(template_name)
    # project_id = 12
    # template_name = "PersonalInfo"
    file = request.FILES['file']
   
    obj_name = f"{template_name}_object"
 
    check = objects.objects.filter(project_id=project_id,obj_name=obj_name)
    if check:
        return Response(status=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE)
 
 
 
    data = pd.read_csv(file)
    column = list(data.columns)
    col = []
    columns = []
    for c in column:
        col.append(c)
        col.append("TEXT")
        columns.append(col)
        col = []
    # print(*columns, sep = ', ')
    data = pd.DataFrame(data)
 
    df = pd.DataFrame(data)
    print(df)
 
    seg_name = f"{template_name}_segment"
    
    project_name = ""
    project_type = ""
    if Project.objects.filter(project_id = project_id):
        prj = Project.objects.get(project_id = project_id)
        project_name = prj.project_name
        project_type = prj.project_type

    obj_data = {
        "obj_name" : obj_name,
        "project_id" : project_id,
        "template_name" : template_name,
        "project_name" : project_name,
        "project_type" : project_type
    }
 
    # print(obj_data)
 
    # print("Heloooooooooooooo")
    obj = ObjectSerializer(data=obj_data)
 
   
    if objects.objects.filter(project_id=obj_data['project_id'],obj_name = obj_data['obj_name']):
        return Response(status=status.HTTP_406_NOT_ACCEPTABLE)
 
    if obj.is_valid():
        obj_instance=obj.save()
        objid = obj_instance.obj_id
 
        # seg =TableName_Modification(seg_name)
        # objName =TableName_Modification(obj_name)
        tab ="t"+"_"+str(project_id)+"_"+str(obj_name)+"_"+str(seg_name)
        seg = seg_name
        seg_obj = {
            "project_id" : project_id,
            "obj_id" : objid,
            "segement_name":seg,
            "table_name" : tab
        }
        seg_instance = SegementSerializer(data=seg_obj)
        if seg_instance.is_valid():
            seg_id_get = seg_instance.save()
            segment_id = seg_id_get.segment_id
 
        else:
            return Response("Error at first segement creation")
       
 
        create_table(tab,columns)
 
        for d in columns:
            field_obj = {
                "project_id" : project_id,
                "obj_id" : objid,
                "segement_id" : segment_id,
                "sap_structure" : "",
                "fields" : d[0],
                "description" : "",
                "isMandatory" : "True",
                "isKey" : ""
            }
            field_instance = FieldSerializer(data=field_obj)
            if field_instance.is_valid():
                field_id_get = field_instance.save()
                field_id = field_id_get.field_id
            else:
                return Response("Error at Field Creation")
 
        insert_data_from_dataframe(df,tab)
 
        return Response(obj.data)
    else:
        return Response(status=status.HTTP_409_CONFLICT) 




@api_view(['PUT'])
def reUploadSuccessFactors(request,oid):
    template_name = request.data['template_name']
    # project_id = 12
    # template_name = "Personal Info"
    template_name = TableName_Modification(template_name)
    file = request.FILES['file']
 
    obj_name = f"{template_name}_object"
 
    obj = objects.objects.filter(obj_id = oid)
 
 
    if obj:
        obj = objects.objects.get(obj_id = oid)
        serializer = ObjectSerializer(obj)
        project_id = serializer.data['project_id']
        if obj.obj_name != obj_name :
            print("Hii inside if")
            check = objects.objects.filter(project_id=project_id,obj_name=obj_name)
            if check:
                return Response(status=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE)
       
        seg_delete = segments.objects.filter(project_id=project_id,obj_id = oid)
        if seg_delete:
            seg_del = segments.objects.get(project_id=project_id,obj_id = oid)
            deleteSqlLiteTable(seg_del.table_name)      
            if seg_del:
                serializer = SegementSerializer(seg_del)
                seg_del.delete()
 
       
 
        obj_data = {
            "obj_name" : obj_name,
            "project_id" : project_id,
            "template_name" : template_name
        }
       
        data = ObjectSerializer(instance=obj, data=obj_data)
        if data.is_valid():
            obj_instance=data.save()
            objid = obj_instance.obj_id
           
 
            data = pd.read_csv(file)
            column = list(data.columns)
            col = []
            columns = []
            for c in column:
                col.append(c)
                col.append("TEXT")
                columns.append(col)
                col = []
            # print(*columns, sep = ', ')
            data = pd.DataFrame(data)
 
            df = pd.DataFrame(data)
            print(df)
 
            seg_name = f"{template_name}_segment"
 
 
 
            # seg =TableName_Modification(seg_name)
            # objName =TableName_Modification(obj_name)
            tab ="t"+"_"+str(project_id)+"_"+str(obj_name)+"_"+str(seg_name)
            seg = seg_name
            seg_obj = {
                "project_id" : project_id,
                "obj_id" : objid,
                "segement_name":seg,
                "table_name" : tab
            }
            seg_instance = SegementSerializer(data=seg_obj)
            if seg_instance.is_valid():
                seg_id_get = seg_instance.save()
                segment_id = seg_id_get.segment_id
 
            else:
                return Response("Error at first segement creation")
           
 
            create_table(tab,columns)
            # create_table(tab+"_src",columns)
 
            for d in columns:
                field_obj = {
                    "project_id" : project_id,
                    "obj_id" : objid,
                    "segement_id" : segment_id,
                    "sap_structure" : "",
                    "fields" : d[0],
                    "description" : "",
                    "isMandatory" : "True",
                    "isKey" : ""
                }
                field_instance = FieldSerializer(data=field_obj)
                if field_instance.is_valid():
                    field_id_get = field_instance.save()
                    field_id = field_id_get.field_id
                else:
                    return Response("Error at Field Creation")
 
            insert_data_from_dataframe(df,tab)
 
            return Response("Done")
       
        else:
            return Response(status=status.HTTP_406_NOT_ACCEPTABLE)
 
    else:
        return Response(status=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE)



def join_json_objects_multiple_keys(obj1, obj2, primary_keys):
    result = []
 
    # Create a dictionary to efficiently look up items in obj2 by combined primary keys
    obj2_lookup = {}
    for item2 in obj2:
        key = tuple(item2[key] for key in primary_keys)  # Create a tuple key
        obj2_lookup[key] = item2
 
    for item1 in obj1:
        key = tuple(item1[key] for key in primary_keys)
        item2 = obj2_lookup.get(key)  # Efficient lookup
 
        if item2:
            merged_object = {**item1, **item2}
            result.append(merged_object)
        else:
            result.append(item1)  # Keep item1 if no match
            print(f"No match found for {key}")
 
    return result


@api_view(['GET'])
def getSapTableData(request):
    #-Begin-----------------------------------------------------------------
 
#-Packages--------------------------------------------------------------
 
 
    #-Constants-------------------------------------------------------------
 
    #-RFC_RC - RFC return codes---------------------------------------------
    class RFC_ERROR_INFO(Structure):
        _fields_ = [("code", c_long),
                    ("group", c_long),
                    ("key", c_wchar * 128),
                    ("message", c_wchar * 512),
                    ("abapMsgClass", c_wchar * 21),
                    ("abapMsgType", c_wchar * 2),
                    ("abapMsgNumber", c_wchar * 4),
                    ("abapMsgV1", c_wchar * 51),
                    ("abapMsgV2", c_wchar * 51),
                    ("abapMsgV3", c_wchar * 51),
                    ("abapMsgV4", c_wchar * 51)]
    class RFC_CONNECTION_PARAMETER(Structure):
        _fields_ = [("name", c_wchar_p),
                    ("value", c_wchar_p)]
    RFC_OK = 0
    RFC_COMMUNICATION_FAILURE = 1
    RFC_LOGON_FAILURE = 2
    RFC_ABAP_RUNTIME_FAILURE = 3
    RFC_ABAP_MESSAGE = 4
    RFC_ABAP_EXCEPTION = 5
    RFC_CLOSED = 6
    RFC_CANCELED = 7
    RFC_TIMEOUT = 8
    RFC_MEMORY_INSUFFICIENT = 9
    RFC_VERSION_MISMATCH = 10
    RFC_INVALID_PROTOCOL = 11
    RFC_SERIALIZATION_FAILURE = 12
    RFC_INVALID_HANDLE = 13
    RFC_RETRY = 14
    RFC_EXTERNAL_FAILURE = 15
    RFC_EXECUTED = 16
    RFC_NOT_FOUND = 17
    RFC_NOT_SUPPORTED = 18
    RFC_ILLEGAL_STATE = 19
    RFC_INVALID_PARAMETER = 20
    RFC_CODEPAGE_CONVERSION_FAILURE = 21
    RFC_CONVERSION_FAILURE = 22
    RFC_BUFFER_TOO_SMALL = 23
    RFC_TABLE_MOVE_BOF = 24
    RFC_TABLE_MOVE_EOF = 25
    RFC_START_SAPGUI_FAILURE = 26
    RFC_ABAP_CLASS_EXCEPTION = 27
    RFC_UNKNOWN_ERROR = 28
    RFC_AUTHORIZATION_FAILURE = 29
 
    #-RFCTYPE - RFC data types----------------------------------------------
    RFCTYPE_CHAR = 0
    RFCTYPE_DATE = 1
    RFCTYPE_BCD = 2
    RFCTYPE_TIME = 3
    RFCTYPE_BYTE = 4
    RFCTYPE_TABLE = 5
    RFCTYPE_NUM = 6
    RFCTYPE_FLOAT = 7
    RFCTYPE_INT = 8
    RFCTYPE_INT2 = 9
    RFCTYPE_INT1 = 10
    RFCTYPE_NULL = 14
    RFCTYPE_ABAPOBJECT = 16
    RFCTYPE_STRUCTURE = 17
    RFCTYPE_DECF16 = 23
    RFCTYPE_DECF34 = 24
    RFCTYPE_XMLDATA = 28
    RFCTYPE_STRING = 29
    RFCTYPE_XSTRING = 30
    RFCTYPE_BOX = 31
    RFCTYPE_GENERIC_BOX = 32
 
    #-RFC_UNIT_STATE - Processing status of a background unit---------------
    RFC_UNIT_NOT_FOUND = 0
    RFC_UNIT_IN_PROCESS = 1
    RFC_UNIT_COMMITTED = 2
    RFC_UNIT_ROLLED_BACK = 3
    RFC_UNIT_CONFIRMED = 4
 
    #-RFC_CALL_TYPE - Type of an incoming function call---------------------
    RFC_SYNCHRONOUS = 0
    RFC_TRANSACTIONAL = 1
    RFC_QUEUED = 2
    RFC_BACKGROUND_UNIT = 3
 
    #-RFC_DIRECTION - Direction of a function module parameter--------------
    RFC_IMPORT = 1
    RFC_EXPORT = 2
    RFC_CHANGING = RFC_IMPORT + RFC_EXPORT
    RFC_TABLES = 4 + RFC_CHANGING
 
    #-RFC_CLASS_ATTRIBUTE_TYPE - Type of an ABAP object attribute-----------
    RFC_CLASS_ATTRIBUTE_INSTANCE = 0
    RFC_CLASS_ATTRIBUTE_CLASS = 1
    RFC_CLASS_ATTRIBUTE_CONSTANT = 2
 
    #-RFC_METADATA_OBJ_TYPE - Ingroup repository----------------------------
    RFC_METADATA_FUNCTION = 0
    RFC_METADATA_TYPE = 1
    RFC_METADATA_CLASS = 2
 
 
    #-Variables-------------------------------------------------------------
    ErrInf = RFC_ERROR_INFO; RfcErrInf = ErrInf()
    ConnParams = RFC_CONNECTION_PARAMETER * 5; RfcConnParams = ConnParams()
    SConParams = RFC_CONNECTION_PARAMETER * 3; RfcSConParams = SConParams()
 
 
    #-Library---------------------------------------------------------------
    # if str(platform.architecture()[0]) == "32bit":
    #   os.environ['PATH'] += ";C:\\SAPRFCSDK\\32bit"
    #   SAPNWRFC = "C:\\SAPRFCSDK\\32bit\\sapnwrfc.dll"
    # elif str(platform.architecture()[0]) == "64bit":
    #   os.environ['PATH'] += ";C:\\SAPRFCSDK\\64bit"
    #   SAPNWRFC = "C:\\SAPRFCSDK\\64bit\\sapnwrfc.dll"
 
    SAPNWRFC = "sapnwrfc.dll"
    SAP = windll.LoadLibrary(SAPNWRFC)
 
    #-Prototypes------------------------------------------------------------
    SAP.RfcAppendNewRow.argtypes = [c_void_p, POINTER(ErrInf)]
    SAP.RfcAppendNewRow.restype = c_void_p
 
    SAP.RfcCreateTable.argtypes = [c_void_p, POINTER(ErrInf)]
    SAP.RfcCreateTable.restype = c_void_p
 
    SAP.RfcCloseConnection.argtypes = [c_void_p, POINTER(ErrInf)]
    SAP.RfcCloseConnection.restype = c_ulong
 
    SAP.RfcCreateFunction.argtypes = [c_void_p, POINTER(ErrInf)]
    SAP.RfcCreateFunction.restype = c_void_p
 
    SAP.RfcCreateFunctionDesc.argtypes = [c_wchar_p, POINTER(ErrInf)]
    SAP.RfcCreateFunctionDesc.restype = c_void_p
 
    SAP.RfcDestroyFunction.argtypes = [c_void_p, POINTER(ErrInf)]
    SAP.RfcDestroyFunction.restype = c_ulong
 
    SAP.RfcDestroyFunctionDesc.argtypes = [c_void_p, POINTER(ErrInf)]
    SAP.RfcDestroyFunctionDesc.restype = c_ulong
 
    SAP.RfcGetChars.argtypes = [c_void_p, c_wchar_p, c_void_p, c_ulong, \
    POINTER(ErrInf)]
    SAP.RfcGetChars.restype = c_ulong
 
    SAP.RfcGetCurrentRow.argtypes = [c_void_p, POINTER(ErrInf)]
    SAP.RfcGetCurrentRow.restype = c_void_p
 
    SAP.RfcGetFunctionDesc.argtypes = [c_void_p, c_wchar_p, POINTER(ErrInf)]
    SAP.RfcGetFunctionDesc.restype = c_void_p
 
    SAP.RfcGetRowCount.argtypes = [c_void_p, POINTER(c_ulong), \
    POINTER(ErrInf)]
    SAP.RfcGetRowCount.restype = c_ulong
 
    SAP.RfcGetStructure.argtypes = [c_void_p, c_wchar_p, \
    POINTER(c_void_p), POINTER(ErrInf)]
    SAP.RfcGetStructure.restype = c_ulong
 
    SAP.RfcGetTable.argtypes = [c_void_p, c_wchar_p, POINTER(c_void_p), \
    POINTER(ErrInf)]
    SAP.RfcGetTable.restype = c_ulong
 
    SAP.RfcGetVersion.argtypes = [POINTER(c_ulong), POINTER(c_ulong), \
    POINTER(c_ulong)]
    SAP.RfcGetVersion.restype = c_wchar_p
 
    SAP.RfcInstallServerFunction.argtypes = [c_wchar_p, c_void_p, \
    c_void_p, POINTER(ErrInf)]
    SAP.RfcInstallServerFunction.restype = c_ulong
 
    SAP.RfcInvoke.argtypes = [c_void_p, c_void_p, POINTER(ErrInf)]
    SAP.RfcInvoke.restype = c_ulong
 
    SAP.RfcListenAndDispatch.argtypes = [c_void_p, c_ulong, POINTER(ErrInf)]
    SAP.RfcListenAndDispatch.restype = c_ulong
 
    SAP.RfcMoveToFirstRow.argtypes = [c_void_p, POINTER(ErrInf)]
    SAP.RfcMoveToFirstRow.restype = c_ulong
 
    SAP.RfcMoveToNextRow.argtypes = [c_void_p, POINTER(ErrInf)]
    SAP.RfcMoveToNextRow.restype = c_ulong
 
    SAP.RfcOpenConnection.argtypes = [POINTER(ConnParams), c_ulong, \
    POINTER(ErrInf)]
    SAP.RfcOpenConnection.restype = c_void_p
 
    SAP.RfcPing.argtypes = [c_void_p, POINTER(ErrInf)]
    SAP.RfcPing.restype = c_ulong
 
    SAP.RfcRegisterServer.argtypes = [POINTER(SConParams), c_ulong, \
    POINTER(ErrInf)]
    SAP.RfcRegisterServer.restype = c_void_p
 
    SAP.RfcSetChars.argtypes = [c_void_p, c_wchar_p, c_wchar_p, c_ulong, \
    POINTER(ErrInf)]
    SAP.RfcSetChars.restype = c_ulong
   
 
    #-Main------------------------------------------------------------------
 
    RfcConnParams[0].name = "ASHOST"; RfcConnParams[0].value = "34.194.191.113"
    RfcConnParams[1].name = "SYSNR" ; RfcConnParams[1].value = "01"
    RfcConnParams[2].name = "CLIENT"; RfcConnParams[2].value = "100"
    RfcConnParams[3].name = "USER"  ; RfcConnParams[3].value = "GAJANANDAM"
    RfcConnParams[4].name = "PASSWD"; RfcConnParams[4].value = "SunR!se@6789"
 
    TableName = "ADR6"
    keyFields = []
    cnt = 0
    l=[]
    hRFC = SAP.RfcOpenConnection(RfcConnParams, 5, RfcErrInf)
    if hRFC != None:
 
      charBuffer = create_unicode_buffer(1048576 + 1)
 
      hFuncDesc = SAP.RfcGetFunctionDesc(hRFC, "CACS_GET_TABLE_FIELD450", RfcErrInf)
      if hFuncDesc != 0:
        hFunc = SAP.RfcCreateFunction(hFuncDesc, RfcErrInf)
        if hFunc != 0:
 
          rc = SAP.RfcSetChars(hFunc, "I_TABNAME", TableName, \
            len(TableName), RfcErrInf)
          print(SAP.RfcInvoke(hRFC, hFunc, RfcErrInf) == RFC_OK)
          if SAP.RfcInvoke(hRFC, hFunc, RfcErrInf) == RFC_OK:
 
            hTable = c_void_p(0)
            if SAP.RfcGetTable(hFunc, "T_KEYFIELD", hTable, RfcErrInf) == RFC_OK:
              RowCount = c_ulong(0)
              rc = SAP.RfcGetRowCount(hTable, RowCount, RfcErrInf)
              print(RowCount, 1)
              rc = SAP.RfcMoveToFirstRow(hTable, RfcErrInf)
              for i in range(0, RowCount.value):
                hRow = SAP.RfcGetCurrentRow(hTable, RfcErrInf)
                rc = SAP.RfcGetChars(hRow, "FIELDNAME", charBuffer, 512, RfcErrInf)
                # print(str(charBuffer.value), end="    ")
                fieldName = str(charBuffer.value)
                # rc = SAP.RfcGetChars(hRow, "LENGTH", charBuffer, 512, RfcErrInf)
                # val = int(charBuffer.value)
                # if (sum + val < 512):
                #   sum += val
                #   l1.append(fieldName.strip())
                #   # print(sum)
                # else:
                keyFields.append(fieldName.strip())
                  # l1 = [fieldName.strip()]
                  # sum = val
                if i < RowCount.value:
                  rc = SAP.RfcMoveToNextRow(hTable, RfcErrInf)
 
          rc = SAP.RfcDestroyFunction(hFunc, RfcErrInf)
 
      # rc = SAP.RfcCloseConnection(hRFC, RfcErrInf)
 
      print(*keyFields)
      keyFieldsCnt = len(keyFields)
      print(keyFieldsCnt)
    else:
      print(RfcErrInf.key)
      print(RfcErrInf.message)
 
 
    ind, keyDict = 0, {}
 
    # hRFC = SAP.RfcOpenConnection(RfcConnParams, 5, RfcErrInf)
    if hRFC != None:
 
      charBuffer = create_unicode_buffer(1048576 + 1)
 
      hFuncDesc = SAP.RfcGetFunctionDesc(hRFC, "Z450RFC_READ_TABLE", RfcErrInf)
      if hFuncDesc != 0:
        hFunc = SAP.RfcCreateFunction(hFuncDesc, RfcErrInf)
        if hFunc != 0:
 
          rc = SAP.RfcSetChars(hFunc, "QUERY_TABLE", TableName, \
            len(TableName), RfcErrInf)
          rc = SAP.RfcSetChars(hFunc, "DELIMITER", "~", 1, RfcErrInf)
          if SAP.RfcInvoke(hRFC, hFunc, RfcErrInf) == RFC_OK:
 
            hTable = c_void_p(0)
            if SAP.RfcGetTable(hFunc, "FIELDS", hTable, RfcErrInf) == RFC_OK:
             
             
              sum, l, l1 = 0, [], keyFields.copy()
              keyFieldsLen = 0
              RowCount = c_ulong(0)
              rc = SAP.RfcGetRowCount(hTable, RowCount, RfcErrInf)
              print(RowCount)
              rc = SAP.RfcMoveToFirstRow(hTable, RfcErrInf)
              for i in range(0, RowCount.value):
                hRow = SAP.RfcGetCurrentRow(hTable, RfcErrInf)
                rc = SAP.RfcGetChars(hRow, "FIELDNAME", charBuffer, 512, RfcErrInf)
                # print(str(charBuffer.value), end="    ")
                fieldName = str(charBuffer.value)
                rc = SAP.RfcGetChars(hRow, "LENGTH", charBuffer, 512, RfcErrInf)
                val = int(charBuffer.value)
                cnt += 1
                # print(fieldName.strip(), cnt)
                if (i < keyFieldsCnt):
                  print(i)
                  i += 1
                  keyFieldsLen += val
                else:
                  if (sum + val + keyFieldsLen < 400):
                    sum += val
                    l1.append(fieldName.strip())
                    # print(sum)
                  else:
                    l.append(l1)
                    l1 = keyFields.copy()
                    l1.append(fieldName.strip())
                    # print(sum + keyFieldsLen)
                    sum = val
                   
                if i < RowCount.value:
                  rc = SAP.RfcMoveToNextRow(hTable, RfcErrInf)
              l.append(l1)
          rc = SAP.RfcDestroyFunction(hFunc, RfcErrInf)
 
      # rc = SAP.RfcCloseConnection(hRFC, RfcErrInf)
 
      # print(l)
    else:
      print(RfcErrInf.key)
      print(RfcErrInf.message)
 
    # for i in l:
    #   print(i[:2])
 
    length = 0
    for ii in l:
      for jj in ii:
        if (jj == 'USERNAME'):
            length += 1
    print(l)
    print("Total Number of  : ",length)
 
    jsonTemp =[]
    jsonPrimary = []
    for splittedFields in l:
      # hRFC = SAP.RfcOpenConnection(RfcConnParams, 5, RfcErrInf)
      if hRFC != None:
 
        charBuffer = create_unicode_buffer(1048576 + 1)
 
        hFuncDesc = SAP.RfcGetFunctionDesc(hRFC, "Z450RFC_READ_TAB_DATA", RfcErrInf)
        if hFuncDesc != 0:
          hFunc = SAP.RfcCreateFunction(hFuncDesc, RfcErrInf)
          if hFunc != 0:
 
            rc = SAP.RfcSetChars(hFunc, "QUERY_TABLE", TableName, \
              len(TableName), RfcErrInf)
            rc = SAP.RfcSetChars(hFunc, "DELIMITER", "~", 1, RfcErrInf)
 
            #MATNR,MTART,ATTYP,SATNR,MATKL,MBRSH,MEINS,SPART,BISMT,DATAB,LIQDT,NORMT,GROES,LABOR,BRGEW,NTGEW,GEWEI,LAENG,BREIT,HOEHE,MEABM,VOLUM,VOLEH,KZKFG,IPRKZ,RAUBE,TEMPB,BEHVO,STOFF,ETIAR,ETIFO,WESCH,XGCHP,MHDHB,MHDRZ,SLED_BBD
 
            field = ','.join(splittedFields)
            # print(field)
            rc = SAP.RfcSetChars(hFunc, "FIELDNAME", field, len(field), RfcErrInf)
 
            # print(SAP.RfcInvoke(hRFC, hFunc, RfcErrInf) == RFC_OK)
            if SAP.RfcInvoke(hRFC, hFunc, RfcErrInf) == RFC_OK:
 
              hTable = c_void_p(0)
              if SAP.RfcGetTable(hFunc, "DATA", hTable, RfcErrInf) == RFC_OK:
 
                RowCount = c_ulong(0)
                rc = SAP.RfcGetRowCount(hTable, RowCount, RfcErrInf)
                rc = SAP.RfcMoveToFirstRow(hTable, RfcErrInf)
                for i in range(0, RowCount.value):
                  hRow = SAP.RfcGetCurrentRow(hTable, RfcErrInf)
                  rc = SAP.RfcGetChars(hRow, "WA", charBuffer, 1024, RfcErrInf)
                  data_row=charBuffer.value
                  # data_dict = {field: value for field, value in zip(splittedFields, data_row)}
                  # print(data_dict)
 
                  data_row = charBuffer.value.split("~")
                #   print(data_row)
   
                                    # Create dictionary using only requested fields
                                # data_dict = {field: value for field, value in zip(field, data_row)}
                                # # print(charBuffer.value)
                                # res.append(data_dict)
                  fields = field.split(",")
                  data_dict = {f: v.strip() for f, v in zip(fields, data_row)}
                  jsonTemp.append(data_dict)
                #   print(jsonTemp)
                #   print("Hiieloooooooooooooooooooooooooooooooooooooo")
 
                  if i < RowCount.value:
                    rc = SAP.RfcMoveToNextRow(hTable, RfcErrInf)
 
            rc = SAP.RfcDestroyFunction(hFunc, RfcErrInf)
        print(len(jsonPrimary))
        if (len(jsonPrimary) == 0):
        #   print("HEYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYYY")
          jsonPrimary = jsonTemp
        else:
        #   print(jsonPrimary)
        #   print("YASHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHHH")
          jsonPrimary = join_json_objects_multiple_keys(jsonPrimary, jsonTemp, keyFields)
        jsonTemp = []
        # rc = SAP.RfcCloseConnection(hRFC, RfcErrInf)
   
 
      else:
        print(RfcErrInf.key)
        print(RfcErrInf.message)
 
 
 
    df = pd.DataFrame(jsonPrimary)
    # print(df)
    new_column_list = df.columns.tolist()
    # print(column_list)
 
    # Create a new list to store the modified column names
    column_list = []
    for cl in new_column_list:
        if cl == "PRIMARY":
            column_list.append("PRIMARY1")
        else:
            column_list.append(cl)
 
    df.columns = column_list
    columns = []
    col =[]
 
    for cl in column_list:
        col.append(cl)
        col.append("TEXT")
        columns.append(col)
        col=[]
    create_table(TableName, columns)
    insert_data_from_dataframe(df,TableName)
    # print("Final JSON : ",jsonPrimary)
    return Response(jsonPrimary)
 
 
    del SAP
 
    # for d in jsonPrimary:
    #   for j in d:
    #     d[j] = d[j].strip()
 
    # print(jsonPrimary[-1])
 


# def getTableAndRuleData(request,pid,oid,sid):

#     connections = fields.objects.filter(project_id=pid,obj_id=oid,segement_id=sid)
#     if connections:
#         serializer = FieldSerializer(connections,many=True)
#         fields_data = serializer.data 
    


import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('SVG')

from io import BytesIO
 
class GetPlot(APIView):
    proj, obj, seg, table_name = "", "", "", ""
    tempGraph = []
   
    def plot(self, l, d):
        plt.figure(figsize=(12, 8))
        missing_values = d.isnull().sum()
        # print("dont know: \n", missing_values.sum())
        if (not missing_values.sum()):
            return 'No Nulls'
        column_descriptions = l
       
        field_names = [col.strip() for col in missing_values.index]
        field_names = [col.replace("_", " ") for col in field_names]
        # print(column_descriptions)
        # print(field_names)
        described_names = [f"{name} ({column_descriptions.get(name, '')})" for name in field_names]
        described_names, field_names = [], []
        for _ in l:
            field_names.append(_)
            described_names.append(l[_])
        # print('sjfakda')
       
        ax = plt.bar(described_names, missing_values.values)
        print('hi')
        plt.title('Missing Values in MARA Data')
        plt.xticks(rotation=45, ha='right')
        plt.ylabel('Count of Missing Values')
        plt.tight_layout()
   
        for rect in ax.patches:
            height = rect.get_height()
            plt.annotate(f'{int(height)}',
                        xy=(rect.get_x() + rect.get_width() / 2, height),
                        xytext=(0, 3),
                        textcoords="offset points",
                        ha='center', va='bottom')
           
        buf = BytesIO()
        plt.savefig(buf, format='png')
        buf.seek(0)  # rewind the buffer to the beginning
        plt.close()  # close the plot to free memory
        return buf.getvalue().decode('latin1')  # encode for
 
    def mandt(self):
 
        mandDic, mandFields = {}, []
       
        conn = SaveRule.objects.filter(project_id=self.proj,object_id=self.obj,segment_id=self.seg, isMandatory=True)
        if conn:
            serializer = SaveRuleSerializer(conn,many=True)
            for i in serializer.data:
                mandDic[i['target_sap_field']] = i['text_description']
                mandFields.append(i['target_sap_field'])
 
        try:
            with connections["default"].cursor() as cursor:
                mandt = ', '.join(mandFields)
                # print("hii")
                # print(mandt)
                data = cursor.execute(f"SELECT {mandt} FROM {self.table_name}")
                data = data.fetchall()
                data = pd.DataFrame(data, columns=mandFields)
                return self.plot(mandDic, data)
           
        except Exception as e:
            print(f"{e}")
       
        return 0
 
    def save_plot_to_bytes(self):
        """Save the current plot to a BytesIO object and return it."""
        buf = BytesIO()
        plt.savefig(buf, format='png')
        buf.seek(0)  # rewind the buffer to the beginning
        plt.close()  # close the plot to free memory
        return buf.getvalue().decode('latin1')  # encode for
   
    def groupPlots(self, name, desc, d):
        if (len(d[name].value_counts()) > 0):
            top_creators = d[name].value_counts().head(10)
    
            # colors = plt.cm.viridis(np.linspace(0, 1, len(top_creators)))
            colors = plt.cm.get_cmap('Paired', len(top_creators))
            # top_creators = d[x].value_counts().head(10)
        
            plt.figure(figsize=(10, 6))  # Create a new figure
            ax = top_creators.plot(kind='bar', color=colors(range(len(top_creators))))
    
            for p in ax.patches:
                ax.annotate(f'{int(p.get_height())}',
                            (p.get_x() + p.get_width() / 2., p.get_height()),
                            ha='center', va='center',
                            fontsize=12, color='black',
                            xytext=(0, 5),  # Vertical offset
                            textcoords='offset points')
    
            plt.title(f'{desc}', fontsize=14)
            plt.ylabel('Count', fontsize=12)
            plt.xticks(rotation=45, ha='right') #rotate x axis labels
            plt.grid(axis='y', linestyle='--', alpha=0.7)
            # plt.show()
            plt.tight_layout() #prevents labels from being cut off
            self.tempGraph.append({name : [desc, self.save_plot_to_bytes()]})
            plt.close()
       
    def profMand(self):
        mandDic, mandFields = {}, []
        self.tempGraph = []
        conn = SaveRule.objects.filter(project_id=self.proj,object_id=self.obj,segment_id=self.seg, isMandatory=True)
        if conn:
            serializer = SaveRuleSerializer(conn,many=True)
            for i in serializer.data:
                mandDic[i['target_sap_field']] = i['text_description']
                mandFields.append(i['target_sap_field'])
        # print(lookFields)
 
        #mandFields -> columnNames, mandDic -> dictonary for description
        try:
            with connections["default"].cursor() as cursor:
                mandt = ', '.join(mandFields)
                # print("hii")
                data = cursor.execute(f"SELECT {mandt} FROM {self.table_name}")
                data = data.fetchall()
                data = pd.DataFrame(data, columns=mandFields)
                # print("Hi hello" ,data)
                for i in mandDic:
                    self.groupPlots(i, mandDic[i], data)
                return {'profMand': self.tempGraph}
       
        except Exception as e:
            print(f"{e}")
       
        return 0
 
    def profLook(self):
        lookUpDic = []  
        self.tempGraph = []      
        conn = SaveRule.objects.filter(project_id=self.proj,object_id=self.obj,segment_id=self.seg)
        if conn:
            serializer = SaveRuleSerializer(conn,many=True)
            for i in serializer.data:
                if (i['lookup_table']):
                    lookUpDic.append({i['target_sap_field']: i['text_description']})
 
        lookFields = []
        for i in lookUpDic:
            for j in i:
                lookFields.append(j)
        try:
            with connections["default"].cursor() as cursor:
                mandLook = ', '.join(lookFields)
                data = cursor.execute(f"SELECT {mandLook} FROM {self.table_name}")
                data = data.fetchall()
                data = pd.DataFrame(data, columns=lookFields)
                for i in lookUpDic:
                    for j in i:
                        self.groupPlots(j, i[j], data)    
            return {'profLook': self.tempGraph}
 
        except Exception as e:
            print(f"{e}")
       
        return 0
       
    def get(self, request, pid, oid, sid):
        graphs = []
        self.proj, self.obj, self.seg = pid, oid, sid
        conn = segments.objects.filter(project_id=pid,obj_id=oid,segment_id=sid).first()
        self.table_name = conn.table_name
        graphs.append({'mandatory': self.mandt()})
        graphs.append(self.profMand())
        graphs.append(self.profLook())
 
        return JsonResponse({'plot': graphs})  
 
 
class GetExactGraph(APIView):
    def get(self, request, pid, oid, sid, fname):
        conn = segments.objects.filter(project_id=pid,obj_id=oid,segment_id=sid).first()
        table_name = conn.table_name
       
        desc = ""
        conn = SaveRule.objects.filter(project_id=pid,object_id=oid,segment_id=sid, target_sap_field = fname)
        if conn:
            serializer = SaveRuleSerializer(conn,many=True)
            for i in serializer.data:
                    desc =  i['text_description']
        # print(table_name, fname)
        try:
            with connections["default"].cursor() as cursor:
                # print("hii")
                data = cursor.execute(f"SELECT {fname} FROM {table_name}")
                data = data.fetchall()
                data = pd.DataFrame(data, columns=[fname])
 
            top_creators = data[fname].value_counts().head(20)
            plt.figure(figsize=(12, 6))
            colors = plt.cm.get_cmap('Paired', len(top_creators))
 
            fig, ax = plt.subplots(figsize=(10, 6))  # Create a single figure and axes
 
            ax = top_creators.plot(kind='bar', ax=ax, legend=False, color=colors(range(len(top_creators))))
 
            for p in ax.patches:
                ax.annotate(f'{int(p.get_height())}',
                            (p.get_x() + p.get_width() / 2., p.get_height()),
                            ha='center', va='center',
                            fontsize=12, color='black',
                            xytext=(0, 5),  # Vertical offset
                            textcoords='offset points')
 
            ax.set_title(f'{desc}', fontsize=14)
            ax.set_ylabel('Count', fontsize=12)
            # ax.tick_params(axis='both', which='major', labelsize=10)
            ax.grid(axis='y', linestyle='--', alpha=0.7)
 
            buf = BytesIO()
            plt.savefig(buf, format='png')
            buf.seek(0)  # rewind the buffer to the beginning
            plt.close()  # close the plot to free memory
            return HttpResponse(buf.getvalue().decode('latin1'))  # encode for
 
        except Exception as e:
            return HttpResponse(f"{e}")
 
 
# class GetPlot(APIView):
#     proj, obj, seg, table_name = "", "", "", ""
#     tempGraph = []
   
#     def plot(self, l, d):
#         plt.figure(figsize=(12, 8))
#         missing_values = d.isnull().sum()
#         if (not missing_values.sum()):
#             return 'No Nulls'
#         column_descriptions = l
       
#         field_names = [col.strip() for col in missing_values.index]
#         field_names = [col.replace("_", " ") for col in field_names]
#         described_names = []
#         for _ in l:
#             described_names.append(l[_])
       
#         ax = plt.bar(described_names, missing_values.values)
#         plt.title('Missing Values in MARA Data')
#         plt.xticks(rotation=45, ha='right')
#         plt.ylabel('Count of Missing Values')
#         plt.tight_layout()
   
#         for rect in ax.patches:
#             height = rect.get_height()
#             plt.annotate(f'{int(height)}',
#                         xy=(rect.get_x() + rect.get_width() / 2, height),
#                         xytext=(0, 3),
#                         textcoords="offset points",
#                         ha='center', va='bottom')
           
#         buf = BytesIO()
#         plt.savefig(buf, format='png')
#         buf.seek(0)
#         plt.close()
#         return buf.getvalue().decode('latin1')

#     def generate_bar_graph(self, field_name, field_description, data):
#         """Generate bar graph for selected field with count"""
#         if field_name not in data.columns:
#             return None
            
#         value_counts = data[field_name].value_counts().head(20)  # Top 20 values
        
#         plt.figure(figsize=(12, 8))
        
#         # Create bar plot
#         bars = plt.bar(range(len(value_counts)), value_counts.values, 
#                       color=plt.cm.Paired(np.linspace(0, 1, len(value_counts))))
        
#         # Customize the plot
#         plt.title(f'Distribution of {field_description}', fontsize=16, pad=20)
#         plt.xlabel(field_description, fontsize=12)
#         plt.ylabel('Count', fontsize=12)
#         plt.xticks(range(len(value_counts)), value_counts.index, rotation=45, ha='right')
#         plt.grid(axis='y', linestyle='--', alpha=0.7)
        
#         # Add value annotations on bars
#         for bar in bars:
#             height = bar.get_height()
#             plt.annotate(f'{int(height)}',
#                         xy=(bar.get_x() + bar.get_width() / 2, height),
#                         xytext=(0, 5),
#                         textcoords="offset points",
#                         ha='center', va='bottom',
#                         fontsize=10, fontweight='bold')
        
#         plt.tight_layout()
        
#         # Save to buffer
#         buf = BytesIO()
#         plt.savefig(buf, format='png', dpi=150, bbox_inches='tight')
#         buf.seek(0)
#         plt.close()
        
#         return buf.getvalue().decode('latin1')

#     def mandt(self):
#         mandDic, mandFields = {}, []
       
#         conn = SaveRule.objects.filter(project_id=self.proj,object_id=self.obj,segment_id=self.seg, isMandatory=True)
#         if conn:
#             serializer = SaveRuleSerializer(conn,many=True)
#             for i in serializer.data:
#                 mandDic[i['target_sap_field']] = i['text_description']
#                 mandFields.append(i['target_sap_field'])

#         try:
#             with connections["default"].cursor() as cursor:
#                 mandt = ', '.join(mandFields)
#                 data = cursor.execute(f"SELECT {mandt} FROM {self.table_name}")
#                 data = data.fetchall()
#                 data = pd.DataFrame(data, columns=mandFields)
#                 return self.plot(mandDic, data)
           
#         except Exception as e:
#             print(f"{e}")
       
#         return 0

#     def profMand(self):
#         mandDic, mandFields = {}, []
#         self.tempGraph = []
#         conn = SaveRule.objects.filter(project_id=self.proj,object_id=self.obj,segment_id=self.seg, isMandatory=True)
#         if conn:
#             serializer = SaveRuleSerializer(conn,many=True)
#             for i in serializer.data:
#                 mandDic[i['target_sap_field']] = i['text_description']
#                 mandFields.append(i['target_sap_field'])

#         try:
#             with connections["default"].cursor() as cursor:
#                 mandt = ', '.join(mandFields)
#                 data = cursor.execute(f"SELECT {mandt} FROM {self.table_name}")
#                 data = data.fetchall()
#                 data = pd.DataFrame(data, columns=mandFields)
                
#                 for field_name, field_desc in mandDic.items():
#                     graph_data = self.generate_bar_graph(field_name, field_desc, data)
#                     if graph_data:
#                         self.tempGraph.append({
#                             'field_name': field_name,
#                             'description': field_desc,
#                             'graph': graph_data
#                         })
#                 return {'profMand': self.tempGraph}
       
#         except Exception as e:
#             print(f"{e}")
       
#         return 0

#     def profLook(self):
#         lookUpDic = {}  
#         self.tempGraph = []      
#         conn = SaveRule.objects.filter(project_id=self.proj,object_id=self.obj,segment_id=self.seg)
#         if conn:
#             serializer = SaveRuleSerializer(conn,many=True)
#             for i in serializer.data:
#                 if i['lookup_table']:
#                     lookUpDic[i['target_sap_field']] = i['text_description']

#         lookFields = list(lookUpDic.keys())
#         try:
#             with connections["default"].cursor() as cursor:
#                 mandLook = ', '.join(lookFields)
#                 data = cursor.execute(f"SELECT {mandLook} FROM {self.table_name}")
#                 data = data.fetchall()
#                 data = pd.DataFrame(data, columns=lookFields)
                
#                 for field_name, field_desc in lookUpDic.items():
#                     graph_data = self.generate_bar_graph(field_name, field_desc, data)
#                     if graph_data:
#                         self.tempGraph.append({
#                             'field_name': field_name,
#                             'description': field_desc,
#                             'graph': graph_data
#                         })
#                 return {'profLook': self.tempGraph}

#         except Exception as e:
#             print(f"{e}")
       
#         return 0
       
#     def get(self, request, pid, oid, sid):
#         graphs = []
#         self.proj, self.obj, self.seg = pid, oid, sid
#         conn = segments.objects.filter(project_id=pid,obj_id=oid,segment_id=sid).first()
#         self.table_name = conn.table_name
#         graphs.append({'mandatory': self.mandt()})
#         graphs.append(self.profMand())
#         graphs.append(self.profLook())

#         return JsonResponse({'plot': graphs})  
 
# class GetExactGraph(APIView):
#     def get(self, request, pid, oid, sid, fname):
#         # Get field description
#         desc = ""
#         conn = SaveRule.objects.filter(project_id=pid,object_id=oid,segment_id=sid, target_sap_field=fname)
#         if conn:
#             serializer = SaveRuleSerializer(conn,many=True)
#             for i in serializer.data:
#                 desc = i['text_description']
        
#         # Get table name
#         conn = segments.objects.filter(project_id=pid,obj_id=oid,segment_id=sid).first()
#         table_name = conn.table_name
        
#         try:
#             with connections["default"].cursor() as cursor:
#                 data = cursor.execute(f"SELECT {fname} FROM {table_name}")
#                 data = data.fetchall()
#                 data = pd.DataFrame(data, columns=[fname])

#             # Generate bar graph
#             value_counts = data[fname].value_counts().head(20)
            
#             plt.figure(figsize=(12, 8))
#             colors = plt.cm.Paired(np.linspace(0, 1, len(value_counts)))
            
#             bars = plt.bar(range(len(value_counts)), value_counts.values, color=colors)
            
#             # Customize plot
#             plt.title(f'Distribution of {desc}', fontsize=16, pad=20)
#             plt.xlabel(desc, fontsize=12)
#             plt.ylabel('Count', fontsize=12)
#             plt.xticks(range(len(value_counts)), value_counts.index, rotation=45, ha='right')
#             plt.grid(axis='y', linestyle='--', alpha=0.7)
            
#             # Add annotations
#             for bar in bars:
#                 height = bar.get_height()
#                 plt.annotate(f'{int(height)}',
#                             xy=(bar.get_x() + bar.get_width() / 2, height),
#                             xytext=(0, 5),
#                             textcoords="offset points",
#                             ha='center', va='bottom',
#                             fontsize=10, fontweight='bold')
            
#             plt.tight_layout()
            
#             buf = BytesIO()
#             plt.savefig(buf, format='png', dpi=150, bbox_inches='tight')
#             buf.seek(0)
#             plt.close()
            
#             return HttpResponse(buf.getvalue().decode('latin1'), content_type='image/png')

#         except Exception as e:
#             return HttpResponse(f"Error: {e}", status=500)

# NEW API for dynamic field selection
class DynamicFieldGraph(APIView):
    def get(self, request, pid, oid, sid):
        field_name = request.GET.get('field_name', '')
        
        if not field_name:
            return JsonResponse({'error': 'Field name is required'}, status=400)
        
        # Get field description
        desc = ""
        conn = SaveRule.objects.filter(project_id=pid,object_id=oid,segment_id=sid, target_sap_field=field_name)
        if conn:
            serializer = SaveRuleSerializer(conn,many=True)
            for i in serializer.data:
                desc = i['text_description']
        
        # Get table name
        conn = segments.objects.filter(project_id=pid,obj_id=oid,segment_id=sid).first()
        table_name = conn.table_name
        
        try:
            with connections["default"].cursor() as cursor:
                data = cursor.execute(f"SELECT {field_name} FROM {table_name}")
                data = data.fetchall()
                data = pd.DataFrame(data, columns=[field_name])

            # Generate bar graph
            value_counts = data[field_name].value_counts().head(20)
            
            plt.figure(figsize=(12, 8))
            colors = plt.cm.Paired(np.linspace(0, 1, len(value_counts)))
            
            bars = plt.bar(range(len(value_counts)), value_counts.values, color=colors)
            
            # Customize plot
            plt.title(f'Distribution of {desc}', fontsize=16, pad=20)
            plt.xlabel(desc, fontsize=12)
            plt.ylabel('Count', fontsize=12)
            plt.xticks(range(len(value_counts)), value_counts.index, rotation=45, ha='right')
            plt.grid(axis='y', linestyle='--', alpha=0.7)
            
            # Add annotations
            for bar in bars:
                height = bar.get_height()
                plt.annotate(f'{int(height)}',
                            xy=(bar.get_x() + bar.get_width() / 2, height),
                            xytext=(0, 5),
                            textcoords="offset points",
                            ha='center', va='bottom',
                            fontsize=10, fontweight='bold')
            
            plt.tight_layout()
            
            buf = BytesIO()
            plt.savefig(buf, format='png', dpi=150, bbox_inches='tight')
            buf.seek(0)
            image_data = buf.getvalue()
            plt.close()
            
            # Return as base64 encoded string for frontend
            image_base64 = base64.b64encode(image_data).decode('latin1')
            
            return JsonResponse({
                'graph': f'data:image/png;base64,{image_base64}',
                'field_name': field_name,
                'description': desc,
                'total_count': len(data),
                'unique_values': len(value_counts)
            })

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
 




@api_view(['GET'])
def demo_execute_queries(request):
    try:
        with connection.cursor() as cursor:
           
 
            # cursor.execute(f"""
            #    INSERT INTO t_26_Product_Basic_Data_mandatory_Ext (PRODUCT)
            #      SELECT MATNR as PRODUCT
            #      FROM MARA
            #      WHERE MTART = 'ROH';
            #  """)
 
 
            # cursor.execute(f"""
            #   UPDATE t_26_Product_Basic_Data_mandatory_Ext
            #     SET MTART = COALESCE(
            #         (SELECT ml.MTART FROM MARA_Legacy ml WHERE ml.MATNR = t.PRODUCT),
            #         (SELECT m.MTART FROM MARA m WHERE m.MATNR = t.PRODUCT)
            #     )
            #     FROM t_26_Product_Basic_Data_mandatory_Ext t;
            #  """)
 
 
            # cursor.execute(f"""
            #     UPDATE t_26_Product_Basic_Data_mandatory_Ext
            #         SET MAKTX = COALESCE(
            #             (SELECT makt.MAKTX
            #             FROM MAKT makt
            #             WHERE TRIM(makt.MATNR) = TRIM(t.PRODUCT)
            #             AND makt.SPRAS = 'E'),
            #             (SELECT makt_de.MAKTX
            #             FROM MAKT makt_de
            #             WHERE TRIM(makt_de.MATNR) = TRIM(t.PRODUCT)
            #             AND makt_de.SPRAS = 'D'),
            #             'NoText'
            #         )
            #         FROM t_26_Product_Basic_Data_mandatory_Ext t;          
 
            #  """)
           
 
 
 
            # cursor.execute(f"""
 
            #      UPDATE t_26_Product_Basic_Data_mandatory_Ext
            #     SET MAKTX = COALESCE(
            #         (SELECT makt.MAKTX
            #         FROM MAKT makt
            #         WHERE TRIM(makt.MATNR) = TRIM(PRODUCT)
            #         AND makt.SPRAS = 'E'),
            #         (SELECT makt_de.MAKTX
            #         FROM MAKT makt_de
            #         WHERE TRIM(makt_de.MATNR) = TRIM(PRODUCT)
            #         AND makt_de.SPRAS = 'D'),
            #         'NoText'
            #     )
            #     FROM MAKT;      
 
            #  """)
 
 
 
            # cursor.execute(f"""
 
            #     UPDATE t_26_Product_Basic_Data_mandatory_Ext
            #     SET MAKTX = (
            #         SELECT makt.MAKTX
            #         FROM MAKT makt
            #         WHERE makt.MATNR=PRODUCT
            #     )
            #     FROM MAKT;        
 
            #  """)
 
           
 
         
            connection.commit() #commit the changes to the database.
            return Response({"message": "Query executed successfully"})
 
    except Exception as e:
        connection.rollback() #rollback the changes if an error occurs.
        return Response({"message": f"Error executing query: {str(e)}"}, status=500)



@api_view(['GET'])
def getSfTableData(request,oid):
    print("Hello in SF Get")
    try:
        segment = segments.objects.filter(obj_id=oid).first()
        if not segment:
            return Response({"error": "Segment not found"}, status=404)

        table_name = segment.table_name
        print(table_name)
        with connections["default"].cursor() as cursor:
            cursor.execute(f"SELECT * FROM {table_name}")
            data = cursor.fetchall()
            cursor.execute(f"PRAGMA table_info({table_name})") # For SQLite, use PRAGMA table_info
            columns_info = cursor.fetchall()
            column_names = [col[1] for col in columns_info] # Get the column names

            # 1. Convert to list of dictionaries (JSON-like)
            results = []
            for row in data:
                row_dict = dict(zip(column_names, row))
                results.append(row_dict)

            return Response({"columns": column_names, "data": results})

    except Exception as e:
        print(f"Error fetching data: {e}")
        return Response({"error": str(e)}, status=500)




def create_new_table_with_existing_columns(existing_table_name, new_table_name, additional_columns=None):

    try:
        with connection.cursor() as cursor:

            # Get the column names and their data types from the existing table
            cursor.execute(f"PRAGMA table_info({existing_table_name})")
            existing_columns_info = cursor.fetchall()
            existing_columns = [info[1] for info in existing_columns_info]

            if not existing_columns:
                print(f"Error: Table '{existing_table_name}' has no columns.")
                return

            deleteSqlLiteTable(new_table_name)
            # Construct the CREATE TABLE statement for the new table
            create_table_statement = f"CREATE TABLE IF NOT EXISTS {new_table_name} ("
            column_definitions = []
            for col_name in existing_columns:
                # Attempt to infer the data type from the existing table's schema.
                # This is a simplified approach; a more robust solution might involve
                # querying the PRAGMA table_info for the exact data type.
                column_definitions.append(f"{col_name}")

            if additional_columns:
                for col_name in additional_columns:
                    column_definitions.append(f"{col_name} TEXT")

            create_table_statement += ", ".join(column_definitions) + ")"

            # Execute the CREATE TABLE statement
            cursor.execute(create_table_statement)
            # print("existing columns :",existing_columns)
            insert_columns = ", ".join(existing_columns)
            cursor.execute(
                f"INSERT INTO {new_table_name} ({insert_columns}) SELECT {insert_columns} FROM {existing_table_name}"
            )

            print(f"Successfully created table '{new_table_name}' with columns from '{existing_table_name}' and additional columns (if any).")

    except Exception as e:
        print(f"Error creating table '{new_table_name}': {e}")
        connection.rollback()
        raise Exception(f"Error creating validation table '{new_table_name}': {e}")


@api_view(['POST'])
def create_Validation_Table(request):
    try:
        sid = request.data['segment_id']
        print(sid) 
        print(type(sid)) 

        # deleteSqlLiteTable("t_68_Product_Additional_Descriptions_validation")
        print("Hello no error")
        seg = segments.objects.filter(segment_id=int(sid)).first()
        if seg:
            table_name = seg.table_name
            # base_name = table_name[:-4]  # Remove the last 4 characters
            base_name = table_name
            new_table_name = f"{base_name}_validation"


            validation_fields = []
            validation_fields.append('Preload_status')
            validation_fields.append('Mandatory_ErrorField')
            validation_fields.append('Lookup_ErrorTable_Field')

                    
            create_new_table_with_existing_columns(table_name,new_table_name,validation_fields)

            validation_table_data = get_complete_table_data(new_table_name)
            return Response(status=status.HTTP_200_OK,data=validation_table_data)
        else:
            print("No segment exists")
            return Response(status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        print(f"Error creating validation table: {e}")
        return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data="Error creating validation table")



@api_view(['GET'])
def Insert_Data_Into_ValidationTable(request):

    # columns = ["PRODUCT", "SPRAS", "MAKTX", "Preload_status", "Mandatory_ErrorField","Lookup_ErrorTable_Field"]

    # sample_records = [
    #     ("RM120", "R", "RAW120,PD,QualityManaged", "", "",""),
    #     (None, "DE", "RAW129,PD", "", "",""),
    #     (None, "E", "Description G", "", "",""),
    #     (None, "FR", "Description H", "", "",""),
    #     ("RM30", "Product T", "Description E","", "", ""),
    #     ("RM1", "Product V", "Description E","", "", ""),
    #     ("varun", "Product V", "RAW30,VB,Batch-Fifo","", "", ""),
    # ]

    # table_name = "t_68_Product_Additional_Descriptions_validation"
    # delete_table_data("t_68_Product_Additional_Descriptions_validation")



    columns = ["PRODUCT", "SPRAS", "MAKTX"]

    sample_records = [
        ("RM120", "R", "RAW120,PD,QualityManaged"),
        (None, "DE", "RAW129,PD"),
        (None, "ER", "Description G"),
        (None, "FR", "Description H"),
        ("RM30", "Product T", "Description E"),
        ("RM1", "Product V", "Description E"),
        ("varun", "Product V", "RAW30,VB,Batch-Fifo"),
        ("Balaji", "Product B", "RAW30,VB,Batch-Fifo"),
    ]

    table_name = "t_68_Product_Additional_Descriptions"
    delete_table_data("t_68_Product_Additional_Descriptions")
    

    try:
        with connection.cursor() as cursor:
        # Construct the INSERT SQL statement dynamically
            placeholders = ", ".join(["?"] * len(columns))
            sql = f"INSERT INTO {table_name} ({', '.join(columns)}) VALUES ({placeholders})"

            # Execute the INSERT statement for each sample record
            cursor.executemany(sql, sample_records)

            # Commit the changes


            print(f"{cursor.rowcount} records inserted successfully into {table_name}")
            return Response("Inserted Succesfully")

    except Exception as e:
        print(f"Error Inserting Data Into validation table: {e}")
        connection.rollback()
        return Response("Error")



@api_view(['POST'])
def create_PreLoad_Tables(request):

    try:
        sid = request.data['segment_id']
        seg = segments.objects.filter(segment_id=sid).first()
        if seg:
            table_name = seg.table_name
            # base_name = table_name[:-4]  # Remove the last 4 characters
            base_name = table_name
            validation_table_name = f"{base_name}_validation"
            validation_pass_table_name = f"{validation_table_name}_pass"    
            validation_fail_table_name = f"{validation_table_name}_fail"
            # deleteSqlLiteTable(validation_pass_table_name)
            # deleteSqlLiteTable(validation_fail_table_name)
            # return Response("OK")

            try:
                with connection.cursor() as cursor:
                        # 1. Check if the table exists

                        cursor.execute(f"SELECT name FROM sqlite_master WHERE type='table' AND name='{validation_table_name}';")
                        table_exists = cursor.fetchone() is not None
        
                        if not table_exists:  #Validation table does not exists
                            print("Validation Table Does not Exists")
                            return Response(status=status.HTTP_400_BAD_REQUEST)


                        # Define the column name for the status
                        status_column = "Preload_status"
    

                        # Fetch the column names from the source table
                        cursor.execute(f"PRAGMA table_info({validation_table_name})")
                        columns_info = cursor.fetchall()
                        column_names = [info[1] for info in columns_info]

                        # Construct the CREATE TABLE statements for pass and fail table dynamically
                        create_table1_columns = ", ".join([f"\"{col}\" TEXT" for col in column_names])
                        create_table2_columns = ", ".join([f"\"{col}\" TEXT" for col in column_names])

                        deleteSqlLiteTable(validation_pass_table_name)
                        deleteSqlLiteTable(validation_fail_table_name)


                        cursor.execute(f"""
                            CREATE TABLE IF NOT EXISTS {validation_pass_table_name} (
                                {create_table1_columns}
                            )
                        """)

                        cursor.execute(f"""
                            CREATE TABLE IF NOT EXISTS {validation_fail_table_name} (
                                {create_table2_columns}
                            )
                        """)

                        # cursor.execute(f"DELETE FROM {validation_pass_table_name}")
                        # cursor.execute(f"DELETE FROM {validation_fail_table_name}")


                        # Construct the SELECT statement to fetch all columns
                        select_columns = ", ".join(column_names)
                        cursor.execute(f"SELECT {select_columns} FROM {validation_table_name}")
                        all_records = cursor.fetchall()

                        # Get the index of the status column
                        try:
                            status_column_index = column_names.index(status_column)
                        except ValueError:
                            print(f"Error: Column '{status_column}' not found in the validation source table.")
                            exit()

                        # Iterate through the records and insert into the respective tables
                        for record in all_records:
                            print("Hello Balaji")
                            print(record)
                            status_value = record[status_column_index]



                            placeholders = ', '.join(['%s'] * len(column_names))
                            # insert_columns = ", ".join(column_names)
                            insert_columns = ', '.join(f'"{col}"' for col in column_names)

                            if status_value == "Pass":
                                insert_sql = f"""
                                    INSERT INTO {validation_pass_table_name} ({insert_columns})
                                    VALUES ({placeholders})
                                """
                                # print("SQL for Pass:", insert_sql)
                                cursor.execute(insert_sql, record)
                            elif status_value == "Fail":
                                insert_sql = f"""
                                    INSERT INTO {validation_fail_table_name} ({insert_columns})
                                    VALUES ({placeholders})
                                """
                                cursor.execute(insert_sql, record)
                            else:
                                print("Validations has not been done")
                                return Response(status=status.HTTP_404_NOT_FOUND,data="Validations has not been done")

                        # Commit the changes

                        print(f"Records from {validation_table_name} processed and moved to {validation_pass_table_name} and {validation_fail_table_name} based on '{status_column}'.")
        
                        return Response(status=status.HTTP_200_OK)

            except Exception as e:
                print(f"Error creating PreLoad tables: {e}")
                connection.rollback()
                return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    except Exception as e:
        print(f"Error in creating PreLoad tables: {e}")
        return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR, data="Error creating PreLoad tables")





# from cryptography.fernet import Fernet
# from django.conf import settings

# def get_encryption_key():
#     """
#     Retrieves the encryption key from Django settings.
#     Ensure this key is securely generated and stored.
#     """
#     key = getattr(settings, 'DATA_ENCRYPTION_KEY', None)
#     if not key:
#         raise ValueError("DATA_ENCRYPTION_KEY not set in Django settings.")
#     return key.encode()  # Key must be bytes

# def encrypt_data(data: str) -> bytes:
#     """Encrypts the given string data."""
#     key = get_encryption_key()
#     f = Fernet(k  ey)
#     encrypted_data = f.encrypt(data.encode())
#     return encrypted_data

# def decrypt_data(encrypted_data: bytes) -> str:
#     """Decrypts the given bytes data."""
#     key = get_encryption_key()
#     f = Fernet(key)
#     decrypted_data = f.decrypt(encrypted_data).decode()
#     return decrypted_data






# from django.conf import settings
# from Crypto.Cipher import AES
# from Crypto.Util.Padding import pad, unpad
# import base64
# def get_encryption_key():
#     key = getattr(settings, 'DATA_ENCRYPTION_KEY', None)
#     if not key:
#         raise ValueError("DATA_ENCRYPTION_KEY not set in Django settings.")
#     return key.encode('utf-8')

# def encrypt_data(data: str) -> str:
#     key = get_encryption_key()
#     cipher = AES.new(key, AES.MODE_CBC)
#     ciphered_data = cipher.encrypt(pad(data.encode('utf-8'), AES.block_size))
#     iv = base64.b64encode(cipher.iv).decode('utf-8')
#     ciphertext = base64.b64encode(ciphered_data).decode('utf-8')
#     return json.dumps({'iv': iv, 'ciphertext': ciphertext})

# def decrypt_data(encrypted_payload: str) -> str:
#     key = get_encryption_key()
#     payload = json.loads(encrypted_payload)
#     iv = base64.b64decode(payload['iv'])
#     ciphertext = base64.b64decode(payload['ciphertext'])
#     cipher = AES.new(key, AES.MODE_CBC, iv)
#     padded_plaintext = cipher.decrypt(ciphertext)
#     plaintext = unpad(padded_plaintext, AES.block_size).decode('utf-8')
#     return plaintext


# @api_view(['GET'])
# def demo_encryption(request):

#     prj = Project.objects.all()

#     data = ProjectSerializer(prj,many=True)
#     # print(data.data)

#     ency_text = encrypt_data(json.dumps(data.data))
#     print("Encrypted Text : ",ency_text)
#     # print("Decrypted Text : ",decrypt_data(ency_text))
#     return Response(ency_text)








def read_sqlite_table_to_dataframe(database_path, table_name):
    try:
        # Connect to the SQLite database
        conn = sqlite3.connect(database_path)

        # Read the table data into a DataFrame
        query = f"SELECT * FROM {table_name}"
        df = pd.read_sql_query(query, conn)

        # Close the database connection
        conn.close()

        return df

    except sqlite3.Error as e:
        print(f"An error occurred: {e}")
        return None  # Return None if there's an error

    except pd.errors.DatabaseError as e: # Catch pandas sql errors as well.
        print(f"Pandas Database error: {e}")
        return None






import io
import openpyxl

@api_view(['POST'])
def sqltable_to_excel(request):
    def save_table_as_excel(table_name, output_dir=r'C:\Users\gajananda.mallidi\Downloads'):
        # Safely get database cursor for default DB
        with connections["default"].cursor() as cursor:
            cursor.execute(f"SELECT * FROM {table_name}")
            rows = cursor.fetchall()
            columns = [col[0] for col in cursor.description]

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = table_name

        # Write header row
        ws.append(columns)
        # Write data rows
        for row in rows:
            ws.append(row)

        # Save file to disk
        output_path = os.path.join(output_dir, f"{table_name}.xlsx")
        wb.save(output_path)
        return output_path
    return Response(save_table_as_excel(request.data['table_name']))

    # # database_file = r"C:\Users\varunbalaji.gada\Desktop\New_LLM\LLM_Backend\db.sqlite3"
    
    # database_file = r"C:\Users\gajananda.mallidi\Downloads\NEWPUSH\NEWPUSH\LLM_Backend\db.sqlite3"
    # print("started")
    # table_name = request.data['table_name']  # Replace with your table name

    # df = read_sqlite_table_to_dataframe(database_file, table_name)


    # model_name  = table_name

    # output = io.BytesIO()
    # with pd.ExcelWriter(output, engine='openpyxl') as writer: #Using openpyxl engine
    #     df.to_excel(writer, index=False, sheet_name="data")  # Add sheet name

    # output.seek(0)  # Rewind to the beginning of the file

    # # Create the HTTP response with appropriate headers
    # response = HttpResponse(
    #     content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    # )
    # response['Content-Disposition'] = f'attachment; filename="{model_name}.xlsx"'
    # response.write(output.getvalue()) #Write the file to response
    # print("done")
    # return response

    # return Response("Hello")




@api_view(['GET'])
def excel_data_to_sqllite(request):

    column = []
    table_name = "t_81_Supplier_Company_Data"
    skip_rows = []

    # Load workbook
    wb = load_workbook(r"C:\Users\gajananda.mallidi\Downloads\t_81_Supplier_Company_Data.xlsx", data_only=False)  # Needed to get raw display formatting
    ws = wb["Sheet1"]

    data = []
    for ii, row in enumerate(ws.iter_rows()):
        if ii in skip_rows:
            continue
        row_data = []
        for cell in row:
            if cell.value is None:
                row_data.append("")
            else:
                # NEW: get exactly how Excel displays it
                if hasattr(cell, 'number_format') and cell.is_date:
                    row_data.append(cell.value.strftime('%d.%m.%Y'))
                else:
                    # Use the formatted value from Excel's UI representation
                    try:
                        display_val = cell.number_format
                        # Trick: use cell._value as stored raw, but better: formatted
                        s = cell._value if isinstance(cell._value, str) else str(cell._value)
                        row_data.append(s)
                    except:
                        row_data.append(str(cell.value))
        data.append(row_data)

    df1 = pd.DataFrame(data)

    # Set first row as header
    df1.columns = df1.iloc[0]
    df1 = df1[1:].reset_index(drop=True)


    # df1 = pd.read_excel(, sheet_name="Sheet1")

    for i in df1.columns:
        column.append([i,"TEXT"])
    cursor = connection.cursor()
    # Drop table if it exists
    cursor.execute(f"DROP TABLE IF EXISTS {table_name};")    
    create_table(table_name , column)    
    print(df1.head())
    insert_data_from_dataframe(df1,table_name)

    return Response("Hello")




from django.http import FileResponse
from django.conf import settings
import os

def download_database(request):
    """
    Allows users to download the SQLite3 database file.
    """
    db_path = os.path.join(settings.BASE_DIR, 'db.sqlite3')  # Adjust path if needed

    if os.path.exists(db_path):
        response = FileResponse(open(db_path, 'rb'))
        response['Content-Disposition'] = 'attachment; filename="database_backup.sqlite3"'
        return response
    else:
        # Handle the case where the database file doesn't exist
        return HttpResponseNotFound("Database file not found.")

from django.http import HttpResponseNotFound





import json
from django.db import connection

def get_table_data(table_name):
    with connection.cursor() as cursor:
        print(table_name)
        cursor.execute(f"SELECT * FROM \"{table_name}\"")
        columns = [col[0] for col in cursor.description]
        rows = cursor.fetchall()
        data = []
        if rows:
            for row in rows:
                row_dict = {}
                for col, value in zip(columns, row):
                    if value is not None and value != "":
                        row_dict[col] = value
                if row_dict:  # Only add if there's at least one non-null field
                    data.append(row_dict)

        # data = [dict(zip(columns, row)) for row in rows]
        print("Hello came out from the If ")
        table_data = {}
        table_data['columns'] = columns
        if len(data)>0:
            # table_data.append({"data":data})
            table_data['data'] = data
        print("Data is returned")
        return table_data


import json
from django.http import HttpResponse, HttpResponseNotFound
from django.core import serializers
from django.apps import apps
from .models import Project, Connection, objects, segments, fields, SaveRule, Chat, Rule, FileConnection, erp_tables_description  # Assuming your models are in the same app
 
@api_view(['GET'])
def download_project_data(request, project_id):
    """Downloads data for a specific project and its associated segment tables as a JSON file."""
    try:
        project = Project.objects.get(project_id=project_id)
    except Project.DoesNotExist:
        return HttpResponseNotFound(f"Project with ID {project_id} not found.") 
 
    all_data = {
        'project': json.loads(serializers.serialize('json', [project])),
        'connections': json.loads(serializers.serialize('json', project.connections.all())),
        'objects': json.loads(serializers.serialize('json', project.objects_project.all())),
        'fields': json.loads(serializers.serialize('json', fields.objects.filter(project_id=project_id))),
        'save_rules': json.loads(serializers.serialize('json', SaveRule.objects.filter(project_id=project_id))),
        'chats': json.loads(serializers.serialize('json', Chat.objects.filter(project_id=project_id))),
        'rules': json.loads(serializers.serialize('json', Rule.objects.filter(project_id=project_id))),
        'file_connections': json.loads(serializers.serialize('json', project.files.all())),
        'erp_tables_description': json.loads(serializers.serialize('json', erp_tables_description.objects.all())), # Assuming you want all descriptions
    }
 
    Segments = segments.objects.filter(project_id=project_id)
    all_data['Segments'] = json.loads(serializers.serialize('json', Segments))
 
    for segment in Segments:
        table_name = segment.table_name  # Assuming your Segment model has a field named 'table_name'
        tables_by_segment = {}
        try:
            table_data = get_table_data(table_name)
            all_data[table_name] = table_data
        except Exception as e:
            tables_by_segment[table_name] = f"Error loading table: {str(e)}"
 
    response = HttpResponse(json.dumps(all_data, indent=4), content_type='application/json')
    response['Content-Disposition'] = f'attachment; filename="project_{project_id}_data.json"'
    return response
 






import json
import os
import django
from django.conf import settings
from django.core.exceptions import ObjectDoesNotExist

@api_view(['POST'])
def import_project_data(request):
    """
    Imports project data from a JSON file into the database.
    """
    json_file_path = "C:\\Users\\varunbalaji.gada\\Downloads\\project_24_data (3).json"
    if not os.path.exists(json_file_path):
        print(f"Error: JSON file not found at {json_file_path}")
        return Response("JSON file not found", status=status.HTTP_400_BAD_REQUEST)

    with open(json_file_path, 'r') as f:
        try:
            data = json.load(f)
        except json.JSONDecodeError:
            print(f"Error: Could not decode JSON from {json_file_path}")
            return Response("Invalid JSON format", status=status.HTTP_400_BAD_REQUEST)

    project_data = data.get('project')
    connections_data = data.get('connections')
    objects_data = data.get('objects')
    segments_data = data.get('Segments')
    fields_data_json = data.get('fields')
    save_rules_data = data.get('save_rules')
    chats_data = data.get('chats')
    rules_data = data.get('rules')
    file_connections_data = data.get('file_connections')
    dd02l_desc_data = data.get('erp_tables_description')

    project_id = 0
    object_ids = {}
    segment_ids = {}
    fields_ids = {}
    target_table_names = {}
    try:
        # Import Project
        if project_data:
            for item in project_data:
                fields_data = item.get('fields', {})  # Ensure fields is always a dict
                # Remove project_id from fields to allow auto-generation, handle the case where fields is None
                fields_data.pop('project_id', None)

                print("Fields : ",fields_data)

                try:
                    # proj = Project.objects.filter(project_name=fields_data.get('project_name')).first()
                    # if proj:
                    #      return Response(status=status.HTTP_400_BAD_REQUEST, data="Project already exists with the same name.")
                    # else:
                    #     project = ProjectSerializer(data=fields_data)

                    #     if project.is_valid():
                    #         project.save()
                    #         project_id = project.instance
                    #     else:
                    #         return Response(status=status.HTTP_400_BAD_REQUEST, data=project.errors)


                    project, created = Project.objects.get_or_create(
                        project_name=fields_data.get('project_name'),
                        defaults=fields_data
                    )
                    print(f"Project '{project.project_name}' {'created' if created else 'found'} with ID {project.project_id}.")
                    print("Created :" ,created)
                    print("Project Details :" ,project)
                    if not created:
                        return Response(status=status.HTTP_400_BAD_REQUEST, data="Project already exists with the same name.")
                    project_id = project  # Store the created/found project instance
                    print("Project Id :",project_id)

                except Exception as e:
                    print(f"Error creating/retrieving project: {e}")
                    return Response(f"Error processing project data: {e}", status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        else:
            print("No project data found in the JSON.")
            return Response("No project data found", status=status.HTTP_400_BAD_REQUEST)


        # Import Connections
        if connections_data:
            for item in connections_data:
                fields_data = item.get('fields', {})
                fields_data.pop('project_id',None)
                try:
                    # Use current_project.pk instead of fields['project_id']
                    Connection.objects.create(project_id=project_id, **fields_data)
                    print(f"Connection '{fields_data.get('connection_name', 'Unnamed')}' created.")
                except Exception as e:
                    print(f"Error creating connection: {e}")

        # Import Objects
        if objects_data:
            for item in objects_data:
                fields_data = item.get('fields', {})
                fields_data.pop('project_id',None)
                try:
                    objects.objects.create(project_id=project_id, **fields_data)
                    object_id = objects.objects.get(project_id=project_id,obj_name=fields_data['obj_name'])
                    object_ids[item.get('pk')] = object_id
                    print(f"Object '{fields_data.get('obj_name', 'Unnamed')}' created.")
                except Exception as e:
                    print(f"Error creating object: {e}")


        print("Object Ids : ",object_ids)

        # Import Segments
        print("Outside Segments...")
        if segments_data:
            print("Inside Segments...")
            for item in segments_data:
                fields_data = item.get('fields', {})
                print(fields_data)
                try:
                    objects_instance = object_ids.get(fields_data.get('obj_id'))
                    fields_data.pop('project_id',None)
                    fields_data.pop('obj_id',None)
                    obj_name = TableName_Modification(objects_instance.obj_name)
                    segment_name = TableName_Modification(fields_data.get('segement_name'))
                    table_name = "t_"+str(project_id.project_id)+"_"+obj_name+"_"+segment_name
                    target_table_names[fields_data.get('table_name')] = table_name
                    fields_data.pop('table_name',None)
                    segment_ids[item.get('pk')] = segments.objects.create(project_id=project_id, obj_id=objects_instance, table_name = table_name ,  **fields_data)
                    print(f"Segment '{fields_data.get('segement_name', 'Unnamed')}' created.")
                except ObjectDoesNotExist as e:
                    print(f"Error creating segment: {e}")
        
        print("Segment Ids : ",segment_ids)

        print("Modified Table Names : ",target_table_names)

        # Import Fields
        if fields_data_json:
            for item in fields_data_json:
                print("Hello varun")
                fields_data_item = item.get('fields', {})
                try:
                    print("Hello in fields")
                    objects_instance = object_ids.get(fields_data_item['obj_id'])
                    print("Hello in fields1")
                    segment_instance = segment_ids.get(fields_data_item['segement_id'])
                    print("Hello in fields2")
                    fields_data_item.pop('project_id',None)
                    print("Hello in fields3")
                    fields_data_item.pop('obj_id',None)
                    print("Hello in fields4")
                    fields_data_item.pop('segement_id',None)
                    print("Hello in fields5")
                    print("Hello")  
                    fields_ids[item.get('pk')] = fields.objects.create(project_id=project_id, obj_id=objects_instance, segement_id=segment_instance, **fields_data_item)
                    print("Hello 2")
                    print(f"Field '{fields_data_item.get('fields', 'Unnamed')}' created.")
                except ObjectDoesNotExist as e:
                    print(f"Error creating field: {e}")

        print("Ouside of Fields")
        # Import Save Rules
        if save_rules_data:
            print("Ouside of Fields")
            for item in save_rules_data:
                fields_data = item.get('fields', {})
                try:
                    objects_instance = object_ids.get(fields_data['object_id'])
                    segment_instance = segment_ids.get(fields_data['segment_id'])
                    field_instance = fields_ids.get(fields_data['field_id']) # Changed to pk
                    fields_data.pop('project_id',None)
                    fields_data.pop('object_id',None)
                    fields_data.pop('segment_id',None)
                    fields_data.pop('field_id',None)
                    SaveRule.objects.create(project_id=project_id, object_id=objects_instance, segment_id=segment_instance, field_id=field_instance, **fields_data)
                    print(f"SaveRule '{fields_data.get('rule_no', 'Unnamed')}' created.")
                except ObjectDoesNotExist as e:
                    print(f"Error creating SaveRule: {e}")

        # # Import Chats
        # if chats_data:
        #     for item in chats_data:
        #         fields = item.get('fields', {})
        #         try:
        #             objects_instance = objects.objects.get(pk=fields['object_id']) # Changed to pk
        #             segment_instance = segments.objects.get(pk=fields['segment_id']) # Changed to pk
        #             Chat.objects.create(project_id=current_project, object_id=objects_instance, segment_id=segment_instance, **fields)
        #             print(f"Chat message created at {fields.get('created_at', '')} {fields.get('created_time', '')}.")
        #         except ObjectDoesNotExist as e:
        #             print(f"Error creating chat: {e}")

        # Import Rules (your second Rule model)
        if rules_data:
            for item in rules_data:
                fields_data = item.get('fields', {})
                try:
                    objects_instance = object_ids.get(fields_data['object_id'])
                    segment_instance = segment_ids.get(fields_data['segment_id'])
                    field_instance = fields_ids.get(fields_data['field_id']) # Changed to pk
                    fields_data.pop('project_id',None)
                    fields_data.pop('object_id',None)
                    fields_data.pop('segment_id',None)
                    fields_data.pop('field_id',None)
                    Rule.objects.create(project_id=project_id, object_id=objects_instance, segment_id=segment_instance, field_id = field_instance , **fields_data)
                    print(f"Rule (version {fields_data.get('version_id', 'N/A')}) created.")
                except ObjectDoesNotExist as e:
                    print(f"Error creating Rule: {e}")

        # Import File Connections
        if file_connections_data:
            for item in file_connections_data:
                fields_data = item.get('fields', {})
                try:
                    fields_data.pop('project_id',None)
                    FileConnection.objects.create(project_id=project_id, **fields_data)
                    print(f"FileConnection '{fields_data.get('fileName', 'Unnamed')}' created.")
                except Exception as e:
                    print(f"Error creating file connection: {e}")

        # # Import dd02l_desc (assuming you want to import all)
        # if dd02l_desc_data:
        #     for item in dd02l_desc_data:
        #         fields = item.get('fields', {})
        #         dd02l_desc.objects.create(**fields)
        #         print(f"dd02l_desc '{fields.get('table', 'Unnamed')}' created.")

        if target_table_names:
            for table_name, new_table_name in target_table_names.items():
                try:
                    # Create the new table with the same structure as the existing one
                    table_data = data.get(table_name)
                    if not table_data:
                        print(f"Error: No data found for table '{table_name}' in the JSON.")
                        continue

                    columns = []
                    col = []
                    column = table_data.get('columns',[])
                    for c in column:
                        col = []
                        col.append(c)
                        col.append("TEXT")
                        columns.append(col)
                    create_table(new_table_name ,columns)

                    if table_data.get('data'):
                        insert_data(table_data.get('data'), new_table_name)
                        print(f"Data inserted into '{new_table_name}' from '{table_name}'.")


                except Exception as e:
                    print(f"Error creating table '{new_table_name}': {e}")

        print("Project data import process completed.")
        return Response("Success", status=status.HTTP_200_OK)

    except Exception as e:
        print(f"An unexpected error occurred during import: {e}")
        return Response(f"Error: {e}", status=status.HTTP_400_BAD_REQUEST)


import sqlite3
 
def insert_data(data_to_insert, table_name):
    """
    Inserts a list of dictionaries into a specified SQLite table.
 
    Args:
        data_to_insert (list): A list of dictionaries, where each dictionary
                               represents a row of data with keys matching the
                               column names of the table.
        table_name (str): The name of the table to insert the data into.
        db_name (str, optional): The name of the SQLite database file.
                                   Defaults to "mydatabase.db".
    """
    try:
        with connection.cursor() as cursor:
 
            if not data_to_insert:
                print("Warning: 'data_to_insert' list is empty, no rows to insert.")
                return
    
            # Assuming the keys of the first dictionary in the list are the column names
            columns = list(data_to_insert[0].keys())
            placeholders = ', '.join(['?'] * len(columns))
            column_names = ', '.join(f'"{col}"' for col in columns)
            sql = f"INSERT INTO {table_name} ({column_names}) VALUES ({placeholders})"
    
            rows_to_insert = []
            for row_dict in data_to_insert:
                row_values = [row_dict.get(col) for col in columns]
                rows_to_insert.append(row_values)
    
            cursor.executemany(sql, rows_to_insert)
    
            print(f"Successfully inserted {cursor.rowcount} rows into table '{table_name}'.")
    
    except sqlite3.Error as e:
        print(f"SQLite error: {e}")



def check_table_existance(table_name):
    try:
        with connection.cursor() as cursor:
                # 1. Check if the table exists

                cursor.execute(f"SELECT name FROM sqlite_master WHERE type='table' AND name='{table_name}';")
                table_exists = cursor.fetchone() is not None

                if not table_exists: 
                    return False
                return True
        
    except Exception as e:
        print(f"Error in Checking Table Exsistance : {e}")
        raise Exception(f"Error in Checking Table Exsistance : {e}")


#Validation for Mandatory Fields
def set_preload_status_fail_if_null(table_name, columns_to_check):
    """
    Sets Preload_status to 'Fail' if any of the specified columns are NULL in the given table.
    Also updates ErrorField_Table with the names of the columns that are NULL,
    appending to any existing values.
 
    Args:
        table_name (str): The name of the table to update.
        columns_to_check (list): A list of column names to check for NULL values.
    """
    try:
        null_conditions = " OR ".join([f"{col} IS NULL" for col in columns_to_check])
        error_fields_case = " || ".join([f"CASE WHEN {col} IS NULL THEN '{col} , ' ELSE '' END" for col in columns_to_check])


        sql = f"""
            UPDATE {table_name}
            SET
                Preload_status = CASE 
                                WHEN {null_conditions} 
                                    THEN 'Fail' 
                                    ELSE CASE 
                                            WHEN Lookup_ErrorTable_Field = '' OR Lookup_ErrorTable_Field IS NULL  
                                            THEN 'Pass' 
                                            ELSE 'Fail' END 
                                END,
                Mandatory_ErrorField = CASE
                    WHEN {null_conditions} THEN
                        RTRIM({error_fields_case},' , ')
                    ELSE Mandatory_ErrorField
                    END
        """
        
        # WHERE {null_conditions} OR Preload_status != 'Pass'

        try:
            with connection.cursor() as cursor:  # Assuming 'connection' is a valid database connection object
                cursor.execute(sql)
                connection.commit() # Added commit to save
        except Exception as e:
            print(f"Error updating table: {e}")
            # Consider re-raising the exception or logging it.  A bare except is bad practice.
            raise # re-raise the exception so the caller knows
    except Exception as e:
        print(f"Error in set_preload_status_fail_if_null: {e}")
        raise



def column_existance(table_name , column_name):
    try:
        cursor = connection.cursor()
        cursor.execute(f"PRAGMA table_info({table_name});")
        columns = [row[1] for row in cursor.fetchall()]
        return column_name in columns 
    except Exception as e:
        print(f"Error checking column existence: {e}")
        raise


@api_view(['POST'])
def validate_mandatory_fields(request):

    try:
        
        project_id = request.data['project_id']
        object_id = request.data['object_id']
        segment_id = request.data['segment_id']
        print("Project Id :",project_id)
        print("Object Id :",object_id)
        print("Segment Id :",segment_id)

        latest_version = Rule.objects.filter(project_id=project_id ,object_id = object_id ,segment_id = segment_id).order_by('-version_id').first()
        if latest_version:
            latest_version_number = latest_version.version_id
            mandatory_fields = Rule.objects.filter(project_id=project_id ,object_id = object_id ,segment_id = segment_id, version_id = latest_version_number , user_mandatory = 1).values('target_sap_field')
            if mandatory_fields:
                # print("Latest Version Id :",latest_version.version_id)
                # print("Mandatory Fields :",mandatory_fields)
                mandatory_fields = list(mandatory_fields)
                mandatory_fields = [field['target_sap_field'] for field in mandatory_fields]
                # print("Mandatory Fields :",mandatory_fields)
                validation_table_name = segments.objects.get(segment_id = segment_id).table_name + "_validation"
                # print("Validation Table Name :",validation_table_name)
                if check_table_existance(validation_table_name):
                    set_preload_status_fail_if_null(validation_table_name,mandatory_fields)
                else:
                    return Response(status=status.HTTP_404_NOT_FOUND,data = "Validation Table not found")

                table_data = get_complete_table_data(validation_table_name)
                return Response(status=status.HTTP_200_OK,data=table_data)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND,data="No Mandatory Fields Found")
        else:
            return Response(status=status.HTTP_404_NOT_FOUND,data = "No Version data found for given Project ,Object & Segment Id's")

    except Exception as e:
        print("Error in validating mandatory fields : ",e)
        return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)





# def set_loookup_table_status(table_name,lookup_fields):
    

#     # Set Lookup_ErrorTable_Field to empty string for all records
#     sql1 =f"""
#             UPDATE {table_name} 
#                 SET Lookup_ErrorTable_Field = ''
#             """

#     try:
#         with connection.cursor() as cursor:  # Assuming 'connection' is a valid database connection object
#             cursor.execute(sql1)
#             connection.commit() # Added commit to save
#     except Exception as e:
#         print(f"Error updating table: {e}")
#         # Consider re-raising the exception or logging it.  A bare except is bad practice.
#         raise # re-raise the exception so the caller knows
#         # connection.rollback() # Rollback in case of error


#     #validating if any table or column is not present
#     for field_data in lookup_fields:
#         field_name = field_data['target_sap_field']
#         lookup_field_name = field_data['lookup_field']
#         lookup_table_name = field_data['lookup_table']

#         if check_table_existance(lookup_table_name):
#             if column_existance(lookup_table_name,lookup_field_name):
#                 pass
#             else:
#                 return f"column {lookup_field_name} not found in table {lookup_table_name}"
#         else:
#             return f"Table {lookup_table_name} not found in database"
    
#     #Executing the Main Logic
#     for field_data in lookup_fields:
#         field_name = field_data['target_sap_field']
#         lookup_field_name = field_data['lookup_field']
#         lookup_table_name = field_data['lookup_table']

#         sql2 =f"""
#                 UPDATE {table_name}
#                 SET 
#                     Preload_status = 
#                         CASE 
#                             WHEN {field_name} IN (SELECT {lookup_field_name} FROM {lookup_table_name}) 
#                                 THEN CASE 
#                                     WHEN ( Mandatory_ErrorField = '' OR Mandatory_ErrorField IS NULL ) AND ( Lookup_ErrorTable_Field = '' OR Lookup_ErrorTable_Field IS NULL )
#                                         THEN 'Pass' 
#                                     ELSE 'Fail' 
#                                 END
#                             ELSE 'Fail' 
#                         END,
#                     Lookup_ErrorTable_Field = 
#                         CASE 
#                             WHEN {field_name} IN (SELECT {lookup_field_name} FROM {lookup_table_name}) 
#                                 THEN Lookup_ErrorTable_Field
#                             ELSE CASE 
#                                     WHEN Lookup_ErrorTable_Field = '' OR Lookup_ErrorTable_Field IS NULL 
#                                     THEN '{lookup_table_name}_{lookup_field_name}'
#                                     ELSE Lookup_ErrorTable_Field || ' , ' || '{lookup_table_name}_{lookup_field_name}'
#                                 END
#                         END
#             """
                    
#         try:
#             with connection.cursor() as cursor:  # Assuming 'connection' is a valid database connection object
#                 cursor.execute(sql2)
#                 connection.commit() # Added commit to save
#         except Exception as e:
#             print(f"Error updating table: {e}")
#             # Consider re-raising the exception or logging it.  A bare except is bad practice.
#             raise # re-raise the exception so the caller knows
#             # connection.rollback() # Rollback in case of error
                   

#     return "Lookup Table Status Updated Successfully"




# @api_view(['POST'])
# def validate_Lookup_fields(request):
#     try:
#         project_id = request.data['project_id']
#         object_id = request.data['object_id']
#         segment_id = request.data['segment_id']
#         print("Project Id :",project_id)
#         print("Object Id :",object_id)
#         print("Segment Id :",segment_id)

#         latest_version = Rule.objects.filter(project_id=project_id ,object_id = object_id ,segment_id = segment_id).order_by('-version_id').first()
#         if latest_version:
#             latest_version_number = latest_version.version_id
#             lookup_rule_fields = Rule.objects.filter(project_id=project_id ,object_id = object_id ,segment_id = segment_id, version_id = latest_version_number , lookup_table__isnull = False , lookup_field__isnull = False).exclude(lookup_field='').exclude(lookup_table='').values('target_sap_field','lookup_table','lookup_field')
#             if lookup_rule_fields: 
#                 print("Latest Version Id :",latest_version.version_id)
#                 print("Mandatory Fields :",lookup_rule_fields)
#                 lookup_fields = list(lookup_rule_fields)
#                 # lookup_fields = [field['target_sap_field'] for field in lookup_rule_fields]
#                 print("Mandatory Fields :",lookup_fields)
#                 validation_table_name = segments.objects.get(segment_id = segment_id).table_name + "_validation"
#                 print("Validation Table Name :",validation_table_name)
#                 if check_table_existance(validation_table_name):
#                     lookup_response=set_loookup_table_status(validation_table_name,lookup_fields)
#                     if lookup_response != "Lookup Table Status Updated Successfully":
#                         return Response(status=status.HTTP_404_NOT_FOUND,data = lookup_response)
#                     seg = segments.objects.get(segment_id=segment_id)
#                     seg.validation_last_changed = timezone.now()
#                     latest_version = Rule.objects.filter(project_id=project_id ,object_id = object_id ,segment_id = segment_id).order_by('-version_id').first()
#                     latest_version_number = latest_version.version_id
#                     seg.validation_version = latest_version_number
#                     seg.save(update_fields=['validation_last_changed', 'validation_version'])
#                     table_data = get_complete_table_data(validation_table_name)
#                     return Response(status=status.HTTP_200_OK,data=table_data)

#                 else:
#                     return Response(status=status.HTTP_404_NOT_FOUND,data = "Validation Table not found")
#             else:
#                 return Response("No Lookup Fields Found")
#         else:
#             return Response(status=status.HTTP_404_NOT_FOUND,data = "No Version data founf for given Project ,Object & Segment Id's")

#     except Exception as e:
#         print("Error in validating Lookup fields : ",e)
#         return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR)



# def deterimine_LookTableButton_status(segment_id):
#     try:
#         if segments.objects.filter(segment_id = segment_id).exists():
#             seg = segments.objects.get(segment_id = segment_id)
#             ext_time = seg.extraction_last_changed
#             valid_time = seg.validation_last_changed
#             valid_version = seg.validation_version
#             latest_version = Rule.objects.filter(project_id = seg.project_id ,object_id = seg.obj_id ,segment_id = segment_id).order_by('-version_id').first()
#             if latest_version:
#                 latest_version_number = latest_version.version_id
#                 if ext_time != None and valid_time != None:
#                     if ext_time > valid_time:
#                         return "True ext time"
#                     else:
#                         print(valid_version,latest_version_number)
#                         if int(str(valid_version).strip()) == int(str(latest_version_number).strip()):  
#                             return "False"
#                         else:
#                             return "True at look up"
#                 elif ext_time != None and valid_time == None:
#                     return "True"
#             else:
#                 return "No LookUp Tables are found in this segment"
#         else:
#             return Response(status=status.HTTP_404_NOT_FOUND,data = "Segment not found")
#     except Exception as e:
#         print("Error in determining Lookup Table Button Status : ",e)
#         return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR,data = "Error in determining Lookup Table Button Status")



def get_row_count(table_name):
    """
    Returns the number of rows in the specified table.

    Args:
        table_name (str): Name of the table.

    Returns:
        int: Number of rows in the table.
    """
    if check_table_existance(table_name):
        cursor = connection.cursor()
        try:
            cursor.execute(f"SELECT COUNT(*) FROM {table_name}")
            count = cursor.fetchone()[0]
            return count
        finally:
            cursor.close()
    else:
        return 0



@api_view(['GET'])
def final_report(request,project_id):

    # return Response(get_row_count("MARA"))

    # project_id = request.data['project_id']
    
    project_report = []
    try:
        if Project.objects.filter(project_id = project_id).exists():
            
            objects_data = objects.objects.filter(project_id = project_id)
            if objects_data.exists():
                for obj in objects_data:
                    single_objects_data = {}
                    single_objects_data['key'] = obj.obj_id
                    single_objects_data['object_OR_segment_name'] = obj.obj_name
                    obj_segments = segments.objects.filter(obj_id = obj.obj_id)
                    map_count = 0
                    valid_count = 0
                    invalid_count = 0
                    segments_data = []
                    if obj_segments.exists():
                        for seg in obj_segments:
                            single_segments_data = {}
                            single_segments_data['key'] = str(obj.obj_id)+"_"+str(seg.segment_id)
                            single_segments_data['object_OR_segment_name'] = seg.segement_name
                            extraction_table_name = seg.table_name
                            extraction_count = get_row_count(extraction_table_name)
                            single_segments_data['map'] = extraction_count
                            pass_table_name = seg.table_name + "_validation_pass"
                            fail_table_name = seg.table_name + "_validation_fail"
                            pass_count = get_row_count(pass_table_name)
                            fail_count = get_row_count(fail_table_name)
                            single_segments_data['valid'] = pass_count
                            single_segments_data['invalid'] = fail_count
                            single_segments_data['timestamp'] = "2015-01-15 12:30:00"
                            if extraction_count != 0:
                                if pass_count == 0 and fail_count == 0:
                                    single_segments_data['valid'] = extraction_count
                                    pass_count = extraction_count
                                single_segments_data['enriched_percentage'] = round((pass_count/extraction_count)*100,2)
                                segments_data.append(single_segments_data)
                                map_count += extraction_count
                                valid_count += pass_count
                                invalid_count += fail_count


                    single_objects_data['map'] = map_count
                    single_objects_data['valid'] = valid_count
                    single_objects_data['invalid'] = invalid_count
                    single_objects_data['timestamp'] = "2015-01-15 12:30:00"
                    if map_count != 0:
                        single_objects_data['enriched_percentage'] = round((valid_count/map_count)*100,2)
                    else:
                        single_objects_data['enriched_percentage'] = 0
                    
                    print("No. of segments : ",len(segments_data))
                    if len(segments_data) != 0:
                        single_objects_data['children'] = segments_data
                    else:
                        temp_data = {}
                        temp_data['key'] = str(obj.obj_id)+"_0"
                        temp_data['object_OR_segment_name'] = "No Data Found in Segments"
                        segments_data.append(temp_data)
                        single_objects_data['children'] = segments_data


                    project_report.append(single_objects_data)
                
                return Response(status=status.HTTP_200_OK,data = project_report)


            else:
                return Response(status=status.HTTP_404_NOT_FOUND,data = "No Objects found for the Selected Project")

        else:
            return Response(status=status.HTTP_404_NOT_FOUND,data = "Project not found")
    except Exception as e:
        print("Error in Project Report Generation : ",e)
        return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR,data = "Error in Project Report Generation")
    


@api_view(['GET'])
def download_final_report(request,segment_id,table_type):

    try:
        download_table_name =""
        if segments.objects.filter(segment_id = segment_id).exists():
            table_name = segments.objects.get(segment_id = segment_id).table_name
            download_table_name = segments.objects.get(segment_id = segment_id).segement_name
            if table_type == "valid":
                table_name = table_name + "_validation_pass"
                download_table_name = download_table_name+"_Valid"
            elif table_type == "invalid":
                table_name = table_name + "_validation_fail"
                download_table_name = download_table_name+"_Invalid"
            elif table_type == "map":
                table_name = table_name + "_validation"
            else:
                return Response(status=status.HTTP_400_BAD_REQUEST,data = "Invalid Request")
            
            if check_table_existance(table_name):
                cursor = connection.cursor()
                cursor.execute(f"SELECT * FROM {table_name}")
                rows = cursor.fetchall()
                columns = [column[0] for column in cursor.description]
                df = pd.DataFrame(rows, columns=columns)
                model_name = download_table_name  # More descriptive
                print(model_name)
        
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name=model_name)
        
                output.seek(0)
        
                response = HttpResponse(
                    output.getvalue(),  # Directly pass the file content
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="{model_name}.xlsx"'
        
                return response
            
            else:
                return Response(status=status.HTTP_404_NOT_FOUND,data = "Table not found")

        else:
            return Response(status=status.HTTP_404_NOT_FOUND,data = "Segment Not Found")

    except Exception as e:
        print("Error in downloading final report : ",e)
        return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR,data = "Error in downloading final report")


@api_view(['GET'])
def get_numberOfLookupValidationFailed_Count(request):
    
    conn = Connection.objects.get(id=7)

    print(conn.connection_name)

    # Update fields as needed
    new_tables = ['table5', 'table6', 'table7']
    existing_tables = conn.imported_tables
    print(existing_tables)
    existing_tables.append('table4')
    conn.imported_tables = existing_tables

    # Save changes
    conn.save()

    return Response("Done")

    # return Response(get_numberOfLookupValidationFailed_Count1("t_68_Product_Additional_Descriptions_validation_fail"))
    # return Response(deterimine_LookTableButton_status(1967))
    # return Response(get_complete_table_data("t_68_Product_Additional_Descriptions_validation_fail"))


def get_numberOfLookupValidationFailed_Count1(table_name):
    try:
        # Check if the table exists
        if not check_table_existance(table_name):
            return "Validation table not found"
        with connection.cursor() as cursor:
            # Lookup error counts
            cursor.execute(f"""
                SELECT Lookup_ErrorTable_Field, COUNT(*) as Lookup_Failed 
                FROM {table_name} 
                GROUP BY Lookup_ErrorTable_Field
            """)
            lookup_data = cursor.fetchall()

            # Mandatory error counts
            cursor.execute(f"""
                SELECT Mandatory_ErrorField, COUNT(*) as Mandatory_Failed 
                FROM {table_name}
                GROUP BY Mandatory_ErrorField
            """)
            mandatory_data = cursor.fetchall()

        # Convert to list of dicts for better API response
        lookup_errors = [
            { row[0] :  row[1]}
            for row in lookup_data if row[0] is not None and row[0]!=""
        ]
        mandatory_errors = [
            { row[0] :  row[1]}
            for row in mandatory_data if row[0] is not None and row[0]!=""
        ]

        final_error_report = {
            "lookup_errors": lookup_errors,
            "mandatory_errors": mandatory_errors
        }
        return final_error_report
    except Exception as e:
        error_msg = {"error": f"An error occurred in finding Number of Failed Count: {str(e)}"}
        print(error_msg)
        raise


def get_complete_table_data(table_name):

    try:
        with connection.cursor() as cursor:
            print(table_name)
            cursor.execute(f'SELECT * FROM "{table_name}"')
            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()
            # Use list comprehension for cleaner code
            return [dict(zip(columns, row)) for row in rows]

    except Exception as e:
        error_msg = {"error": f"An error occurred in returnung table data: {str(e)}"}
        print(error_msg)
        raise Exception(f"Error while Fetching Table Data: {str(e)}")


@api_view(['GET'])
def get_report_table(request,segment_id,table_type):

    try:

        if segments.objects.filter(segment_id = segment_id).exists():
            
            seg = segments.objects.get(segment_id = segment_id)
            table_name = seg.table_name
            if table_type == "Full":
                table_name = table_name + "_validation"
            elif table_type == "Valid":
                table_name = table_name + "_validation_pass"
            elif table_type == "Invalid":
                table_name = table_name + "_validation_fail"
            else:
                return Response(status=status.HTTP_400_BAD_REQUEST,data = "Invalid Request")
            if check_table_existance(table_name):
                return Response(status=status.HTTP_200_OK,data = get_complete_table_data(table_name))
            else:
                return Response(status=status.HTTP_404_NOT_FOUND,data = "Table not found")

        else:
            return Response(status=status.HTTP_404_NOT_FOUND,data = "Segment not found")
        
        
    except Exception as e:
        print("Error in getting report table : ",e)
        return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR,data = "Error in getting report table")

# @api_view(['GET'])
# def get_validation_table(request,seg_id):

#     try:
#         if segments.objects.filter(segment_id = seg_id).exists():
#             seg = segments.objects.get(segment_id = seg_id)
#             table_name = seg.table_name + "_validation"
#             if check_table_existance(table_name):
#                 return Response(status=status.HTTP_200_OK,data = get_complete_table_data(table_name))
#             else:
#                 return Response(status=status.HTTP_404_NOT_FOUND,data = "Validation Table not found")
#         else:
#             return Response(status=status.HTTP_404_NOT_FOUND,data = "Segment not found")
#     except Exception as e:
#         print("Error in getting validation table : ",e)
#         return Response(status=status.HTTP_500_INTERNAL_SERVER_ERROR,data = "Error in getting validation table")


def commit():
    return "Demo Updated 06-06-2025"



import pandas as pd

def dynamic_side_by_side_select(table_a, table_b, join_columns):
    import sqlite3
    cur = connection.cursor()
    cur.execute(f"PRAGMA table_info({table_a})")
    columns_info = cur.fetchall()
    columns = [col[1] for col in columns_info]

    select_list = []
    col_names = []
    for col in columns:
        select_list.append(f"src.{col}")
        col_names.append(f"src_{col}")
        select_list.append(f"trans.{col}")
        col_names.append(f"trans_{col}")
    select_clause = ", ".join(select_list)

    join_conditions = " AND ".join([f"src.{col} = trans.{col}" for col in join_columns])

    query = f"""
    SELECT {select_clause}
    FROM {table_a} src
    INNER JOIN {table_b} trans
    ON {join_conditions}
    """

    cur.execute(query)
    rows = cur.fetchall()

    df = pd.DataFrame(rows, columns=col_names)
    return df,list(df.columns)

import json

@api_view(['GET'])
def get_preLoad_table(request, pid, oid, sid):
    segmentForTable = segments.objects.filter(segment_id=sid, project_id=pid, obj_id=oid).first()
    target_table_name = segmentForTable.table_name
    fields_list = fields.objects.filter(
        project_id=pid,
        obj_id=oid,
        segement_id=sid,
        isKey=True
    ).values_list('fields', flat=True)
    fields_list = list(fields_list)
    df, columns = dynamic_side_by_side_select(target_table_name + '_src', target_table_name, fields_list)
    json_columns = []
    # Clean DataFrame
    df = df.replace(r'^\s*$', np.nan, regex=True)
    df = df.dropna(axis=1, how='all')
    columns= list(df.columns)
    for i in range(0,len(columns)):
        json_columns.append({'key':i,'value':columns[i]})
    df=pd.DataFrame(df)  
    json_data = df.to_dict(orient='records')
    return_json = {}
    return_json["rows"] = json_columns
    return_json["columns"] = json_data
    # print(columns)
    df = pd.DataFrame(df)
    
    # Column names as JSON-serializable structure
    columns = list(df.columns)
    json_columns = [{'key': i, 'value': col} for i, col in enumerate(columns)]

    # Data rows, WITHOUT to_dict
    json_data = json.loads(df.to_json(orient='records'))  # use json.loads to get a Python list, not a string

    return_json = {
        "columns" : json_data,
        "rows": json_columns
    }
    return JsonResponse(return_json, safe=False)



from fuzzywuzzy import fuzz

# def validate_and_copy_errors(data_table, error_table, pk_fields):
#     conn = connection.connection  # Use raw DBAPI sqlite3 connection
#     cur = connection.cursor()     # Use Django cursor only for executing raw SQL
#     cur.execute(f"DROP TABLE IF EXISTS {error_table}")
#     connection.commit()
#     # Ensure error_table exists with reason column (same as before)
#     cur.execute(f"PRAGMA table_info({data_table})")
#     cols_info = cur.fetchall()
#     columns = [col[1] for col in cols_info]
#     cols_def = []
#     for col in cols_info:
#         col_name = col[1]
#         col_type = col[2] or "TEXT"
#         cols_def.append(f"{col_name} {col_type}")
#     cols_def.append("reason TEXT")
#     create_error_sql = f"""
#     CREATE TABLE IF NOT EXISTS {error_table} (
#         {', '.join(cols_def)}
#     );
#     """
#     cur.execute(create_error_sql)
#     connection.commit()  # Commit via Django connection

#     # Load data from data_table into pandas using raw sqlite3 connection for compatibility
#     df = pd.read_sql_query(f"SELECT * FROM {data_table}", conn)

#     error_df = pd.DataFrame(columns=list(df.columns) + ['reason'])

#     def build_reason(issue, row):
#         keys = ", ".join([f"{k}={row[k]}" for k in pk_fields])
#         return f"{issue} ({keys})"

#     # Your validations (unchanged)
#     null_pk_mask = df[pk_fields].isnull().any(axis=1)
#     null_rows = df[null_pk_mask].copy()
#     null_rows['reason'] = null_rows.apply(lambda row: build_reason("Null in PK", row), axis=1)
#     error_df = pd.concat([error_df, null_rows])

#     not_null_mask = ~df[pk_fields].isnull().any(axis=1)
#     dup_pk_mask = df[not_null_mask].duplicated(subset=pk_fields, keep=False)
#     dup_pk_rows = df[not_null_mask][dup_pk_mask].copy()
#     dup_pk_rows['reason'] = dup_pk_rows.apply(lambda row: build_reason("Duplicate PK", row), axis=1)
#     error_df = pd.concat([error_df, dup_pk_rows])

#     full_dup_mask = df.duplicated(keep=False)
#     full_dup_rows = df[full_dup_mask].copy()
#     full_dup_rows['reason'] = full_dup_rows.apply(lambda row: build_reason("Full row duplicate", row), axis=1)
#     error_df = pd.concat([error_df, full_dup_rows])

#     error_df = error_df.drop_duplicates()

#     if not error_df.empty:
#         # Write errors to error_table passing raw connection
#         error_df.to_sql(error_table, conn, if_exists='append', index=False)

#     print(f"Validated {data_table} and copied {len(error_df)} error rows to {error_table}.")


from fuzzywuzzy import fuzz

from fuzzywuzzy import fuzz
import pandas as pd
from django.db import connection

def validate_and_copy_errors(data_table, error_table, pk_fields):
    conn = connection.connection  # raw DBAPI sqlite3 connection
    cur = connection.cursor()
    if table_exists(error_table) :
        return
    cur.execute(f"PRAGMA table_info({data_table})")
    cols_info = cur.fetchall()
    columns = [col[1] for col in cols_info]
    cols_def = []
    for col in cols_info:
        col_name = col[1]
        col_type = col[2] or "TEXT"
        cols_def.append(f"{col_name} {col_type}")
    cols_def.append("reason TEXT")
    create_error_sql = f"""
    CREATE TABLE {error_table} (
        {', '.join(cols_def)}
    );
    """
    cur.execute(create_error_sql)
    connection.commit()

    # Load table data
    df = pd.read_sql_query(f"SELECT * FROM {data_table}", conn)
    error_df = pd.DataFrame(columns=list(df.columns) + ['reason'])

    def build_reason(issue, row):
        keys = ", ".join([f"{k}={row[k]}" for k in pk_fields])
        return f"{issue} ({keys})"

    # 1. Null in PK
    null_pk_mask = df[pk_fields].isnull().any(axis=1)
    null_rows = df[null_pk_mask].copy()
    null_rows['reason'] = null_rows.apply(lambda row: build_reason("Null in PK", row), axis=1)
    error_df = pd.concat([error_df, null_rows])

    # 2. Duplicate PK excluding null PKs
    not_null_mask = ~df[pk_fields].isnull().any(axis=1)
    dup_pk_mask = df[not_null_mask].duplicated(subset=pk_fields, keep=False)
    dup_pk_rows = df[not_null_mask][dup_pk_mask].copy()
    dup_pk_rows['reason'] = dup_pk_rows.apply(lambda row: build_reason("Duplicate PK", row), axis=1)
    error_df = pd.concat([error_df, dup_pk_rows])

    # 3. Full row duplicates
    full_dup_mask = df.duplicated(keep=False)
    full_dup_rows = df[full_dup_mask].copy()
    full_dup_rows['reason'] = full_dup_rows.apply(lambda row: build_reason("Full row duplicate", row), axis=1)
    error_df = pd.concat([error_df, full_dup_rows])

    # # 4. Optimized Fuzzy match block
    # df_str = df.astype(str)
    # n = len(df_str)
    # fuzzy_flagged = set()
    # fuzzy_errors = []
    # cc=0
    # for i in range(n):
    #     if i in fuzzy_flagged:
    #         continue
    #     row_i_str = " ".join(df_str.iloc[i].values)
    #     fflg=0
    #     for j in range(i + 1, n):
    #         if j in fuzzy_flagged:
    #             continue
    #         row_j_str = " ".join(df_str.iloc[j].values)
    #         ratio = fuzz.ratio(row_i_str, row_j_str)
    #         if ratio > 80:
    #             fflg=1
    #             cc+=1
    #             print(cc)
    #             # mark both as flagged and append once each
    #             row_i = df.iloc[i].copy()
    #             row_j = df.iloc[j].copy()
    #             reason_i = f"Fuzzy"
    #             reason_j = f"Fuzzy"
    #             row_i['reason'] = reason_i
    #             row_j['reason'] = reason_j
    #             fuzzy_errors.append(row_j)
    #             fuzzy_flagged.add(j)
    #             # break  # stop comparing row_i further once flagged
    #         if fflg == 1:
    #             fuzzy_flagged.add(i)
    #             fuzzy_errors.append(row_i)

    # if fuzzy_errors:
    #     fuzzy_df = pd.DataFrame(fuzzy_errors)
    #     error_df = pd.concat([error_df, fuzzy_df])

    error_df = error_df.drop_duplicates()

    if not error_df.empty:
        error_df.to_sql(error_table, conn, if_exists='append', index=False)

    print(f"Validated {data_table} and copied {len(error_df)} error rows to {error_table}.")



# def table_to_custom_json(table_name):
#     # Get column names
#     with connection.cursor() as cursor:
#         cursor.execute(f"PRAGMA table_info({table_name})")
#         columns_info = cursor.fetchall()
#     columns = [col[1] for col in columns_info]

#     # Fetch all rows into pandas DataFrame
#     query = f"SELECT * FROM {table_name}"
#     df = pd.read_sql_query(query, connection)

#     # Build json_columns as list of {key, value}
#     json_columns = [{'key': i, 'value': col} for i, col in enumerate(columns)]

#     # Convert DataFrame rows into dict list
#     json_data = df.to_dict(orient='records')

#     # Prepare return JSON structure
#     # json_columns.insert(0,json_columns[-1])
#     # json_columns.pop(-1)
#     return_json = {
#         "rows": json_columns,
#         "columns": json_data
#     }
#     print(columns)  # for debug/logging
#     return JsonResponse(return_json, safe=False)


@api_view(['GET'])
def get_error_table(request,pid,oid,sid):
    segmentForTable = segments.objects.filter(segment_id=sid,project_id=pid,obj_id=oid).first()
    target_table_name = segmentForTable.table_name
    fields_list = fields.objects.filter(
    project_id=pid,
    obj_id=oid,
    segement_id=sid,
    isKey=True
        ).values_list('fields', flat=True)
    #primary keys for this target table
    fields_list = list(fields_list)
    print(fields_list)
    validate_and_copy_errors(target_table_name+"_src",target_table_name+"_err",fields_list)
    return table_to_custom_json(target_table_name+"_err")



@api_view(['POST', 'GET'])
def file_upload_bussiness(request):
    excel_file = request.FILES['file'] 
    df = pd.read_excel(excel_file)
    df = df.fillna('')

    # Make isKey and isMandatory columns string type
    if 'isKey' in df.columns:
        df['isKey'] = df['isKey'].astype(str)
    if 'isMandatory' in df.columns:
        df['isMandatory'] = df['isMandatory'].astype(str)
    data = df.to_dict(orient='records')
    return JsonResponse(data, safe=False)


def table_has_data(table_name):
    with connection.cursor() as cursor:
        cursor.execute(f"SELECT EXISTS(SELECT 1 FROM {table_name} LIMIT 1);")
        result = cursor.fetchone()
        return bool(result[0]) if result else False

@api_view(['GET'])
def check_is_initial_version(request,pid,oid,sid):
    latest_version = Rule.objects.filter(
                    project_id=pid,  
                    object_id=oid,
                    segment_id=sid
                ).order_by('-version_id').first()
    try:
        latest_version.version_id
        noVersions=False
    except:
        noVersions=True
    
    # if not noVersions: return Response(3)
    segmentForTable = segments.objects.filter(segment_id=sid,project_id=pid,obj_id=oid).first()
    target_table_name = segmentForTable.table_name
    table_contains_data = table_has_data(target_table_name+"_src")
    
    if (table_contains_data and noVersions) or (table_contains_data and noVersions==False) : 
        # already uploaded DMC with DATA
        # we should not open initial rules
        return JsonResponse({"value" : "Success"}, status=status.HTTP_200_OK,safe=False)
    return  JsonResponse({"value" : "Error"} ,status=status.HTTP_404_NOT_FOUND,safe=False)
    


@api_view(['GET'])
def get_validation_table(request,pid,oid,sid):
    segmentForTable = segments.objects.filter(segment_id=sid,project_id=pid,obj_id=oid).first()
    target_table_name = segmentForTable.table_name
    if(table_exists(target_table_name+"_val")):
        return table_to_custom_json(target_table_name+"_val")
    return JsonResponse({
        "rows": [],
        "columns": []
        }
    ,safe=False)






        

    













